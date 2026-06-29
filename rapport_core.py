#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
  PROGRAMME DE TRAITEMENT DES RAPPORTS DE POINTAGE
  Générateur de rapport enrichi (heures sup / respect horaire)
=============================================================================
  Entrée  : Fichier Excel (.xlsx) de présence
  Sortie  : PDF enrichi (portrait) avec :
            - Rapports individuels par employé
            - Rapport de présence global
            - Classement retards & absences
            - Graphique d'assiduité
=============================================================================
  Usage : python3 rapport_heures.py [chemin_du_xlsx]
=============================================================================
"""

import sys, os, re, math, json
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, black
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    KeepTogether, CondPageBreak
)
from reportlab.platypus.flowables import Flowable

class SmartPageBreak(Flowable):
    """PageBreak conditionnel : ne saute QUE si la frame courante a du contenu déjà rempli.
    Évite les pages vierges. v56 : adaptation auto selon orientation."""
    def __init__(self):
        Flowable.__init__(self)
        self.width = 0
        self.height = 0
    
    def wrap(self, availWidth, availHeight):
        # Détecte orientation : portrait si largeur < hauteur, sinon paysage
        # Portrait A4 ≈ 595x842 → frame ≈ 760pt → seuil 700
        # Paysage A4 ≈ 842x595 → frame ≈ 549pt → seuil 500
        threshold = 500 if availWidth > 700 else 700
        if availHeight >= threshold:
            return (0, 0)  # En haut de page → pas de break
        return (availWidth, availHeight)
    
    def draw(self):
        pass
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing, Wedge, String, Circle, Rect
from reportlab.graphics import renderPDF
from datetime import datetime
from collections import defaultdict, OrderedDict

# ======================== COULEURS ========================
TEAL = HexColor("#1A7A6D")
DARK_TEAL = HexColor("#0D6B5E")
ORANGE = HexColor("#E8672A")
GREEN = HexColor("#008000")
RED = HexColor("#CC0000")
BLUE = HexColor("#0000CC")
LGRAY = HexColor("#F5F5F5")
MGRAY = HexColor("#DDDDDD")

# ======================== UTILITAIRES ========================

def t2m(t):
    """Convertit HH:MM en minutes."""
    if not t or str(t).strip() in ['-','nan','','None']: return 0
    s = str(t).strip().replace('\n','')
    p = s.split(':')
    try: return int(p[0])*60+int(p[1]) if len(p)==2 else 0
    except: return 0

def m2h(m):
    """Convertit minutes en HH:MM."""
    if m <= 0: return "00:00"
    return f"{int(m)//60:02d}:{int(m)%60:02d}"

def safe(s):
    """Échappe pour XML ReportLab."""
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

# ======================== EXTRACTION EXCEL ========================

def extract_from_excel(xlsx_path):
    """Extrait les données depuis le fichier Excel."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    # Détection du nom du client
    title = str(ws.cell(1,1).value or "")
    client_name = ""
    for r in range(3, min(6, ws.max_row+1)):
        svc = str(ws.cell(r, 4).value or "")
        if '>' in svc:
            client_name = svc.split('>')[-1].strip()
            break
    if not client_name:
        client_name = "ENTREPRISE"
    
    # Extraction des données par employé
    employees = OrderedDict()
    
    for r in range(3, ws.max_row + 1):
        prenom = str(ws.cell(r, 1).value or "").strip()
        nom = str(ws.cell(r, 2).value or "").strip()
        eid = str(ws.cell(r, 3).value or "").strip()
        
        if not prenom and not nom:
            continue
        
        full_name = f"{prenom} {nom}".strip()
        date_val = str(ws.cell(r, 5).value or "").strip()
        sched_start = str(ws.cell(r, 6).value or "").strip()
        sched_end = str(ws.cell(r, 7).value or "").strip()
        actual_arr = str(ws.cell(r, 8).value or "").strip()
        actual_dep = str(ws.cell(r, 9).value or "").strip()
        duration = str(ws.cell(r, 10).value or "").strip()
        
        # Nettoyer les dates datetime
        if 'datetime' in str(type(ws.cell(r, 5).value)):
            date_val = ws.cell(r, 5).value.strftime('%Y-%m-%d')
        
        # Nettoyer les heures datetime  
        for col_idx, col_name in [(6,'sched_start'),(7,'sched_end'),(8,'actual_arr'),(9,'actual_dep'),(10,'duration')]:
            val = ws.cell(r, col_idx).value
            if val and 'datetime' in str(type(val)):
                locals()[col_name] = val.strftime('%H:%M')
            elif val and 'time' in str(type(val)):
                locals()[col_name] = val.strftime('%H:%M')
        
        # Re-read after potential conversion
        sched_start = str(ws.cell(r, 6).value or "").strip()
        sched_end = str(ws.cell(r, 7).value or "").strip()
        actual_arr = str(ws.cell(r, 8).value or "").strip()
        actual_dep = str(ws.cell(r, 9).value or "").strip()
        duration = str(ws.cell(r, 10).value or "").strip()
        
        # Gérer les formats datetime
        for field in [sched_start, sched_end, actual_arr, actual_dep, duration]:
            pass  # Already strings
        
        key = (full_name, eid)
        if key not in employees:
            employees[key] = {
                'name': full_name,
                'ref': eid,
                'records': []
            }
        
        employees[key]['records'].append({
            'date': date_val[:10] if len(date_val) >= 10 else date_val,
            'sched_start': sched_start[:5] if len(sched_start) >= 5 else sched_start,
            'sched_end': sched_end[:5] if len(sched_end) >= 5 else sched_end,
            'arrival': actual_arr[:5] if len(actual_arr) >= 5 else actual_arr,
            'departure': actual_dep[:5] if len(actual_dep) >= 5 else actual_dep,
            'duration': duration[:5] if len(duration) >= 5 else duration,
        })
    
    # Trier les records par date pour chaque employé
    for key in employees:
        employees[key]['records'].sort(key=lambda x: x['date'])
    
    return list(employees.values()), client_name

# ======================== CALCULS ========================

def calc_employee_stats(emp, hp=0, hp_weekend=0, hourly_cost=0, rest_days=None,
                         days_required_override=None, period_total_days=None,
                         pause_minutes=0, auto_invert_night=False):
    """Calcule les statistiques complètes d'un employé. 
    rest_days=liste des jours de repos (0=lundi..6=dimanche).
    days_required_override : si fourni, force le nombre de jours obligatoires (sinon calcul auto).
    period_total_days : si fourni, utilise ce total commun à TOUS les employés au lieu
    de len(records) qui varie par employé (cas où l'Excel a des dates manquantes pour
    certains employés).
    v126 : auto_invert_night : si True, inverse automatiquement l'EDT de jour en EDT
    de nuit quand les pointages indiquent clairement un poste de nuit
    (arrivée >= 18:00 ET départ < 12:00). Désactivé par défaut."""
    if rest_days is None: rest_days = []
    records = emp['records']
    total_required = 0
    total_worked = 0
    total_overtime = 0
    total_deficit = 0
    total_late_mins = 0
    days_present = 0
    days_late = 0
    days_punctual = 0
    days_absent = 0
    days_badge_error = 0
    days_rest = 0
    hm = hp * 60
    hm_we = hp_weekend * 60
    
    enriched = []
    
    for rec in records:
        # v123 : Logique corrigée — la PRIORITÉ est claire :
        #   1. sched_start_matched : EDT saisi EXPLICITEMENT par l'utilisateur dans
        #      l'interface (override) → si présent, l'utilisateur a explicitement demandé
        #      à modifier les heures, on utilise sa saisie.
        #   2. sched_start_original : valeur du fichier source PRÉSERVÉE
        #      (pour le cas où sched_start aurait été modifié ailleurs).
        #   3. sched_start : valeur courante (= fichier source si pas écrasé).
        #
        # Ainsi : sans saisie explicite, les heures du fichier source ne bougent JAMAIS.
        # Avec saisie explicite, l'override de l'utilisateur s'applique.
        eff_start = (rec.get('sched_start_matched')
                     or rec.get('sched_start_original')
                     or rec.get('sched_start', ''))
        eff_end = (rec.get('sched_end_matched')
                   or rec.get('sched_end_original')
                   or rec.get('sched_end', ''))
        ss = t2m(eff_start)
        se = t2m(eff_end)
        aa = t2m(rec['arrival'])
        ad = t2m(rec['departure'])
        dur = t2m(rec['duration'])
        
        # v126 : INVERSION AUTOMATIQUE EDT pour poste de nuit détecté (OPTIONNEL)
        # Cette inversion s'active SEULEMENT si `auto_invert_night=True` (bouton coché).
        # Conditions strictes pour éviter les faux positifs :
        #   - arrivée >= 18:00 (1080 min) : vraiment en soirée
        #   - départ > 0 ET départ < 12:00 (720 min) : vraiment le matin du lendemain
        #   - planning saisi est de jour (ss < se, ex: 07:00-20:00)
        # Si les 3 conditions sont remplies, inverser : 07:00-20:00 → 20:00-07:00.
        night_swap_applied = False
        if auto_invert_night:
            is_night_shift_detected = (aa >= 1080 and ad > 0 and ad < 720)
            is_planning_already_night = (ss > 0 and se > 0 and ss > se)
            if is_night_shift_detected and not is_planning_already_night and ss > 0 and se > 0:
                # Inversion : ce qui était début devient fin, et vice-versa
                ss, se = se, ss
                eff_start, eff_end = eff_end, eff_start
                night_swap_applied = True
        
        # Déterminer le jour de la semaine
        is_weekend = False
        is_rest_day = False
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(rec['date'][:10], '%Y-%m-%d')
            is_weekend = d.weekday() >= 5
            if d.weekday() in rest_days:
                is_rest_day = True
        except:
            pass
        
        # Sélectionner les heures obligatoires selon le jour
        if is_rest_day:
            required = 0  # Jour de repos — pas d'heures obligatoires
        elif is_weekend and hp_weekend > 0:
            required = hm_we
        elif not is_weekend and hp > 0:
            required = hm
        elif hp > 0 and hp_weekend == 0:
            required = hm
        else:
            # v124 : Détection planning de nuit (ss > se → traverse minuit)
            if ss > 0 and se > 0 and ss > se:
                # Planning de nuit : durée = (24h - ss) + se
                required = (1440 - ss) + se
            else:
                required = se - ss if se > ss else 0
            # v72 : appliquer la pause globale si demandée
            if pause_minutes > 0 and required > pause_minutes:
                required -= pause_minutes
        
        if not is_rest_day:
            total_required += required
        
        # v124 + v125 : Affichage = priorité à la saisie EXPLICITE de l'utilisateur (matched),
        # sinon EDT d'origine du fichier. Si swap nuit détecté → inverser l'affichage aussi.
        display_start = (rec.get('sched_start_matched')
                         or rec.get('sched_start_original')
                         or rec.get('sched_start', ''))
        display_end = (rec.get('sched_end_matched')
                       or rec.get('sched_end_original')
                       or rec.get('sched_end', ''))
        if night_swap_applied:
            display_start, display_end = display_end, display_start
        schedule_str = f"({display_start}_{display_end})"
        
        # Déterminer l'état
        if is_rest_day:
            state = "Repos"
            days_rest += 1
            worked = t2m(rec['duration']) if t2m(rec['duration']) > 0 else 0
            if worked > 0:
                total_worked += worked
                days_present += 1
                total_overtime += worked  # Tout travail en jour de repos = heures sup
            overtime = worked
            late = 0
            respect = "REPOS"
        elif (aa > 0 and ad == 0) or (aa == 0 and ad > 0):
            # NOUVEAU v51 : Badge incomplet - 1 seul pointage (arrivée OU départ, pas les deux)
            state = "Erreur badge"
            days_badge_error += 1
            worked = 0
            overtime = 0
            late = 0
            respect = "ERR"
        elif (aa == 0 and ad == 0):
            # v124 : Aucun pointage du tout = Absent
            # (test simplifié : on ne se base plus sur dur==0 car la duration peut être
            # à 00:00 pour un poste de nuit valide. On vérifie uniquement les badges.)
            state = "Absent(e)"
            days_absent += 1
            worked = 0
            overtime = 0
            late = 0
            respect = "ABS"
        else:
            days_present += 1
            
            # v124 : DÉTECTION POSTE DE NUIT
            # Si arrivée tard (>= 14h00) ET départ tôt (< 14h00 et > 0) → c'est un poste de nuit.
            # L'employé est arrivé en soirée et reparti le lendemain matin.
            # Ex: arrivée 19:37, départ 08:07 → travaillé = (24h - 19:37) + 08:07 = 12h30
            is_night_shift = (aa >= 840 and ad > 0 and ad < 840)
            
            # Aussi : si le PLANNING saisi est un poste de nuit (ex: 20:00-07:00)
            # → ss > se (1200 > 420), on considère aussi comme nuit
            is_planning_night = (ss > 0 and se > 0 and ss > se)
            
            if is_night_shift:
                # === HEURES TRAVAILLÉES (poste de nuit) ===
                # de l'arrivée le soir jusqu'à minuit + de minuit jusqu'au départ matin
                worked = (1440 - aa) + ad
                
                # Retard : si planning est nuit aussi, comparer aa avec ss (planning début)
                if is_planning_night:
                    # ss = début prévu nuit (ex: 20:00 = 1200), aa = arrivée réelle (ex: 19:37 = 1177)
                    # Si arrivée APRÈS début prévu = retard
                    late = aa - ss if aa > ss else 0
                else:
                    # Planning non détecté comme nuit, on ne peut pas calculer le retard
                    late = 0
                
                if late > 0:
                    total_late_mins += late
                    days_late += 1
                    state = "Retard"
                else:
                    days_punctual += 1
                    state = "Présent(e)"
                
                # === HEURES SUPPLÉMENTAIRES (poste de nuit) ===
                if is_planning_night:
                    # Durée prévue = de ss à se (en passant par minuit)
                    scheduled_duration = (1440 - ss) + se
                    if worked > scheduled_duration:
                        overtime = worked - scheduled_duration
                    else:
                        overtime = 0
                else:
                    overtime = 0
                total_overtime += overtime
            else:
                # === COMPORTEMENT NORMAL (jour) ===
                # Le comptage commence au début du planning, PAS avant
                effective_start = max(aa, ss)
                worked = ad - effective_start if ad > effective_start else 0
                total_worked = total_worked  # update plus bas
                
                # Retard : arrivée après le début prévu
                if aa > ss:
                    late = aa - ss
                    total_late_mins += late
                    days_late += 1
                    state = "Retard"
                else:
                    late = 0
                    days_punctual += 1
                    state = "Présent(e)"
                
                # === HEURES SUPPLÉMENTAIRES ===
                if ad > se:
                    overtime = ad - se
                else:
                    overtime = 0
                total_overtime += overtime
            
            total_worked += worked
            
            # === RESPECT HORAIRE ===
            if worked >= (required - 5):
                respect = "OUI"
            else:
                deficit = required - worked
                total_deficit += deficit
                respect = f"NON (-{m2h(deficit)})"
        
        enriched.append({
            'date': rec['date'],
            'schedule': schedule_str,
            'state': state,
            'arrival': rec['arrival'],
            'departure': rec['departure'],
            # v67 : pause début/fin (peut être absent pour les anciens flux DPCI/Excel)
            'pause_start': rec.get('pause_start', '') or '',
            'pause_end': rec.get('pause_end', '') or '',
            'worked': m2h(worked),
            'late': m2h(late),
            'required': m2h(required),
            'respect': respect,
            'overtime': m2h(overtime),
        })
    
    # === ASSIDUITÉ (basée sur le taux de présence) ===
    # Détermination des jours obligatoires : override prioritaire, sinon calcul auto basé sur la période
    if days_required_override is not None and days_required_override > 0:
        days_required = int(days_required_override)
        days_required_source = 'manuel'
    else:
        # NOUVEAU v49 : utiliser le total de la période (commun à tous) au lieu de len(records)
        # qui varie par employé. Cela garantit que TOUS les employés du même mois ont
        # le même nombre de jours obligatoires.
        if period_total_days is not None and period_total_days > 0:
            days_required = period_total_days - days_rest
        else:
            # Fallback ancien comportement (si period_total_days non fourni)
            days_required = len(records) - days_rest
        days_required_source = 'auto'
    
    # Recalcul absences en fonction de days_required (si override > calcul auto, on a plus d'absences)
    # Si override < calcul auto (ex: 22 j obligatoires sur 30 jours, le delta = jours non comptés)
    # Pour ne pas changer la logique de calcul historique, on garde days_absent tel quel
    # (calculé jour par jour depuis records). days_required sert juste de référence pour le taux.
    
    # NOUVEAU v53 : cap à 100% (cas où days_present > days_required)
    presence_rate = min((days_present / days_required * 100) if days_required > 0 else 0, 100.0)
    if presence_rate >= 95:
        observation = "Assidu"
    elif presence_rate >= 80:
        observation = "Moyennement assidu"
    else:
        observation = "Non assidu"
    
    # NOUVEAU v52 : days_absent_real = obligatoires - effectués
    # Les "Absent(e)" détectés via records (ligne dans le fichier sans pointage) sont conservés
    # dans days_absent (compatibilité), mais le calcul réel d'absences manquantes est :
    # absences réelles = nb jours obligatoires - nb jours où l'employé a été VU dans le fichier
    days_effectues_real = days_present + days_badge_error  # days_present inclut days_late
    days_absent_real = max(days_required - days_effectues_real, 0)
    
    # Recalcul du taux de présence avec les absences réelles
    # NOUVEAU v53 : capper à 100% (cas où days_present > days_required, ex: heures sup ou
    # jours travaillés au-delà de l'obligation)
    presence_rate_real = min((days_present / days_required * 100) if days_required > 0 else 0, 100.0)
    
    stats = {
        'days_required': days_required,
        'days_required_source': days_required_source,
        'days_present': days_present,
        'days_late': days_late,
        'days_punctual': days_punctual,
        'days_absent': days_absent_real,  # v52 : on remplace par les absences réelles
        'days_absent_records': days_absent,  # ancien comptage (pour debug si besoin)
        'days_effectues': days_effectues_real,  # v52 : nouveau champ explicite
        'days_badge_error': days_badge_error,
        'days_rest': days_rest,
        'total_required': total_required,
        'total_worked': total_worked,
        'total_overtime': total_overtime,
        'total_deficit': total_deficit,
        'total_late_mins': total_late_mins,
        'presence_rate': round(presence_rate_real, 1),
        'observation': observation,
        'hourly_cost': hourly_cost,
        'cost_late': round(total_late_mins / 60 * hourly_cost) if hourly_cost > 0 else 0,
        'cost_deficit': round(total_deficit / 60 * hourly_cost) if hourly_cost > 0 else 0,
        'cost_absent': round(days_absent_real * (total_required / max(len(records), 1)) / 60 * hourly_cost) if hourly_cost > 0 and len(records) > 0 else 0,
        'cost_overtime': round(total_overtime / 60 * hourly_cost) if hourly_cost > 0 else 0,
    }
    
    return enriched, stats

# ======================== STYLES PDF ========================

def make_styles():
    return {
        'co': ParagraphStyle('co', fontName='Helvetica-Bold', fontSize=10, textColor=TEAL, leading=12),
        'cl': ParagraphStyle('cl', fontName='Helvetica-Bold', fontSize=12, textColor=ORANGE, alignment=TA_RIGHT),
        'ti': ParagraphStyle('ti', fontName='Helvetica-Bold', fontSize=14, textColor=TEAL, alignment=TA_CENTER, spaceAfter=4),
        'st': ParagraphStyle('st', fontName='Helvetica', fontSize=9, alignment=TA_CENTER, spaceAfter=8),
        'ei': ParagraphStyle('ei', fontName='Helvetica-Bold', fontSize=9, spaceAfter=2),
        'eb': ParagraphStyle('eb', fontName='Helvetica-Bold', fontSize=9, textColor=BLUE, spaceAfter=2),
        'c': ParagraphStyle('c', fontName='Helvetica', fontSize=7, alignment=TA_CENTER, leading=8),
        'cb': ParagraphStyle('cb', fontName='Helvetica-Bold', fontSize=7, alignment=TA_CENTER, leading=8),
        'h': ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=7, textColor=white, alignment=TA_CENTER, leading=8),
        'g': ParagraphStyle('g', fontName='Helvetica-Bold', fontSize=7, textColor=GREEN, alignment=TA_CENTER, leading=8),
        'r': ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=7, textColor=RED, alignment=TA_CENTER, leading=8),
        'b': ParagraphStyle('b', fontName='Helvetica-Bold', fontSize=7, textColor=BLUE, alignment=TA_CENTER, leading=8),
        'sh': ParagraphStyle('sh', fontName='Helvetica-Bold', fontSize=6.5, textColor=white, alignment=TA_CENTER, leading=8),
        'sv': ParagraphStyle('sv', fontName='Helvetica', fontSize=7, alignment=TA_CENTER, leading=8),
        'ft': ParagraphStyle('ft', fontName='Helvetica', fontSize=6, alignment=TA_RIGHT, textColor=HexColor("#888")),
        # Styles pour les pages résumé
        'big_ti': ParagraphStyle('big_ti', fontName='Helvetica-Bold', fontSize=14, textColor=TEAL, alignment=TA_CENTER, spaceAfter=8),
        'med_ti': ParagraphStyle('med_ti', fontName='Helvetica-Bold', fontSize=11, textColor=TEAL, alignment=TA_LEFT, spaceAfter=4),
        'rh': ParagraphStyle('rh', fontName='Helvetica-Bold', fontSize=7, textColor=white, alignment=TA_CENTER, leading=9),
        'rv': ParagraphStyle('rv', fontName='Helvetica', fontSize=7, alignment=TA_CENTER, leading=9),
        'rvb': ParagraphStyle('rvb', fontName='Helvetica-Bold', fontSize=7, alignment=TA_CENTER, leading=9),
        'rvo': ParagraphStyle('rvo', fontName='Helvetica-Bold', fontSize=7, textColor=ORANGE, alignment=TA_CENTER, leading=9),
    }

# ======================== HEADER COMMUN ========================

def make_header(S, provider_name, provider_info, client_name, client_info=""):
    """Crée l'en-tête commun pour chaque page."""
    right_text = safe(client_name)
    if client_info:
        right_text += f"<br/><font size=6>{safe(client_info)}</font>"
    h = Table([
        [Paragraph(f"{safe(provider_name)}<br/><font size=6>{safe(provider_info)}</font>", S['co']),
         Paragraph(right_text, S['cl'])]
    ], colWidths=[110*mm, 80*mm])
    h.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                           ('LINEBELOW',(0,0),(-1,0),1,TEAL)]))
    return h

# ======================== PAGE 1-N : RAPPORTS INDIVIDUELS ========================

def gen_individual_pages(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, period, now, show_pause=True):
    """Génère les pages de rapport individuel."""
    total_emps = len(emps)
    
    for idx, emp in enumerate(emps):
        if idx > 0: story.append(SmartPageBreak())
        
        enriched, stats = all_stats[idx]
        emp_num = idx + 1
        
        story.append(make_header(S, provider_name, provider_info, client_name, client_info))
        story.append(Spacer(1, 1*mm))
        story.append(Paragraph(f"RAPPORT INDIVIDUEL — {emp['name']} (Réf: {emp['ref']}) — Fiche {emp_num}/{total_emps} — {period}",
                              ParagraphStyle('ti2', fontName='Helvetica-Bold', fontSize=12, textColor=TEAL, alignment=TA_CENTER, spaceAfter=2)))
        story.append(Spacer(1, 1*mm))
        
        # === v67 : Mode compact si beaucoup de records (>22) pour tenir en 2 pages ===
        nb_recs_emp = len(enriched)
        is_compact = nb_recs_emp > 22
        
        # Bandeau bien visible : nombre de jours obligatoires + source
        # v67 : version compactée si beaucoup de records
        source_label = {
            'manuel': 'Configuration manuelle',
            'auto': 'Calcul automatique (calendrier - repos)'
        }.get(stats.get('days_required_source','auto'), 'Calcul automatique')
        
        if is_compact:
            # Version réduite : 1 ligne courte
            oblig_compact = Table([[Paragraph(
                f"<b>JOURS OBLIGATOIRES : {stats['days_required']} jours</b> &nbsp;&nbsp; <i style='color:#888'>{source_label}</i>",
                ParagraphStyle('oblc', fontName='Helvetica-Bold', fontSize=9,
                              textColor=DARK_TEAL, alignment=TA_CENTER))]],
                colWidths=[190*mm])
            oblig_compact.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,-1),HexColor('#fff3e0')),
                ('BOX',(0,0),(-1,-1),0.6,DARK_TEAL),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ]))
            story.append(oblig_compact)
            story.append(Spacer(1, 1.5*mm))
        else:
            oblig_data = [[
                Paragraph(f"<b>NOMBRE DE JOURS OBLIGATOIRES À EFFECTUER</b>", 
                          ParagraphStyle('oblh', fontName='Helvetica-Bold', fontSize=9,
                                        textColor=colors.white, alignment=TA_CENTER)),
                Paragraph(f"<b>{stats['days_required']} jours</b>",
                          ParagraphStyle('oblv', fontName='Helvetica-Bold', fontSize=14,
                                        textColor=DARK_TEAL, alignment=TA_CENTER)),
                Paragraph(f"<i>{source_label}</i>",
                          ParagraphStyle('obls', fontName='Helvetica-Oblique', fontSize=7,
                                        textColor=colors.grey, alignment=TA_CENTER)),
            ]]
            oblig_t = Table(oblig_data, colWidths=[80*mm, 40*mm, 70*mm])
            oblig_t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(0,0),DARK_TEAL),
                ('BACKGROUND',(1,0),(1,0),HexColor('#fff3e0')),
                ('BACKGROUND',(2,0),(2,0),HexColor('#f8faf9')),
                ('BOX',(0,0),(-1,-1),0.8,DARK_TEAL),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
            ]))
            story.append(oblig_t)
            story.append(Spacer(1, 3*mm))
        
        # Résumé compact
        # v52 : stats['days_absent'] contient déjà les absences réelles (= obligatoires - effectués)
        # v52 : stats['days_effectues'] est nouveau (= present + erreur badge)
        days_obligatoires = stats['days_required']
        days_effectues = stats.get('days_effectues', stats['days_present'] + stats['days_badge_error'])
        
        sum_hdrs = ["Jours<br/>obligat.","Jours<br/>effectués","Présent","Retard","Absent","Err.<br/>badge",
                    "","H. obligat.","H. travail.","H. retard","H. absent"]
        sum_vals = [
            f"{days_obligatoires}j", f"{days_effectues}j",
            f"{stats['days_present']}j",
            f"{stats['days_late']}j", f"{stats['days_absent']}j", f"{stats['days_badge_error']}j",
            "",
            m2h(stats['total_required']), m2h(stats['total_worked']),
            m2h(stats['total_late_mins']),
            m2h(stats['days_absent'] * (stats['total_required'] // max(stats['days_required'],1)))
        ]
        sh = [Paragraph(x, S['sh']) for x in sum_hdrs]
        sv = [Paragraph(x, S['sv']) for x in sum_vals]
        sw = [15*mm,15*mm,13*mm,13*mm,13*mm,12*mm, 4*mm, 16*mm,16*mm,15*mm,15*mm]
        stbl = Table([sh, sv], colWidths=sw)
        stbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(5,0),TEAL),('BACKGROUND',(7,0),(10,0),TEAL),
            # Mettre les 2 premières colonnes (obligatoires/effectués) en orange pour les distinguer
            ('BACKGROUND',(0,0),(1,0),HexColor('#e8672a')),
            ('GRID',(0,0),(5,-1),0.4,colors.grey),('GRID',(7,0),(10,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),1),('BOTTOMPADDING',(0,0),(-1,-1),1),
            ('LEFTPADDING',(0,0),(-1,-1),1),('RIGHTPADDING',(0,0),(-1,-1),1),
        ]))
        # is_compact a été défini plus haut (avant le bandeau jours oblig)
        story.extend([stbl, Spacer(1, 1*mm if is_compact else 3*mm)])
        
        # Tableau détail
        # v67 : ajout colonnes Pause début + Pause fin entre Arrivée et Départ
        # v69 : lisibilité prioritaire — police minimum 7pt (au lieu de 6pt en v67)
        nb_recs_for_style = len(enriched)
        # v160 : seuil dense relevé (25 → 31) pour qu'un mois complet (~26 j) garde la police lisible
        is_dense = nb_recs_for_style > 31

        if is_dense:
            # v160 : mode dense — police 8pt (était 7pt), leading 9.5 pour des lignes plus lisibles
            s_c = ParagraphStyle('s_c', fontName='Helvetica', fontSize=8, alignment=TA_CENTER, leading=9.5)
            s_cb = ParagraphStyle('s_cb', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER, leading=9.5)
            s_h = ParagraphStyle('s_h', fontName='Helvetica-Bold', fontSize=8, textColor=white, alignment=TA_CENTER, leading=10)
            s_g = ParagraphStyle('s_g', fontName='Helvetica-Bold', fontSize=8, textColor=GREEN, alignment=TA_CENTER, leading=9.5)
            s_r = ParagraphStyle('s_r', fontName='Helvetica-Bold', fontSize=8, textColor=RED, alignment=TA_CENTER, leading=9.5)
            s_b = ParagraphStyle('s_b', fontName='Helvetica-Bold', fontSize=8, textColor=BLUE, alignment=TA_CENTER, leading=9.5)
        else:
            # v160 : mode normal — police 9.5pt (était 8pt), leading 12.5 pour des lignes bien visibles
            s_c = ParagraphStyle('s_c2', fontName='Helvetica', fontSize=9.5, alignment=TA_CENTER, leading=12.5)
            s_cb = ParagraphStyle('s_cb2', fontName='Helvetica-Bold', fontSize=9.5, alignment=TA_CENTER, leading=12.5)
            s_h = ParagraphStyle('s_h2', fontName='Helvetica-Bold', fontSize=9.5, textColor=white, alignment=TA_CENTER, leading=12.5)
            s_g = ParagraphStyle('s_g2', fontName='Helvetica-Bold', fontSize=9.5, textColor=GREEN, alignment=TA_CENTER, leading=12.5)
            s_r = ParagraphStyle('s_r2', fontName='Helvetica-Bold', fontSize=9.5, textColor=RED, alignment=TA_CENTER, leading=12.5)
            s_b = ParagraphStyle('s_b2', fontName='Helvetica-Bold', fontSize=9.5, textColor=BLUE, alignment=TA_CENTER, leading=12.5)
        
        # v73 : Colonnes Pause masquées si show_pause=False
        if show_pause:
            hdrs = ["N°","Date","Planning","État","Arrivée","P.<br/>début","P.<br/>fin",
                    "Départ","H.<br/>travail.","Retard",
                    "H.<br/>obligat.","H.<br/>Respectée","H. sup."]
            # v160 : colonnes élargies (total 190mm ≤ 194mm utiles en portrait A4, marges 8mm)
            cw = [7*mm,17*mm,21*mm,17*mm,13*mm,12*mm,12*mm,13*mm,15*mm,13*mm,15*mm,19*mm,16*mm]
        else:
            hdrs = ["N°","Date","Planning","État","Arrivée",
                    "Départ","H.<br/>travail.","Retard",
                    "H.<br/>obligat.","H.<br/>Respectée","H. sup."]
            # v160 : colonnes élargies (total 190mm ≤ 194mm utiles en portrait A4, marges 8mm)
            cw = [8*mm,19*mm,24*mm,19*mm,16*mm,16*mm,17*mm,15*mm,17*mm,21*mm,18*mm]
        
        td = [[Paragraph(x, s_h) for x in hdrs]]
        
        for i, rec in enumerate(enriched, 1):
            resp = rec['respect']
            if resp == "OUI":
                rp = Paragraph("OUI", s_g)
            elif resp == "ABS":
                rp = Paragraph("ABS", s_r)
            elif resp.startswith("NON"):
                rp = Paragraph(resp.replace(" ","<br/>"), s_r)
            else:
                rp = Paragraph(resp, s_c)
            
            ot_mins = t2m(rec['overtime'])
            ot = Paragraph(rec['overtime'], s_b) if ot_mins > 0 else Paragraph(rec['overtime'], s_c)
            
            # v67 : afficher pause début/fin avec fallback '-' si absent
            ps_val = rec.get('pause_start', '') or '-'
            pe_val = rec.get('pause_end', '') or '-'
            
            if show_pause:
                td.append([
                    Paragraph(str(i), s_c),
                    Paragraph(rec['date'], s_c),
                    Paragraph(rec['schedule'], s_c),
                    Paragraph(rec['state'], s_cb),
                    Paragraph(rec['arrival'], s_c),
                    Paragraph(ps_val, s_c),
                    Paragraph(pe_val, s_c),
                    Paragraph(rec['departure'], s_c),
                    Paragraph(rec['worked'], s_cb),
                    Paragraph(rec['late'], s_c),
                    Paragraph(rec['required'], s_cb),
                    rp, ot
                ])
            else:
                td.append([
                    Paragraph(str(i), s_c),
                    Paragraph(rec['date'], s_c),
                    Paragraph(rec['schedule'], s_c),
                    Paragraph(rec['state'], s_cb),
                    Paragraph(rec['arrival'], s_c),
                    Paragraph(rec['departure'], s_c),
                    Paragraph(rec['worked'], s_cb),
                    Paragraph(rec['late'], s_c),
                    Paragraph(rec['required'], s_cb),
                    rp, ot
                ])
        
        dt = Table(td, colWidths=cw, repeatRows=1)
        # v67 : ajuster les indices BACKGROUND (header indices décalés de 2)
        # v67 : compactage adaptatif pour rester sur 2 pages max
        if is_dense:
            top_pad = 1; bot_pad = 1       # v160 : un peu d'air même en mode dense
        else:
            top_pad = 2.5; bot_pad = 2.5   # v160 : lignes plus hautes / lisibles
        sc = [('BACKGROUND',(0,0),(-1,0),TEAL),('BACKGROUND',(10,0),(12,0),DARK_TEAL),
              ('GRID',(0,0),(-1,-1),0.3,colors.grey),
              ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
              ('TOPPADDING',(0,0),(-1,-1),top_pad),('BOTTOMPADDING',(0,0),(-1,-1),bot_pad),
              ('LEFTPADDING',(0,0),(-1,-1),1),('RIGHTPADDING',(0,0),(-1,-1),1)]
        for i in range(2, len(td), 2):
            sc.append(('BACKGROUND',(0,i),(-1,i),LGRAY))
        # Separate days with thick border when date changes
        prev_date = None
        for i, rec in enumerate(enriched, 1):
            cur_date = rec['date'][:10] if rec['date'] else ''
            if prev_date and cur_date != prev_date:
                sc.append(('LINEABOVE',(0,i),(-1,i),1.5,TEAL))
            prev_date = cur_date
        dt.setStyle(TableStyle(sc))
        story.append(dt)
        
        # Totaux : v67 — version compacte si beaucoup de records
        if is_dense:
            story.append(Spacer(1, 1*mm))
        else:
            story.append(Spacer(1, 2*mm))
        tt = Table([[
            Paragraph(f"<b>TOTAL H. SUPPLÉMENTAIRES : {m2h(stats['total_overtime'])}</b>",
                ParagraphStyle('x',fontName='Helvetica-Bold',fontSize=(7 if is_dense else 8),textColor=BLUE)),
            Paragraph(f"<b>TOTAL DÉFICIT : {m2h(stats['total_deficit'])}</b>",
                ParagraphStyle('x',fontName='Helvetica-Bold',fontSize=(7 if is_dense else 8),textColor=RED)),
        ]], colWidths=[95*mm,95*mm])
        tt_pad = 1 if is_dense else 3
        tt.setStyle(TableStyle([('BOX',(0,0),(-1,-1),0.8,TEAL),
            ('INNERGRID',(0,0),(-1,-1),0.4,TEAL),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),tt_pad),('BOTTOMPADDING',(0,0),(-1,-1),tt_pad),
            ('LEFTPADDING',(0,0),(-1,-1),4)]))
        story.extend([tt, Spacer(1, 1*mm if is_dense else 2*mm)])
        
        # === ENCADRÉ COÛT (si hourly_cost > 0) ===
        # v67 : version compacte 1 ligne si rapport dense
        if stats.get('hourly_cost', 0) > 0:
            fmt_cost = lambda x: f"{x:,.0f} FCFA"
            if is_dense:
                # Version compacte 1 seule ligne
                total_perdu = stats['cost_late'] + stats['cost_deficit'] + stats['cost_absent']
                cost_compact = Table([[
                    Paragraph(f"<b>💰 IMPACT FINANCIER</b> &nbsp;&nbsp; Coût/h : <b>{fmt_cost(stats['hourly_cost'])}</b> &nbsp; Retards : <b style='color:red'>{fmt_cost(stats['cost_late'])}</b> &nbsp; Déficit : <b style='color:red'>{fmt_cost(stats['cost_deficit'])}</b> &nbsp; Absences : <b style='color:red'>{fmt_cost(stats['cost_absent'])}</b> &nbsp; <b style='color:red'>TOTAL : {fmt_cost(total_perdu)}</b>",
                              ParagraphStyle('ctc', fontName='Helvetica', fontSize=7, textColor=DARK_TEAL))
                ]], colWidths=[190*mm])
                cost_compact.setStyle(TableStyle([
                    ('BACKGROUND',(0,0),(-1,-1),HexColor('#fff3e0')),
                    ('BOX',(0,0),(-1,-1),0.6,DARK_TEAL),
                    ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                    ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
                    ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ]))
                story.extend([cost_compact, Spacer(1, 1*mm)])
            else:
                # Version complète (rapports peu denses)
                cost_data = [
                    [Paragraph("<b>💰 IMPACT FINANCIER</b>", ParagraphStyle('ct',fontName='Helvetica-Bold',fontSize=8,textColor=colors.white)),
                     Paragraph(f"<b>Coût horaire : {fmt_cost(stats['hourly_cost'])}</b>", ParagraphStyle('ct2',fontName='Helvetica-Bold',fontSize=8,textColor=colors.white,alignment=2))],
                    [Paragraph(f"Perte retards ({m2h(stats['total_late_mins'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                     Paragraph(f"<b>{fmt_cost(stats['cost_late'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                    [Paragraph(f"Perte déficit horaire ({m2h(stats['total_deficit'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                     Paragraph(f"<b>{fmt_cost(stats['cost_deficit'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                    [Paragraph(f"Perte absences ({stats['days_absent']} jour(s))", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                     Paragraph(f"<b>{fmt_cost(stats['cost_absent'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                    [Paragraph("<b>TOTAL GAIN PERDU</b>", ParagraphStyle('ct3',fontName='Helvetica-Bold',fontSize=8,textColor=RED)),
                     Paragraph(f"<b>{fmt_cost(stats['cost_late'] + stats['cost_deficit'] + stats['cost_absent'])}</b>",
                        ParagraphStyle('ct4',fontName='Helvetica-Bold',fontSize=9,textColor=RED,alignment=2))],
                ]
                if stats['cost_overtime'] > 0:
                    cost_data.append([
                        Paragraph(f"Heures sup. ({m2h(stats['total_overtime'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                        Paragraph(f"<b>+{fmt_cost(stats['cost_overtime'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=BLUE,alignment=2))
                    ])
                
                ct = Table(cost_data, colWidths=[120*mm, 70*mm])
                ct_style = [
                    ('BACKGROUND',(0,0),(-1,0),DARK_TEAL),
                    ('BACKGROUND',(0,-1),(-1,-1),HexColor('#fff3e0')),
                    ('BOX',(0,0),(-1,-1),1,DARK_TEAL),
                    ('INNERGRID',(0,0),(-1,-1),0.3,colors.grey),
                    ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                    ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
                    ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ]
                ct.setStyle(TableStyle(ct_style))
                story.extend([ct, Spacer(1,2*mm)])
        
        story.append(
            Paragraph(f"Généré le {now} | {safe(client_name)} - Rapport {safe(emp['name'])} {emp_num}/{total_emps}", S['ft']))

# ======================== PAGE : RAPPORT DE PRÉSENCE ========================

def gen_rapport_presence(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now, subtitle=None):
    # v163 : ne rien générer si aucun employé (évite un tableau vide)
    if not emps:
        return
    story.append(SmartPageBreak())
    story.append(make_header(S, provider_name, provider_info, client_name, client_info))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("RAPPORT DE PRÉSENCE", S['big_ti']))
    if subtitle:
        story.append(Paragraph(subtitle, ParagraphStyle('rp_sub', fontName='Helvetica-Bold', fontSize=10, textColor=TEAL, alignment=TA_CENTER, leading=13)))
    story.append(Spacer(1, 4*mm))
    
    hdrs = ["N°","Employé","Jours<br/>obligat.","Jours de<br/>présence",
            "Taux<br/>présence","Jours de<br/>retards","Jours<br/>ponctuel","Jours<br/>d'absences",
            "Observation"]
    hrow = [Paragraph(h, S['rh']) for h in hdrs]
    cw = [8*mm, 30*mm, 16*mm, 16*mm, 16*mm, 16*mm, 16*mm, 16*mm, 28*mm]
    
    td = [hrow]
    for i, (emp, (enriched, stats)) in enumerate(zip(emps, all_stats), 1):
        obs = stats['observation']
        if obs == "Non assidu":
            obs_style = S['rvo']
        elif obs == "Moyennement assidu":
            obs_style = ParagraphStyle('rvblue', fontName='Helvetica-Bold', fontSize=7, textColor=BLUE, alignment=TA_CENTER, leading=9)
        else:
            obs_style = ParagraphStyle('rvgreen', fontName='Helvetica-Bold', fontSize=7, textColor=GREEN, alignment=TA_CENTER, leading=9)
        td.append([
            Paragraph(str(i), S['rv']),
            Paragraph(emp['name'], S['rvb']),
            Paragraph(f"{stats['days_required']} j", S['rv']),
            Paragraph(f"{stats['days_present']} j", S['rv']),
            Paragraph(f"{stats.get('presence_rate', 0):.0f}%", S['rv']),
            Paragraph(f"{stats['days_late']} j", S['rv']),
            Paragraph(f"{stats['days_punctual']} j", S['rv']),
            Paragraph(f"{stats['days_absent']} j", S['rv']),
            Paragraph(obs, obs_style),
        ])
    
    t = Table(td, colWidths=cw, repeatRows=1)
    sc = [('BACKGROUND',(0,0),(-1,0),TEAL),
          ('GRID',(0,0),(-1,-1),0.4,colors.grey),
          ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
          ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]
    for i in range(2, len(td), 2):
        sc.append(('BACKGROUND',(0,i),(-1,i),LGRAY))
    t.setStyle(TableStyle(sc))
    story.extend([t, Spacer(1,4*mm),
        Paragraph(f"Généré le {now} | {safe(client_name)} - Rapport de Présence", S['ft'])])

# ======================== PAGE : CLASSEMENT RETARDS & ABSENCES ========================

def gen_classement(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now):
    story.append(SmartPageBreak())
    story.append(make_header(S, provider_name, provider_info, client_name, client_info))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("CLASSEMENT PAR DEGRÉ DE RETARDS ET D'ABSENCES", S['big_ti']))
    story.append(Spacer(1, 6*mm))
    
    # Classement par retards
    retards = [(emp['name'], stats['total_late_mins']) 
               for emp, (_, stats) in zip(emps, all_stats) if stats['total_late_mins'] > 0]
    retards.sort(key=lambda x: -x[1])
    
    story.append(Paragraph("Classement par Retards", S['med_ti']))
    hdrs = [Paragraph(h, S['rh']) for h in ["Rang","Nom Employé","Total heure de retard"]]
    td_r = [hdrs]
    for i, (name, mins) in enumerate(retards[:10], 1):
        td_r.append([Paragraph(str(i), S['rv']), Paragraph(name, S['rvb']),
                     Paragraph(m2h(mins), S['rv'])])
    
    if len(td_r) > 1:
        tr = Table(td_r, colWidths=[15*mm, 80*mm, 40*mm])
        tr.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),TEAL),
            ('GRID',(0,0),(-1,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
        story.append(tr)
    
    story.append(Spacer(1, 8*mm))
    
    # Classement par absences
    absences = [(emp['name'], stats['days_absent'])
                for emp, (_, stats) in zip(emps, all_stats) if stats['days_absent'] > 0]
    absences.sort(key=lambda x: -x[1])
    
    story.append(Paragraph("Classement par Absences", S['med_ti']))
    hdrs2 = [Paragraph(h, S['rh']) for h in ["Rang","Nom Employé","Nombre de jours d'absence"]]
    td_a = [hdrs2]
    for i, (name, days) in enumerate(absences[:10], 1):
        td_a.append([Paragraph(str(i), S['rv']), Paragraph(name, S['rvb']),
                     Paragraph(str(days), S['rv'])])
    
    if len(td_a) > 1:
        ta = Table(td_a, colWidths=[15*mm, 80*mm, 45*mm])
        ta.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),TEAL),
            ('GRID',(0,0),(-1,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
        story.append(ta)
    
    story.extend([Spacer(1,4*mm), Paragraph(f"Généré le {now}", S['ft'])])

# ======================== PRÉPARATION LOGO ========================

def _prepare_logo(logo_path, work_dir=None):
    """Supprime le fond noir du logo et retourne le chemin du logo nettoyé."""
    try:
        from PIL import Image
        import numpy as np
        
        img = Image.open(logo_path).convert('RGBA')
        data = np.array(img)
        
        dark = (data[:,:,0] < 45) & (data[:,:,1] < 45) & (data[:,:,2] < 45)
        data[dark, 3] = 0
        
        img_clean = Image.fromarray(data)
        bbox = img_clean.getbbox()
        if bbox:
            img_clean = img_clean.crop(bbox)
        
        w, h = img_clean.size
        size = max(w, h)
        square = Image.new('RGBA', (size, size), (255, 255, 255, 0))
        square.paste(img_clean, ((size - w) // 2, (size - h) // 2), img_clean)
        
        out_dir = work_dir or os.path.dirname(os.path.abspath(logo_path))
        clean_path = os.path.join(out_dir, 'logo_clean.png')
        square.save(clean_path)
        
        return clean_path
    except Exception as e:
        print(f"  ⚠️  Erreur traitement logo: {e}")
        return None


def _generate_chart_image(pct_presence, pct_absence, logo_path=None, work_dir=None,
                          total_presence_min=0, total_absence_min=0):
    """Génère un graphique style donut moderne avec léger volume — HAUTE DÉFINITION (3x).
    Logo au centre, labels de % flottants au-dessus de chaque secteur, légende en bas.
    Total minutes optionnels pour afficher 'XXXh' dans la légende.
    Image générée en 2400x2100 pixels (3x) pour rendu net dans le PDF (300dpi)."""
    # NOUVEAU v53 : sécurité - cap les pourcentages à [0, 100] et ajuste pour qu'ils somment à 100
    pct_presence = max(0.0, min(float(pct_presence), 100.0))
    pct_absence = max(0.0, min(float(pct_absence), 100.0))
    # Si la somme dépasse 100, redimensionner pour que ça matche
    total_pct = pct_presence + pct_absence
    if total_pct > 100.01:  # marge pour erreurs flottantes
        pct_presence = (pct_presence / total_pct) * 100
        pct_absence = (pct_absence / total_pct) * 100
    try:
        from PIL import Image, ImageDraw, ImageFont, ImageFilter
        import math
        
        # === FACTEUR D'ÉCHELLE pour rendu HD (3x) ===
        SCALE = 3
        W, H = 800 * SCALE, 700 * SCALE
        cx, cy = 400 * SCALE, 280 * SCALE
        outer_r = 200 * SCALE     # rayon extérieur du donut
        inner_r = 95 * SCALE      # rayon trou central (pour logo)
        depth = 8 * SCALE         # léger volume 3D (subtil, pas exagéré)
        
        # Image RGB blanche
        base = Image.new('RGB', (W, H), (255, 255, 255))
        
        # Couleurs (style image utilisateur)
        teal = (15, 121, 100)        # vert teal foncé pour Présence
        teal_dark = (10, 90, 75)
        red = (235, 50, 60)          # rouge vif pour Absence
        red_dark = (180, 35, 45)
        text_dark = (40, 60, 75)
        text_grey = (110, 110, 110)
        
        # === 1. OMBRE PORTÉE LÉGÈRE (subtile) ===
        shadow = Image.new('RGBA', (W, H), (0, 0, 0, 0))
        sdraw = ImageDraw.Draw(shadow)
        sdraw.ellipse([cx - outer_r + 4*SCALE, cy - outer_r + 8*SCALE + depth,
                       cx + outer_r + 4*SCALE, cy + outer_r + 8*SCALE + depth],
                      fill=(0, 0, 0, 50))
        try: shadow = shadow.filter(ImageFilter.GaussianBlur(radius=10*SCALE))
        except: pass
        base.paste(shadow, (0, 0), shadow)
        
        draw = ImageDraw.Draw(base)
        
        # === 2. TRANCHE 3D LÉGÈRE ===
        bbox_top = [cx - outer_r, cy - outer_r, cx + outer_r, cy + outer_r]
        if pct_absence <= 0:
            for d in range(depth, 0, -1):
                draw.ellipse([cx - outer_r, cy - outer_r + d, cx + outer_r, cy + outer_r + d], fill=teal_dark)
        elif pct_presence <= 0:
            for d in range(depth, 0, -1):
                draw.ellipse([cx - outer_r, cy - outer_r + d, cx + outer_r, cy + outer_r + d], fill=red_dark)
        else:
            angle_p = 360 * pct_presence / 100
            for d in range(depth, 0, -1):
                bbox_d = [cx - outer_r, cy - outer_r + d, cx + outer_r, cy + outer_r + d]
                draw.pieslice(bbox_d, start=-90, end=-90 + angle_p, fill=teal_dark)
                draw.pieslice(bbox_d, start=-90 + angle_p, end=270, fill=red_dark)
        
        # === 3. DISQUE DU DESSUS (donut avec secteurs) ===
        if pct_absence <= 0:
            draw.ellipse(bbox_top, fill=teal)
        elif pct_presence <= 0:
            draw.ellipse(bbox_top, fill=red)
        else:
            angle_p = 360 * pct_presence / 100
            draw.pieslice(bbox_top, start=-90, end=-90 + angle_p, fill=teal)
            draw.pieslice(bbox_top, start=-90 + angle_p, end=270, fill=red)
        
        # === 4. TROU CENTRAL (donut) - blanc ===
        draw.ellipse([cx - inner_r, cy - inner_r, cx + inner_r, cy + inner_r], fill=(255, 255, 255))
        draw.ellipse([cx - inner_r, cy - inner_r, cx + inner_r, cy + inner_r],
                    outline=(220, 220, 220), width=1*SCALE)
        
        # === 5. LOGO RAMYA AU CENTRE (dans le trou) ===
        if logo_path and os.path.exists(logo_path):
            try:
                clean_path = _prepare_logo(logo_path, work_dir)
                if clean_path:
                    logo = Image.open(clean_path).convert('RGBA')
                    logo_size = int(inner_r * 1.5)
                    logo.thumbnail((logo_size, logo_size), Image.LANCZOS)
                    lw, lh = logo.size
                    base.paste(logo, (cx - lw // 2, cy - lh // 2), logo)
            except Exception as e:
                print(f"  Logo: {e}")
        
        # Petit cercle blanc autour du logo (effet badge)
        badge_r = inner_r - 8*SCALE
        draw.ellipse([cx - badge_r, cy - badge_r, cx + badge_r, cy + badge_r],
                    outline=teal, width=2*SCALE)
        
        # === 6. LABELS DE % FLOTTANTS (sur chaque secteur) ===
        try:
            font_pct = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 32*SCALE)
            font_pct_red = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 26*SCALE)
            font_legend = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14*SCALE)
            font_legend_b = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14*SCALE)
            font_title = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 22*SCALE)
        except:
            font_pct = font_pct_red = font_legend = font_legend_b = font_title = ImageFont.load_default()
        
        if pct_presence > 0 and pct_absence > 0:
            # NOUVEAU v52 : labels parfaitement centrés dans l'épaisseur du donut
            # et taille de police adaptée si le secteur est petit
            angle_p = 360 * pct_presence / 100
            mid_p = -90 + angle_p / 2
            mid_p_rad = math.radians(mid_p)
            # Centre exact entre rayon interne et externe (pas de débordement)
            label_r = (outer_r + inner_r) / 2
            lpx = cx + int(label_r * math.cos(mid_p_rad))
            lpy = cy + int(label_r * math.sin(mid_p_rad))
            text_p = f"{pct_presence:.1f}%"
            # Adapter la taille selon la portion (si petit secteur, font plus petite)
            font_p_used = font_pct if pct_presence >= 15 else font_pct_red
            bbox = draw.textbbox((0, 0), text_p, font=font_p_used)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            draw.text((lpx - tw // 2, lpy - th // 2), text_p, fill=(255, 255, 255), font=font_p_used)
            
            angle_a = 360 - angle_p
            mid_a = -90 + angle_p + angle_a / 2
            mid_a_rad = math.radians(mid_a)
            lax = cx + int(label_r * math.cos(mid_a_rad))
            lay = cy + int(label_r * math.sin(mid_a_rad))
            text_a = f"{pct_absence:.1f}%"
            font_a_used = font_pct if pct_absence >= 15 else font_pct_red
            bbox = draw.textbbox((0, 0), text_a, font=font_a_used)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            draw.text((lax - tw // 2, lay - th // 2), text_a, fill=(255, 255, 255), font=font_a_used)
        elif pct_presence > 0:
            text_p = f"{pct_presence:.1f}%"
            bbox = draw.textbbox((0, 0), text_p, font=font_pct)
            tw = bbox[2] - bbox[0]
            draw.text((cx - tw // 2, cy - outer_r - 35*SCALE), text_p, fill=teal, font=font_pct)
        else:
            text_a = f"{pct_absence:.1f}%"
            bbox = draw.textbbox((0, 0), text_a, font=font_pct)
            tw = bbox[2] - bbox[0]
            draw.text((cx - tw // 2, cy - outer_r - 35*SCALE), text_a, fill=red, font=font_pct)
        
        # === 7. LABELS SOUS LE GRAPHIQUE ===
        h_p = total_presence_min / 60.0 if total_presence_min else 0
        h_a = total_absence_min / 60.0 if total_absence_min else 0
        
        label_y = cy + outer_r + 40*SCALE
        text_left = f"Présence ({h_p:.1f}h - {pct_presence:.0f}%)" if total_presence_min else f"Présence : {pct_presence:.1f}%"
        text_right = f"Absence ({h_a:.1f}h - {pct_absence:.0f}%)" if total_absence_min else f"Absence : {pct_absence:.1f}%"
        
        bbox = draw.textbbox((0, 0), text_left, font=font_legend_b)
        tw_l = bbox[2] - bbox[0]
        draw.text((cx - outer_r - 40*SCALE, label_y), text_left, fill=text_dark, font=font_legend_b)
        
        bbox = draw.textbbox((0, 0), text_right, font=font_legend_b)
        tw_r = bbox[2] - bbox[0]
        draw.text((cx + outer_r - tw_r + 40*SCALE, label_y), text_right, fill=text_dark, font=font_legend_b)
        
        # === 8. LÉGENDE EN BAS (cartouches verts/rouges) ===
        legend_y = cy + outer_r + 90*SCALE
        leg_w = 280*SCALE
        leg_h = 36*SCALE
        # Vert
        leg_x_g = cx - leg_w - 15*SCALE
        draw.rectangle([leg_x_g, legend_y, leg_x_g + leg_w, legend_y + leg_h], fill=teal)
        h_p_int = int(total_presence_min // 60) if total_presence_min else 0
        m_p_int = int(total_presence_min % 60) if total_presence_min else 0
        text_g = f"Total heure de présence : {h_p_int:02d}:{m_p_int:02d}"
        bbox = draw.textbbox((0, 0), text_g, font=font_legend_b)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        draw.text((leg_x_g + (leg_w - tw) // 2, legend_y + (leg_h - th) // 2 - 2*SCALE),
                 text_g, fill=(255, 255, 255), font=font_legend_b)
        
        # Rouge
        leg_x_r = cx + 15*SCALE
        draw.rectangle([leg_x_r, legend_y, leg_x_r + leg_w, legend_y + leg_h], fill=red)
        h_a_int = int(total_absence_min // 60) if total_absence_min else 0
        m_a_int = int(total_absence_min % 60) if total_absence_min else 0
        text_r = f"Total heure d'absence : {h_a_int:02d}:{m_a_int:02d}"
        bbox = draw.textbbox((0, 0), text_r, font=font_legend_b)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        draw.text((leg_x_r + (leg_w - tw) // 2, legend_y + (leg_h - th) // 2 - 2*SCALE),
                 text_r, fill=(255, 255, 255), font=font_legend_b)
        
        out_dir = work_dir or (os.path.dirname(os.path.abspath(logo_path)) if logo_path else '/tmp')
        chart_path = os.path.join(out_dir, '_chart_donut.png')
        # Sauvegarde avec qualité maximum + DPI métadonnée
        base.save(chart_path, 'PNG', quality=100, optimize=False, dpi=(300, 300))
        return chart_path
        
    except Exception as e:
        print(f"  Erreur chart: {e}")
        import traceback; traceback.print_exc()
        return None


def gen_graphique(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now, logo_path=None, work_dir=None):
    story.append(SmartPageBreak())
    story.append(make_header(S, provider_name, provider_info, client_name, client_info))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("GRAPHIQUE D'ASSIDUITÉ DU MOIS", S['big_ti']))
    story.append(Spacer(1, 8*mm))
    
    # Calculs globaux
    total_presence = sum(s['total_worked'] for _, s in all_stats)
    total_required = sum(s['total_required'] for _, s in all_stats)
    total_absence = max(0, total_required - total_presence)
    
    if total_required > 0:
        # NOUVEAU v53 : cap à 100% (cas où des heures sup font dépasser le total obligatoire)
        pct_presence = min((total_presence / total_required) * 100, 100.0)
        pct_absence = max(0.0, 100.0 - pct_presence)
    else:
        pct_presence = 100
        pct_absence = 0
    
    # Générer le graphique en image PIL pour gérer la transparence du logo
    chart_path = _generate_chart_image(pct_presence, pct_absence, logo_path, work_dir,
                                       total_presence_min=total_presence,
                                       total_absence_min=total_absence)
    
    if chart_path:
        from reportlab.platypus import Image as PLImage
        img = PLImage(chart_path, width=160*mm, height=140*mm)
        story.append(img)
    
    story.append(Spacer(1, 2*mm))
    
    # Légendes texte
    story.append(Paragraph(
        f"<font color='#E85D4A'><b>Absence ({m2h(total_absence)} - {pct_absence:.0f}%)</b></font>",
        ParagraphStyle('la', fontSize=10, alignment=TA_RIGHT, spaceAfter=2)))
    story.append(Paragraph(
        f"<font color='#1A7A6D'><b>Présence ({m2h(total_presence)} - {pct_presence:.0f}%)</b></font>",
        ParagraphStyle('lp', fontSize=10, alignment=TA_LEFT, spaceAfter=4)))
    
    story.append(Spacer(1, 2*mm))
    
    # Légende en bas
    leg = Table([[
        Paragraph(f"<font color='#1A7A6D'><b>■</b></font>  Total heure de présence: {m2h(total_presence)}", 
                  ParagraphStyle('l1', fontSize=9)),
        Paragraph(f"<font color='#E85D4A'><b>■</b></font>  <font color='#E85D4A'>Total heure d'absence: {m2h(total_absence)}</font>",
                  ParagraphStyle('l2', fontSize=9)),
    ]], colWidths=[95*mm, 95*mm])
    leg.setStyle(TableStyle([('BOX',(0,0),(-1,-1),0.5,colors.grey),
        ('INNERGRID',(0,0),(-1,-1),0.5,colors.grey),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
        ('LEFTPADDING',(0,0),(-1,-1),6)]))
    story.extend([leg, Paragraph(f"Généré le {now} | {safe(client_name)} - Graphique d'Assiduité", S['ft'])])

# ======================== FICHE DE PRÉSENCE SIMPLE ========================

def gen_simple_pages(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, period, now):
    """Génère une fiche de présence simple : uniquement N°, Date, Planning, Arrivée, Départ — sans retards, absences, totaux."""
    
    for idx, emp in enumerate(emps):
        if idx > 0: story.append(SmartPageBreak())
        
        enriched, stats = all_stats[idx]
        
        story.append(make_header(S, provider_name, provider_info, client_name, client_info))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph("RAPPORT INDIVIDUEL", S['ti']))
        story.append(Paragraph(period, S['st']))
        story.append(Paragraph(f"Employé: {emp['name']}  |  Réf: {emp['ref']}", S['ei']))
        story.append(Spacer(1, 3*mm))
        
        # Résumé ultra-compact : jours prévus seulement
        sum_data = [[
            Paragraph("<b>Nbre de jours à Effectuer</b>", S['sh']),
            Paragraph(f"<b>{stats['days_required']} jours</b>", S['sv']),
        ]]
        st = Table(sum_data, colWidths=[95*mm, 95*mm])
        st.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(0,0),TEAL), ('BACKGROUND',(1,0),(1,0),HexColor('#f8faf9')),
            ('BOX',(0,0),(-1,-1),0.5,TEAL),
            ('ALIGN',(0,0),(-1,-1),'CENTER'), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        story.extend([st, Spacer(1, 3*mm)])
        
        # Tableau simple : N°, Date, Emploi du temps, Heure d'arrivée, Heure de départ
        hdrs = ["N°", "Date", "Emploi du temps", "Heure d'arrivée", "Heure de départ"]
        cw = [10*mm, 25*mm, 40*mm, 55*mm, 55*mm]
        
        td = [[Paragraph(x, S['h']) for x in hdrs]]
        
        for i, rec in enumerate(enriched, 1):
            td.append([
                Paragraph(str(i), S['c']),
                Paragraph(rec['date'], S['c']),
                Paragraph(rec['schedule'], S['c']),
                Paragraph(rec['arrival'] if rec['arrival'] and rec['state'] != 'Absent(e)' else '-', S['c']),
                Paragraph(rec['departure'] if rec['departure'] and rec['state'] != 'Absent(e)' else '', S['c']),
            ])
        
        dt = Table(td, colWidths=cw, repeatRows=1)
        sc = [('BACKGROUND',(0,0),(-1,0),TEAL),
              ('GRID',(0,0),(-1,-1),0.3,colors.grey),
              ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
              ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
              ('LEFTPADDING',(0,0),(-1,-1),2),('RIGHTPADDING',(0,0),(-1,-1),2)]
        for i in range(2, len(td), 2):
            sc.append(('BACKGROUND',(0,i),(-1,i),LGRAY))
        dt.setStyle(TableStyle(sc))
        story.append(dt)
        
        # Coût si applicable
        if stats.get('hourly_cost', 0) > 0 and (stats['cost_late'] > 0 or stats['cost_deficit'] > 0 or stats['cost_absent'] > 0):
            story.append(Spacer(1, 3*mm))
            fmt_cost = lambda x: f"{x:,.0f} FCFA"
            total_loss = stats['cost_late'] + stats['cost_deficit'] + stats['cost_absent']
            cost_line = Table([[
                Paragraph(f"<b>💰 Coût horaire: {fmt_cost(stats['hourly_cost'])}</b>", ParagraphStyle('x',fontName='Helvetica-Bold',fontSize=8,textColor=DARK_TEAL)),
                Paragraph(f"<b>Gain perdu: {fmt_cost(total_loss)}</b>", ParagraphStyle('x',fontName='Helvetica-Bold',fontSize=8,textColor=RED,alignment=2)),
            ]], colWidths=[95*mm, 95*mm])
            cost_line.setStyle(TableStyle([('BOX',(0,0),(-1,-1),0.8,DARK_TEAL),
                ('INNERGRID',(0,0),(-1,-1),0.4,DARK_TEAL),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6)]))
            story.append(cost_line)
        
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(f"Imprimé par : RH, le {now}", S['ft']))

# ======================== FICHE DE PRÉSENCE SIMPLE ========================

def gen_simple_pages(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, period, now):
    """Génère une fiche simple : N°, Date, Planning, Arrivée, Départ — sans retard/absence/totaux."""
    
    for idx, emp in enumerate(emps):
        if idx > 0: story.append(SmartPageBreak())
        
        enriched, stats = all_stats[idx]
        
        story.append(make_header(S, provider_name, provider_info, client_name, client_info))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph("RAPPORT INDIVIDUEL", S['ti']))
        story.append(Paragraph(period, S['st']))
        story.append(Paragraph(f"Employé : {emp['name']}  |  Réf : {emp['ref']}", S['ei']))
        story.append(Spacer(1, 2*mm))
        
        # Summary: just days count
        sum_hdrs = ["Nbre de jours à Effectuer", "Ponctuel", "Retard", "Absent", "Erreurs de Badge"]
        sum_vals = [
            f"{stats['days_required']} jours", f"{stats['days_punctual']} jours",
            f"{stats['days_late']} jours", f"{stats['days_absent']} jours", f"{stats['days_badge_error']} jours",
        ]
        sh = [Paragraph(x, S['sh']) for x in sum_hdrs]
        sv = [Paragraph(x, S['sv']) for x in sum_vals]
        sw = [36*mm, 30*mm, 28*mm, 28*mm, 28*mm]
        stbl = Table([sh, sv], colWidths=sw)
        stbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),TEAL),
            ('GRID',(0,0),(-1,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ]))
        story.extend([stbl, Spacer(1, 2*mm)])
        
        # Hours summary row
        hrs_hdrs = ["Total heure obligatoire", "Présence", "Retard", "Absent"]
        hrs_vals = [
            m2h(stats['total_required']) + " heures",
            m2h(stats['total_worked']) + " heures",
            m2h(stats['total_late_mins']) + " heures",
            m2h(stats['days_absent'] * (stats['total_required'] // max(stats['days_required'],1))) + " heures",
        ]
        hh = [Paragraph(x, S['sh']) for x in hrs_hdrs]
        hv = [Paragraph(x, S['sv']) for x in hrs_vals]
        hw = [40*mm, 40*mm, 35*mm, 35*mm]
        htbl = Table([hh, hv], colWidths=hw)
        htbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),TEAL),
            ('GRID',(0,0),(-1,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ]))
        story.extend([htbl, Spacer(1, 3*mm)])
        
        # Detail table: simple — N°, Date, Emploi du temps, Arrivée, Départ
        hdrs = ["N°", "Date", "Emploi du temps", "Heure d'arrivée", "Heure de départ"]
        cw = [12*mm, 28*mm, 40*mm, 40*mm, 40*mm]
        
        td = [[Paragraph(x, S['h']) for x in hdrs]]
        
        for i, rec in enumerate(enriched, 1):
            td.append([
                Paragraph(str(i), S['c']),
                Paragraph(rec['date'], S['c']),
                Paragraph(rec['schedule'], S['c']),
                Paragraph(rec['arrival'] if rec['arrival'] != '00:00' else '-', S['c']),
                Paragraph(rec['departure'] if rec['departure'] != '00:00' else '-', S['c']),
            ])
        
        dt = Table(td, colWidths=cw, repeatRows=1)
        sc = [('BACKGROUND',(0,0),(-1,0),TEAL),
              ('GRID',(0,0),(-1,-1),0.3,colors.grey),
              ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
              ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
              ('LEFTPADDING',(0,0),(-1,-1),2),('RIGHTPADDING',(0,0),(-1,-1),2)]
        for i in range(2, len(td), 2):
            sc.append(('BACKGROUND',(0,i),(-1,i),LGRAY))
        dt.setStyle(TableStyle(sc))
        story.append(dt)
        
        # Cost box if applicable
        if stats.get('hourly_cost', 0) > 0:
            story.append(Spacer(1, 2*mm))
            fmt_cost = lambda x: f"{x:,.0f} FCFA"
            total_lost = stats['cost_late'] + stats['cost_deficit'] + stats['cost_absent']
            cost_data = [
                [Paragraph("<b>💰 IMPACT FINANCIER</b>", ParagraphStyle('ct',fontName='Helvetica-Bold',fontSize=8,textColor=colors.white)),
                 Paragraph(f"<b>Coût horaire : {fmt_cost(stats['hourly_cost'])}</b>", ParagraphStyle('ct2',fontName='Helvetica-Bold',fontSize=8,textColor=colors.white,alignment=2))],
                [Paragraph(f"Perte retards ({m2h(stats['total_late_mins'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_late'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                [Paragraph(f"Perte déficit ({m2h(stats['total_deficit'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_deficit'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                [Paragraph(f"Perte absences ({stats['days_absent']}j)", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_absent'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                [Paragraph("<b>TOTAL GAIN PERDU</b>", ParagraphStyle('ct3',fontName='Helvetica-Bold',fontSize=8,textColor=RED)),
                 Paragraph(f"<b>{fmt_cost(total_lost)}</b>", ParagraphStyle('ct4',fontName='Helvetica-Bold',fontSize=9,textColor=RED,alignment=2))],
            ]
            ct = Table(cost_data, colWidths=[100*mm, 60*mm])
            ct.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),DARK_TEAL),('BACKGROUND',(0,-1),(-1,-1),HexColor('#fff3e0')),
                ('BOX',(0,0),(-1,-1),1,DARK_TEAL),('INNERGRID',(0,0),(-1,-1),0.3,colors.grey),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
            ]))
            story.append(ct)
        
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(f"Imprimé par : RH, le {now}", S['ft']))


# ======================== GENERATION PDF COMPLETE ========================

def generate_full_pdf(emps, output_path, provider_name, provider_info, client_name, period, logo_path=None, hp=0, client_info="", work_dir=None, hp_weekend=0, hourly_cost=0, employee_costs=None, rest_days=None, employee_rest_days=None, days_required_default=None, employee_days_required=None, employee_hours=None, pause_minutes=0, auto_invert_night=False):
    if not employee_costs: employee_costs = {}
    if rest_days is None: rest_days = []
    if employee_rest_days is None: employee_rest_days = {}
    if employee_days_required is None: employee_days_required = {}
    if employee_hours is None: employee_hours = {}
    if not work_dir:
        work_dir = os.path.dirname(os.path.abspath(output_path))
    # NOUVEAU v56 : retour en portrait A4 — 1 page par employé garanti via layout compact
    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=8*mm, rightMargin=8*mm, topMargin=6*mm, bottomMargin=6*mm)
    S = make_styles()
    story = []
    now = datetime.now().strftime("%d/%m/%Y à %H:%M")
    
    # Pré-calculer toutes les stats avec coût par employé, jours de repos et jours obligatoires
    # NOUVEAU v49 : period_total_days = jours uniques entre la date min et max de TOUTE la période
    # → garantit le même nombre de jours obligatoires pour tous les employés
    all_dates_set = set()
    for e in emps:
        for r in e.get('records', []):
            d = r.get('date')
            if d: all_dates_set.add(d)
    period_total_days = len(all_dates_set) if all_dates_set else 30
    
    # v58 : matching tolérant aux espaces/casse pour les overrides par employé
    def _norm_name(s): return ''.join((s or '').strip().upper().split())
    employee_costs_norm = {_norm_name(k): v for k, v in employee_costs.items()}
    employee_rest_norm = {_norm_name(k): v for k, v in employee_rest_days.items()}
    employee_days_norm = {_norm_name(k): v for k, v in employee_days_required.items()}
    employee_hours_norm = {_norm_name(k): v for k, v in employee_hours.items()}
    
    all_stats = []
    for emp in emps:
        norm_key = _norm_name(emp['name'])
        emp_cost = employee_costs_norm.get(norm_key, hourly_cost)
        emp_rest = employee_rest_norm.get(norm_key, rest_days)
        # days_required : override individuel > défaut entreprise > calcul auto
        emp_days_req = employee_days_norm.get(norm_key)
        if emp_days_req is None and days_required_default is not None and days_required_default > 0:
            emp_days_req = days_required_default
        # NOUVEAU v57 : heures par jour par employé (matching tolérant v58)
        emp_hp = employee_hours_norm.get(norm_key, hp)
        all_stats.append(calc_employee_stats(emp, emp_hp, hp_weekend, emp_cost, rest_days=emp_rest,
                                             days_required_override=emp_days_req,
                                             period_total_days=period_total_days,
                                             pause_minutes=pause_minutes,
                                             auto_invert_night=auto_invert_night))
    
    # 1. Rapports individuels
    # v73 : afficher les colonnes Pause uniquement si la pause est activée
    show_pause = pause_minutes > 0
    gen_individual_pages(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, period, now, show_pause=show_pause)
    
    # 2. Rapport de présence
    gen_rapport_presence(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now)
    
    # 3. Classement retards & absences
    gen_classement(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now)
    
    # 4. Graphique d'assiduité
    gen_graphique(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now, logo_path, work_dir)
    
    doc.build(story)

# ======================== MAIN ========================

def main():
    print("\n╔══════════════════════════════════════════════════════════════╗")
    print("║   Générateur de Rapport de Pointage Enrichi               ║")
    print("║   (heures sup / respect horaire / classement / graphique) ║")
    print("╚══════════════════════════════════════════════════════════════╝\n")
    
    # ---- 1. Fichier Excel ----
    file_path = (sys.argv[1] if len(sys.argv)>1 
                 else input("📄 Fichier Excel (.xlsx) — glissez-déposez : ").strip().strip('"').strip("'"))
    
    if not os.path.exists(file_path):
        print(f"\n❌ Fichier '{file_path}' introuvable."); sys.exit(1)
    print(f"  ✅ {os.path.basename(file_path)}")
    
    # ---- 2. Extraction ----
    print(f"\n  🔄 Extraction des données...")
    emps, detected_client = extract_from_excel(file_path)
    
    if not emps:
        print("\n  ❌ Aucun employé trouvé."); sys.exit(1)
    
    print(f"  ✅ {len(emps)} employé(s) détecté(s)")
    
    # ---- 3. Noms des entreprises ----
    print(f"\n  🏢 VOTRE SOCIÉTÉ (Entrée = RAMYA TECHNOLOGIE & INNOVATION) :")
    new_prov = input("     → ").strip()
    provider_name = new_prov if new_prov else "RAMYA TECHNOLOGIE & INNOVATION"
    
    if new_prov:
        new_info = input("     Tél & Email : ").strip()
        provider_info = new_info if new_info else "Tél: 2722204498 | Email: techniqueramya@gmail.com"
    else:
        provider_info = "Tél: 2722204498 | Email: techniqueramya@gmail.com"
    
    print(f"\n  🏬 ENTREPRISE CLIENTE (Entrée = {detected_client}) :")
    new_client = input("     → ").strip()
    client_name = new_client if new_client else detected_client
    
    # ---- 4. Période ----
    # Détecter depuis les dates des enregistrements
    all_dates = []
    for emp in emps:
        for rec in emp['records']:
            all_dates.append(rec['date'])
    if all_dates:
        all_dates.sort()
        period = f"Période du {all_dates[0]} au {all_dates[-1]}"
    else:
        period = "Rapport de pointage"
    
    print(f"\n  📅 {period}")
    
    # ---- 5. Logo ----
    # Chercher le logo dans le même dossier que le fichier Excel, ou le dossier du script
    logo_path = None
    search_dirs = [os.path.dirname(os.path.abspath(file_path)), os.getcwd(), 
                   os.path.dirname(os.path.abspath(__file__))]
    logo_names = ['logo_ramya_ROIND.png', 'logo_ramya.png', 'logo.png']
    for d in search_dirs:
        for ln in logo_names:
            candidate = os.path.join(d, ln)
            if os.path.exists(candidate):
                logo_path = candidate
                break
        if logo_path:
            break
    
    if logo_path:
        print(f"  🖼️  Logo trouvé : {os.path.basename(logo_path)}")
    else:
        print(f"  ℹ️  Pas de logo trouvé (placez logo_ramya_ROIND.png dans le même dossier)")
    
    # ---- 6. Liste des employés ----
    print(f"\n  👥 Employés :")
    for i, emp in enumerate(emps, 1):
        print(f"     {i:2d}. {emp['name']:<25s} ({emp['ref']}) → {len(emp['records'])} jours")
    
    # ---- 6. Génération ----
    base = os.path.splitext(os.path.basename(file_path))[0]
    out_dir = os.path.dirname(os.path.abspath(file_path))
    out = os.path.join(out_dir, f"{base}_RAPPORT_COMPLET.pdf")
    try:
        with open(out, 'wb') as f: pass
        os.remove(out)
    except OSError:
        out = os.path.join(os.getcwd(), f"{base}_RAPPORT_COMPLET.pdf")
    
    print(f"\n  🔄 Génération du PDF complet...")
    generate_full_pdf(emps, out, provider_name, provider_info, client_name, period, logo_path)
    
    print(f"\n  ✅ SUCCÈS → {out}")
    print(f"     🏢 {provider_name} → {client_name}")
    print(f"     👥 {len(emps)} employés")
    print(f"     📄 Contenu : Rapports individuels + Présence + Classement + Graphique\n")
    return out

if __name__ == "__main__":
    main()
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fusion de fichiers Enregistrement des arrivées/départs + Transactions
→ Génère le fichier Présence au format attendu par rapport_core.py
"""

import openpyxl
from collections import defaultdict
from datetime import datetime, timedelta
import os


def parse_time_str(val):
    """Convertit une valeur en string HH:MM."""
    if val is None or str(val).strip() in ('', '-', 'None'):
        return None
    s = str(val).strip()
    # Format datetime
    if hasattr(val, 'strftime'):
        return val.strftime('%H:%M')
    # Format "HH:MM"
    if ':' in s and len(s) <= 5:
        return s
    return s


def time_to_minutes(t_str):
    """Convertit HH:MM en minutes depuis minuit."""
    if not t_str or t_str == '-':
        return None
    parts = t_str.split(':')
    return int(parts[0]) * 60 + int(parts[1])


def minutes_to_hhmm(mins):
    """Convertit des minutes en HH:MM."""
    if mins is None or mins < 0:
        return '00:00'
    h = int(mins) // 60
    m = int(mins) % 60
    return f"{h:02d}:{m:02d}"


def parse_enregistrement(filepath):
    """
    Parse le fichier Enregistrement des arrivées et départs.
    Retourne: dict[employee_id] = {
        'prenom': str, 'nom': str, 'service': str,
        'dates': dict[date_str] = {
            'sched_start': 'HH:MM', 'sched_end': 'HH:MM',
            'arrival': 'HH:MM' or None, 'departure': 'HH:MM' or None,
            'duration': 'HH:MM' or None
        }
    }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    
    # Trouver la ligne d'en-tête
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'prénom' in vals or 'prenom' in vals:
            header_row = i
            break
    
    if not header_row:
        return {}
    
    employees = {}
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        prenom = str(row[0]).strip() if row[0] else None
        nom = str(row[1]).strip() if row[1] else '-'
        emp_id = str(row[2]).strip() if row[2] else None
        service = str(row[3]).strip() if row[3] else ''
        date_val = str(row[4]).strip() if row[4] else None
        
        if not prenom or not emp_id or not date_val:
            continue
        
        # Horaire obligatoire
        sched_start = parse_time_str(row[6])  # Heure d'arrivée obligatoire
        sched_end = parse_time_str(row[8])     # Heure de départ obligatoire
        arrival = parse_time_str(row[9])       # Heure de contrôle d'arrivée
        departure = parse_time_str(row[10])    # Sortie à
        
        # Durée
        dur_raw = str(row[11]).strip() if row[11] else '00:00'
        dur_raw = dur_raw.replace(' : ', ':').replace(' :', ':').replace(': ', ':')
        
        if emp_id not in employees:
            employees[emp_id] = {
                'prenom': prenom,
                'nom': nom,
                'service': service,
                'dates': {},
                'schedules': []
            }
        
        # Normaliser date
        date_str = str(date_val)[:10]
        
        employees[emp_id]['dates'][date_str] = {
            'sched_start': sched_start,
            'sched_end': sched_end,
            'arrival': arrival if arrival != '-' else None,
            'departure': departure if departure != '-' else None,
            'duration': dur_raw
        }
        
        # Stocker le planning pour déduire plus tard
        if sched_start and sched_end:
            employees[emp_id]['schedules'].append((sched_start, sched_end))
    
    return employees


def parse_transactions(filepath):
    """
    Parse le fichier Transactions.
    Retourne: dict[employee_id] = {
        'prenom': str, 'nom': str, 'service': str,
        'dates': dict[date_str] = [list de 'HH:MM' triées]
    }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'prénom' in vals or 'prenom' in vals:
            header_row = i
            break
    
    if not header_row:
        return {}
    
    employees = {}
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        prenom = str(row[0]).strip() if row[0] else None
        nom = str(row[1]).strip() if row[1] else '-'
        emp_id = str(row[2]).strip() if row[2] else None
        service = str(row[3]).strip() if row[3] else ''
        date_val = str(row[4]).strip() if row[4] else None
        heure = parse_time_str(row[5])
        
        if not prenom or not emp_id or not date_val or not heure:
            continue
        
        if emp_id not in employees:
            employees[emp_id] = {
                'prenom': prenom,
                'nom': nom,
                'service': service,
                'dates': defaultdict(list)
            }
        
        date_str = str(date_val)[:10]
        employees[emp_id]['dates'][date_str].append(heure)
    
    # Trier les heures par date
    for emp_id in employees:
        for date_str in employees[emp_id]['dates']:
            employees[emp_id]['dates'][date_str].sort()
    
    return employees


def get_typical_schedule(enr_emp):
    """Détermine le planning typique d'un employé depuis l'Enregistrement."""
    if not enr_emp or not enr_emp.get('schedules'):
        return ('07:00', '17:00')  # Défaut
    
    # Prendre le planning le plus fréquent
    from collections import Counter
    counter = Counter(enr_emp['schedules'])
    return counter.most_common(1)[0][0]


def merge_files(enr_path, trans_path):
    """
    Fusionne Enregistrement + Transactions → données Présence.
    Retourne une liste de lignes au format Présence.
    """
    enr_data = parse_enregistrement(enr_path)
    trans_data = parse_transactions(trans_path)
    
    # Collecter tous les IDs d'employés
    all_ids = set(list(enr_data.keys()) + list(trans_data.keys()))
    
    rows = []
    
    for emp_id in sorted(all_ids):
        enr_emp = enr_data.get(emp_id, {})
        trans_emp = trans_data.get(emp_id, {})
        
        prenom = enr_emp.get('prenom') or trans_emp.get('prenom', '')
        nom = enr_emp.get('nom') or trans_emp.get('nom', '-')
        service = enr_emp.get('service') or trans_emp.get('service', '')
        
        # Plannings typiques de cet employé
        typical_start, typical_end = get_typical_schedule(enr_emp)
        is_night_typical = time_to_minutes(typical_start) > time_to_minutes(typical_end)
        
        # Dédupliquer les transactions
        trans_dates = {}
        if trans_emp and 'dates' in trans_emp:
            for d, times in trans_emp['dates'].items():
                trans_dates[d] = sorted(set(times))
        
        # Collecter toutes les dates
        all_dates = set()
        if enr_emp and 'dates' in enr_emp:
            all_dates.update(enr_emp['dates'].keys())
        all_dates.update(trans_dates.keys())
        
        for date_str in sorted(all_dates, reverse=True):
            enr_day = enr_emp.get('dates', {}).get(date_str, {}) if enr_emp else {}
            times = trans_dates.get(date_str, [])
            
            # --- Planning : toujours garder les heures EXACTES du fichier Enregistrement ---
            # NOUVEAU v55 : règle stricte → on prend exclusivement les valeurs du fichier
            # Enregistrement. Si la date n'a pas d'EDT, on cherche l'EDT du jour le plus
            # proche pour CET employé (au lieu d'un typical inventé). Si aucun EDT n'existe
            # pour cet employé, on laisse vide (cas extrêmement rare).
            if enr_day:
                sched_start = enr_day.get('sched_start') or ''
                sched_end = enr_day.get('sched_end') or ''
            else:
                # Jour absent → chercher l'EDT du jour le plus proche dans l'enregistrement
                sched_start = ''
                sched_end = ''
                if enr_emp and enr_emp.get('dates'):
                    enr_day_keys = sorted(enr_emp['dates'].keys())
                    # Trouver la date la plus proche
                    closest = None
                    min_diff = None
                    for k in enr_day_keys:
                        try:
                            from datetime import datetime as dt2
                            diff = abs((dt2.strptime(date_str, '%Y-%m-%d') - dt2.strptime(k, '%Y-%m-%d')).days)
                            if min_diff is None or diff < min_diff:
                                min_diff = diff
                                closest = k
                        except: continue
                    if closest:
                        clos_day = enr_emp['dates'][closest]
                        sched_start = clos_day.get('sched_start') or ''
                        sched_end = clos_day.get('sched_end') or ''
                # Si toujours rien, fallback typical
                if not sched_start:
                    sched_start = typical_start
                if not sched_end:
                    sched_end = typical_end
            
            ss_m = time_to_minutes(sched_start)
            se_m = time_to_minutes(sched_end)
            # Si pas d'horaire planifié, considéré comme jour normal (pas nuit)
            is_night = (ss_m is not None and se_m is not None) and ss_m > se_m
            
            # --- Arrivée & Départ depuis Transactions ---
            arrival = None
            departure = None
            
            if times:
                if is_night:
                    # Poste de nuit : arrivée = badge >= 14h, départ = badge < 14h
                    evening_badges = [t for t in times if time_to_minutes(t) >= 840]  # >= 14h
                    morning_badges = [t for t in times if time_to_minutes(t) < 840]   # < 14h
                    
                    arrival = evening_badges[0] if evening_badges else None
                    
                    # Départ = premier badge du lendemain matin
                    next_date = None
                    try:
                        from datetime import datetime as dt, timedelta
                        d = dt.strptime(date_str, '%Y-%m-%d')
                        next_date = (d + timedelta(days=1)).strftime('%Y-%m-%d')
                    except:
                        pass
                    
                    if next_date and next_date in trans_dates:
                        next_morning = [t for t in trans_dates[next_date] if time_to_minutes(t) < 840]
                        if next_morning:
                            departure = next_morning[0]
                    
                    # Fallback sur les badges matin du même jour
                    if not departure and morning_badges:
                        departure = morning_badges[0]
                    
                    # Si pas d'arrivée soir mais des badges matin → c'est le départ d'un poste précédent, ignorer
                    if not arrival and morning_badges:
                        continue  # Ce jour est juste le départ d'une nuit précédente
                        
                else:
                    # Poste de jour : premier badge = arrivée, dernier = départ
                    arrival = times[0]
                    departure = times[-1] if len(times) > 1 else times[0]
            
            # --- Fallback sur Enregistrement si Transactions incomplet ---
            if not arrival and enr_day:
                arr_val = enr_day.get('arrival')
                if arr_val and arr_val != '-':
                    arrival = arr_val
            
            if not departure and enr_day:
                dep_val = enr_day.get('departure')
                if dep_val and dep_val != '-':
                    departure = dep_val
            
            if not arrival:
                arrival = '-'
            if not departure:
                departure = '-'
            
            # --- Calculer la durée ---
            if arrival != '-' and departure != '-':
                arr_m = time_to_minutes(arrival)
                dep_m = time_to_minutes(departure)
                if arr_m is not None and dep_m is not None:
                    if is_night or dep_m < arr_m:
                        dur_m = (24 * 60 - arr_m) + dep_m
                    else:
                        dur_m = dep_m - arr_m
                    # Éviter les durées absurdes (> 18h)
                    if dur_m > 18 * 60:
                        dur_m = 0
                    duration = minutes_to_hhmm(dur_m)
                else:
                    duration = '00:00'
            else:
                duration = '00:00'
            
            rows.append([
                prenom, nom, emp_id, service, date_str,
                sched_start, sched_end, arrival, departure, duration
            ])
    
    return rows


def generate_presence_xlsx(enr_path, trans_path, output_path):
    """Génère le fichier Présence .xlsx à partir des 2 fichiers source."""
    rows = merge_files(enr_path, trans_path)
    
    if not rows:
        return None
    
    # Détecter le nom du service/client
    services = set()
    for r in rows:
        if r[3]:
            parts = r[3].split('>')
            if len(parts) >= 2:
                services.add(parts[1].strip())
    client_name = list(services)[0] if services else "CLIENT"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Présence"
    
    # Titre
    ws.append([f"Présence - {client_name}"])
    
    # En-têtes
    ws.append([
        'Prénom', 'Nom de famille', 'ID', 'Service', 'Date',
        "Heure d'arrivée obligatoire", 'Heure de départ obligatoire',
        "Heure de contrôle d'arrivée", 'Sortie à', 'Durée'
    ])
    
    # Données
    for row in rows:
        ws.append(row)
    
    # Ajuster largeurs
    col_widths = [15, 18, 10, 45, 12, 12, 12, 12, 10, 10]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w
    
    wb.save(output_path)
    
    emp_count = len(set(r[2] for r in rows))
    emp_names = sorted(set(f"{r[0]} {r[1]}".strip() for r in rows if r[0]))
    # Group by service for frontend
    emp_by_service = {}
    for r in rows:
        name = f"{r[0]} {r[1]}".strip() if r[0] else None
        service = r[3] if len(r) > 3 else ''
        if name and name not in emp_by_service:
            emp_by_service[name] = service or 'Non défini'
    return {
        'path': output_path,
        'client': client_name,
        'employees': emp_names,
        'emp_services': emp_by_service,
        'emp_count': emp_count,
        'rows': len(rows)
    }


# ======================== TEST ========================
if __name__ == '__main__':
    result = generate_presence_xlsx(
        '/mnt/user-data/uploads/Enregistrement_des_arrivées_et_départs_2026-02-01_2026-02-28.xlsx',
        '/mnt/user-data/uploads/Transactions_2026-02-01_2026-02-28.xlsx',
        '/home/claude/test_presence.xlsx'
    )
    print(f"Résultat: {result}")
    
    # Vérifier
    wb = openpyxl.load_workbook('/home/claude/test_presence.xlsx')
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        print(f"  Row {i+1}: {list(row)}")
    print(f"  ... Total: {ws.max_row} lignes")
#!/usr/bin/env python3
"""Module DPCI — Calcul d'heures avec 4 temps (arrivée, pause début, pause fin, départ)."""

import os
from datetime import datetime
from collections import OrderedDict
import openpyxl

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, black
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

TEAL = HexColor('#1a7a6d')
DARK_TEAL = HexColor('#0d5b50')
ORANGE = HexColor('#e8672a')
RED = HexColor('#c53030')
BLUE = HexColor('#1565c0')
LGRAY = HexColor('#f5f6fa')


def t2m(t):
    """Convertit HH:MM en minutes."""
    if not t or t in ('', '-', '00:00'):
        return 0
    try:
        parts = t.replace('h', ':').split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return 0


def m2h(mins):
    """Convertit minutes en HH:MM."""
    if not mins or mins <= 0:
        return "00:00"
    mins = int(mins)
    return f"{mins // 60:02d}:{mins % 60:02d}"


def parse_dpci_excel(xlsx_path):
    """Parse le fichier Excel DPCI et retourne les employés groupés par département."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Find header row
    header_row = None
    time_period = ""
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        vals = [str(v or '') for v in row]
        if 'Time Period' in vals[0]:
            time_period = vals[0].replace('Time Period: ', '')
        if vals[0] == 'First Name' or 'First' in vals[0]:
            header_row = True
            break

    employees = OrderedDict()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not row[0] or not row[2]:
            continue
        first = str(row[0]).strip()
        last = str(row[1] or '').strip()
        emp_id = str(row[2]).strip()
        dept = str(row[3] or '').strip()
        date_val = str(row[4] or '').strip()
        record = str(row[5] or '').strip()

        if first == 'First Name' or emp_id == 'ID':
            continue

        # Parse date
        if len(date_val) < 8:
            continue

        # Parse 4 times from record
        times = record.split(';')
        arrival = times[0].strip() if len(times) > 0 else ''
        pause_start = times[1].strip() if len(times) > 1 else ''
        pause_end = times[2].strip() if len(times) > 2 else ''
        departure = times[3].strip() if len(times) > 3 else ''

        full_name = f"{first} {last}".strip()
        key = emp_id

        if key not in employees:
            employees[key] = {
                'name': full_name,
                'id': emp_id,
                'department': dept,
                'records': []
            }

        employees[key]['records'].append({
            'date': date_val[:10],
            'arrival': arrival[:5] if len(arrival) >= 5 else arrival,
            'pause_start': pause_start[:5] if len(pause_start) >= 5 else pause_start,
            'pause_end': pause_end[:5] if len(pause_end) >= 5 else pause_end,
            'departure': departure[:5] if len(departure) >= 5 else departure,
        })

    wb.close()

    # Sort records by date
    for emp in employees.values():
        emp['records'].sort(key=lambda x: x['date'])

    return list(employees.values()), time_period


def calc_dpci_stats(emp, schedule=None, hourly_cost=0, hp=0, hp_weekend=0, schedules_per_day=None):
    """Calcule les stats pour un employé DPCI. hp/hp_weekend en heures.
    
    v64 : schedules_per_day = {'lundi': {start_time, end_time, break_start, break_end}, ...}
    Si fourni, le schedule est choisi en fonction du jour de la semaine du record.
    """
    records = emp['records']
    total_worked = 0
    total_pause = 0
    total_late = 0
    total_overtime = 0
    total_required = 0
    days_present = 0
    days_late = 0
    days_absent = 0

    # Default schedule from DB or fallback
    # v71 : pause optionnelle — si break_start/break_end est None, pas de pause appliquée
    default_sched_start = t2m(schedule.get('start_time', '07:00')) if schedule else t2m('07:00')
    default_sched_end = t2m(schedule.get('end_time', '17:00')) if schedule else t2m('17:00')
    _def_bs = schedule.get('break_start') if schedule else None
    _def_be = schedule.get('break_end') if schedule else None
    default_sched_break_start = t2m(_def_bs) if _def_bs else 0
    default_sched_break_end = t2m(_def_be) if _def_be else 0
    default_has_pause = bool(_def_bs and _def_be)
    
    hm = hp * 60  # heures obligatoires semaine en minutes
    hm_we = hp_weekend * 60

    enriched = []
    _DAY_NAMES = ['lundi','mardi','mercredi','jeudi','vendredi','samedi','dimanche']

    for rec in records:
        arr = t2m(rec['arrival'])
        ps = t2m(rec['pause_start'])
        pe = t2m(rec['pause_end'])
        dep = t2m(rec['departure'])

        # Detect weekend + day name
        is_weekend = False
        day_name = None
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(rec['date'][:10], '%Y-%m-%d')
            is_weekend = d.weekday() >= 5
            day_name = _DAY_NAMES[d.weekday()]
        except:
            pass
        
        # v64 : choisir le schedule du jour si disponible
        # v71 : pause optionnelle (None = pas de pause)
        if schedules_per_day and day_name and schedules_per_day.get(day_name):
            day_s = schedules_per_day[day_name]
            sched_start = t2m(day_s.get('start_time') or day_s.get('start') or '07:00')
            sched_end = t2m(day_s.get('end_time') or day_s.get('end') or '17:00')
            _bs = day_s.get('break_start') or day_s.get('pause_start')
            _be = day_s.get('break_end') or day_s.get('pause_end')
            sched_break_start = t2m(_bs) if _bs else 0
            sched_break_end = t2m(_be) if _be else 0
            has_pause = bool(_bs and _be)
        else:
            sched_start = default_sched_start
            sched_end = default_sched_end
            sched_break_start = default_sched_break_start
            sched_break_end = default_sched_break_end
            has_pause = default_has_pause

        # Determine required hours for this day
        # v71 : ne soustrait la pause QUE si elle est définie
        if is_weekend and hp_weekend > 0:
            required = hm_we
        elif not is_weekend and hp > 0:
            required = hm
        else:
            base = sched_end - sched_start
            if has_pause:
                base -= (sched_break_end - sched_break_start)
            required = base
        
        total_required += required

        if arr == 0 and dep == 0:
            days_absent += 1
            enriched.append({
                'date': rec['date'],
                'arrival': '-', 'pause_start': '-', 'pause_end': '-', 'departure': '-',
                'worked': '00:00', 'pause': '00:00', 'presence': '00:00',
                'required': m2h(required), 'state': 'Absent', 'respect': 'ABS',
                # v76 : propager l'EDT d'origine pour l'affichage
                'sched_start_original': rec.get('sched_start_original'),
                'sched_end_original': rec.get('sched_end_original'),
            })
            continue

        days_present += 1

        # Pause duration
        pause = pe - ps if pe > ps else 0
        total_pause += pause

        # Worked = morning + afternoon
        morning = ps - arr if ps > arr else 0
        afternoon = dep - pe if dep > pe else 0
        worked = morning + afternoon
        total_worked += worked

        # Presence (total on site)
        presence = dep - arr if dep > arr else 0

        # Late (tracked internally for cost but not displayed)
        late = arr - sched_start if arr > sched_start else 0
        if late > 0:
            total_late += late
            days_late += 1

        # Overtime
        overtime = dep - sched_end if dep > sched_end else 0
        total_overtime += overtime

        # Respect hours
        if worked >= required - 5:
            respect = "OUI"
        else:
            respect = "NON"

        enriched.append({
            'date': rec['date'],
            'arrival': rec['arrival'],
            'pause_start': rec['pause_start'],
            'pause_end': rec['pause_end'],
            'departure': rec['departure'],
            'worked': m2h(worked),
            'pause': m2h(pause),
            'presence': m2h(presence),
            'required': m2h(required),
            'state': 'Présent',
            'respect': respect,
            # v76 : propager l'EDT d'origine pour l'affichage
            'sched_start_original': rec.get('sched_start_original'),
            'sched_end_original': rec.get('sched_end_original'),
        })

    presence_rate = (days_present / len(records) * 100) if len(records) > 0 else 0

    stats = {
        'days_required': len(records),
        'days_present': days_present,
        'days_late': days_late,
        'days_punctual': days_present - days_late,
        'days_absent': days_absent,
        'total_required': total_required,
        'total_worked': total_worked,
        'total_pause': total_pause,
        'total_late': total_late,
        'total_overtime': total_overtime,
        'presence_rate': round(presence_rate, 1),
        'sched_str': f"{m2h(sched_start)}-{m2h(sched_end)}",
        'hourly_cost': hourly_cost,
        'cost_late': round(total_late / 60 * hourly_cost) if hourly_cost > 0 else 0,
        'cost_absent': round(days_absent * required / 60 * hourly_cost) if hourly_cost > 0 and len(records) > 0 else 0,
    }
    return enriched, stats


def generate_dpci_pdf(emps, output_path, client_name, period, schedules_map=None, employee_costs=None, default_cost=0, hp=0, hp_weekend=0, provider_name='', treated_by='', period_mode='all', rest_days=None, schedules_per_day_map=None, show_pause=True):
    """Génère le rapport PDF DPCI — design identique à la fiche de présence.
    
    v64 : schedules_per_day_map = {nom_employé: {jour: {start_time, end_time, ...}}} (par jour)
    Si fourni, l'EDT est choisi en fonction du jour de la semaine de chaque record.
    """
    if not schedules_map:
        schedules_map = {}
    if not employee_costs:
        employee_costs = {}
    if rest_days is None:
        rest_days = []
    if schedules_per_day_map is None:
        schedules_per_day_map = {}

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=12 * mm, rightMargin=12 * mm, topMargin=10 * mm, bottomMargin=10 * mm)

    # Couleurs exactes de l'image
    HEADER_BG = HexColor('#44546A')   # Barre en-tête gris-bleu foncé
    BLUE_HDR  = HexColor('#4472C4')   # En-têtes tableaux résumé + détail
    BLUE_DARK = HexColor('#305496')   # Sous-en-tête résumé 2
    BORDER_BL = HexColor('#8EAADB')   # Bordures bleu clair
    WHITE     = white
    BLK       = HexColor('#333333')
    LGREY     = HexColor('#F2F2F2')

    hw  = ParagraphStyle('hw', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE, alignment=TA_CENTER)
    hv  = ParagraphStyle('hv', fontSize=9, alignment=TA_CENTER, textColor=BLK)
    th  = ParagraphStyle('th', fontName='Helvetica-Bold', fontSize=7.5, textColor=WHITE, alignment=TA_CENTER, leading=9)
    tc  = ParagraphStyle('tc', fontSize=8, alignment=TA_CENTER, textColor=BLK, leading=10)
    tcb = ParagraphStyle('tcb', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER, textColor=BLK, leading=10)
    ft_s = ParagraphStyle('ft', fontSize=7, textColor=HexColor('#888'), alignment=TA_LEFT)

    story = []
    period_labels = {'all': 'mois', 'week': 'semaine', 'day': 'jour', 'custom': 'période'}
    period_label = period_labels.get(period_mode, 'période')
    now = datetime.now().strftime("%d/%m/%Y \u00e0 %H:%M")

    depts = OrderedDict()
    for emp in emps:
        dept = emp.get('department', 'Non assign\u00e9')
        if dept not in depts:
            depts[dept] = []
        depts[dept].append(emp)

    first_page = True
    pw = 186 * mm
    total_emps = sum(len(v) for v in depts.values())
    emp_counter = 0

    for dept_name, dept_emps in depts.items():
        for emp in dept_emps:
            emp_counter += 1
            if not first_page:
                story.append(PageBreak())
            first_page = False

            sched = schedules_map.get(emp['name'], None)
            cost = employee_costs.get(emp['name'], default_cost)
            sched_per_day = schedules_per_day_map.get(emp['name'])  # v64
            enriched, stats = calc_dpci_stats(emp, schedule=sched, hourly_cost=cost, hp=hp, hp_weekend=hp_weekend, schedules_per_day=sched_per_day)

            # BARRE EN-TETE
            prov = provider_name or 'RAMYA TECHNOLOGIE & INNOVATION'
            hbar = Table([[
                Paragraph(f"<b>{prov}</b>", ParagraphStyle('hl', fontName='Helvetica-Bold', fontSize=10, textColor=WHITE)),
                Paragraph(f"<b>{client_name}</b>", ParagraphStyle('hr', fontName='Helvetica-Bold', fontSize=10, textColor=WHITE, alignment=TA_RIGHT)),
            ]], colWidths=[pw * 0.55, pw * 0.45])
            hbar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 10), ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('LEFTPADDING', (0, 0), (-1, -1), 14), ('RIGHTPADDING', (0, 0), (-1, -1), 14),
            ]))
            story.extend([hbar, Spacer(1, 8 * mm)])

            # TITRE + PERIODE (s\u00e9par\u00e9s)
            story.append(Paragraph("<b>RAPPORT INDIVIDUEL</b>",
                ParagraphStyle('tit', fontName='Helvetica-Bold', fontSize=18, textColor=BLK, alignment=TA_CENTER, spaceAfter=3*mm)))
            story.append(Paragraph(period,
                ParagraphStyle('sub', fontSize=9, textColor=HexColor('#666'), alignment=TA_CENTER, spaceBefore=1*mm)))
            story.append(Spacer(1, 5 * mm))

            # EMPLOYE
            story.append(Paragraph(f"<b>Employ\u00e9 : {emp['name']}</b>",
                ParagraphStyle('emp', fontName='Helvetica-Bold', fontSize=11, textColor=BLK)))
            story.append(Paragraph(f"R\u00e9f\u00e9rence : {emp['id']}",
                ParagraphStyle('ref', fontSize=9, textColor=HexColor('#555'))))
            story.append(Spacer(1, 4 * mm))

            # RESUME 1 : JOURS
            s1_h = ["Nbre de jours \u00e0 Effectuer", "Ponctualité", "Absence"]
            s1_v = [f"{stats['days_required']} jours", f"{stats['days_punctual']} jours", f"{stats['days_absent']} jours"]
            cw1 = [pw * 0.40, pw * 0.30, pw * 0.30]
            t1 = Table([
                [Paragraph(x, hw) for x in s1_h],
                [Paragraph(x, hv) for x in s1_v],
            ], colWidths=cw1)
            t1.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_HDR),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.4, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.extend([t1, Spacer(1, 2 * mm)])

            # RESUME 2 : HEURES
            s2_h = ["Total heure obligatoire", "Pr\u00e9sence", "Absence"]
            abs_hrs = m2h(stats['days_absent'] * (stats['total_required'] // max(stats['days_required'], 1)))
            s2_v = [f"{m2h(stats['total_required'])} heures", f"{m2h(stats['total_worked'])} heures", f"{abs_hrs} heures"]
            t2 = Table([
                [Paragraph(x, hw) for x in s2_h],
                [Paragraph(x, hv) for x in s2_v],
            ], colWidths=cw1)
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_DARK),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.4, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.extend([t2, Spacer(1, 4 * mm)])

            # TABLEAU DETAIL
            # v73 : colonnes pause masquées si show_pause=False
            if show_pause:
                hdrs = ["Jour", "Date", "Emploi du\ntemps", "Heure\nd'arriv\u00e9e",
                        "D\u00e9but de\npause", "Retour de\npause", "Heure de\nd\u00e9part", "Pause\neffectu\u00e9e",
                        "H.\nobligatoire", "H.\ntravaill\u00e9es", "Emploi du temps\nrespect\u00e9"]
                cw_d = [9*mm, 18*mm, 20*mm, 16*mm, 15*mm, 15*mm, 16*mm, 15*mm, 16*mm, 16*mm, 18*mm]
            else:
                hdrs = ["Jour", "Date", "Emploi du\ntemps", "Heure\nd'arriv\u00e9e",
                        "Heure de\nd\u00e9part",
                        "H.\nobligatoire", "H.\ntravaill\u00e9es", "Emploi du temps\nrespect\u00e9"]
                cw_d = [10*mm, 22*mm, 26*mm, 22*mm, 22*mm, 22*mm, 22*mm, 26*mm]

            td = [[Paragraph(x.replace("\n", "<br/>"), th) for x in hdrs]]

            total_pause_mins = 0

            for i, rec in enumerate(enriched, 1):
                # v76 : afficher l'EDT d'origine du fichier si préservé, sinon le sched_str global
                _orig_s = rec.get('sched_start_original')
                _orig_e = rec.get('sched_end_original')
                if _orig_s and _orig_e:
                    sched_str = f"{_orig_s}-{_orig_e}"
                else:
                    sched_str = stats['sched_str']
                resp = rec['respect']
                if resp == 'OUI':
                    rp = Paragraph("OUI", ParagraphStyle('g', fontName='Helvetica-Bold', fontSize=7, textColor=HexColor('#2e7d32'), alignment=TA_CENTER))
                elif resp == 'ABS':
                    rp = Paragraph("ABS", ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=7, textColor=HexColor('#c53030'), alignment=TA_CENTER))
                else:
                    rp = Paragraph("NON", ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=7, textColor=HexColor('#c53030'), alignment=TA_CENTER))

                req_display = rec.get('required', '') or m2h(stats['total_required'] // max(stats['days_required'], 1))

                # Track pause total
                total_pause_mins += t2m(rec.get('pause', '00:00'))

                if show_pause:
                    td.append([
                        Paragraph(str(i), tc),
                        Paragraph(rec['date'], tc),
                        Paragraph(f"({sched_str.replace('-', '_')})", tc),
                        Paragraph(rec['arrival'] if rec['arrival'] != '-' else '-', tcb),
                        Paragraph(rec['pause_start'] if rec['pause_start'] != '-' else '-', tc),
                        Paragraph(rec['pause_end'] if rec['pause_end'] != '-' else '-', tc),
                        Paragraph(rec['departure'] if rec['departure'] != '-' else '-', tcb),
                        Paragraph(rec.get('pause', '00:00'), tc),
                        Paragraph(req_display, tc),
                        Paragraph(rec['worked'], tcb),
                        rp,
                    ])
                else:
                    td.append([
                        Paragraph(str(i), tc),
                        Paragraph(rec['date'], tc),
                        Paragraph(f"({sched_str.replace('-', '_')})", tc),
                        Paragraph(rec['arrival'] if rec['arrival'] != '-' else '-', tcb),
                        Paragraph(rec['departure'] if rec['departure'] != '-' else '-', tcb),
                        Paragraph(req_display, tc),
                        Paragraph(rec['worked'], tcb),
                        rp,
                    ])

            dt = Table(td, colWidths=cw_d, repeatRows=1)
            sc = [
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_HDR),
                ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.3, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 1),
                ('RIGHTPADDING', (0, 0), (-1, -1), 1),
            ]
            for i in range(2, len(td), 2):
                sc.append(('BACKGROUND', (0, i), (-1, i), LGREY))
            dt.setStyle(TableStyle(sc))
            story.append(dt)

            # RÉSUMÉ CUMULS EN BAS
            story.append(Spacer(1, 3 * mm))
            cum_h = [f"Cumul pause ({period_label})", f"Cumul H. travaill\u00e9es ({period_label})", f"Cumul H. obligatoire ({period_label})", "Taux pr\u00e9sence"]
            cum_v = [f"{m2h(total_pause_mins)}", f"{m2h(stats['total_worked'])}", f"{m2h(stats['total_required'])}", f"{stats['presence_rate']}%"]
            ct_cum = Table([
                [Paragraph(x, hw) for x in cum_h],
                [Paragraph(x, ParagraphStyle('cv', fontName='Helvetica-Bold', fontSize=9, alignment=TA_CENTER, textColor=BLK)) for x in cum_v],
            ], colWidths=[pw * 0.25] * 4)
            ct_cum.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_DARK),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.4, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(ct_cum)

            # IMPACT FINANCIER (absences uniquement)
            if cost > 0 and stats['cost_absent'] > 0:
                story.append(Spacer(1, 3 * mm))
                fmt = lambda x: f"{x:,.0f} FCFA"
                cd = [
                    [Paragraph("<b>IMPACT FINANCIER</b>", ParagraphStyle('x', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE)),
                     Paragraph(f"<b>Co\u00fbt : {fmt(cost)}/h</b>", ParagraphStyle('x2', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE, alignment=TA_RIGHT))],
                    [Paragraph(f"Perte absences ({stats['days_absent']} jour(s))", ParagraphStyle('x3', fontSize=8, textColor=BLK)),
                     Paragraph(f"<b>{fmt(stats['cost_absent'])}</b>", ParagraphStyle('x4', fontSize=9, fontName='Helvetica-Bold', textColor=HexColor('#c53030'), alignment=TA_RIGHT))],
                    [Paragraph("<b>TOTAL GAIN PERDU</b>", ParagraphStyle('x5', fontName='Helvetica-Bold', fontSize=9, textColor=HexColor('#c53030'))),
                     Paragraph(f"<b>{fmt(stats['cost_absent'])}</b>", ParagraphStyle('x6', fontName='Helvetica-Bold', fontSize=10, textColor=HexColor('#c53030'), alignment=TA_RIGHT))],
                ]
                ct = Table(cd, colWidths=[pw * 0.65, pw * 0.35])
                ct.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
                    ('BACKGROUND', (0, -1), (-1, -1), HexColor('#FFF2CC')),
                    ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                    ('INNERGRID', (0, 0), (-1, -1), 0.3, BORDER_BL),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ]))
                story.append(ct)

            # FOOTER
            story.append(Spacer(1, 6 * mm))
            story.append(Paragraph(f"Généré le {now} | {client_name} - Rapport {emp['name']} {emp_counter}/{total_emps}  —  Traité par : {treated_by or 'Admin'}", ft_s))

    doc.build(story)
#!/usr/bin/env python3
"""Générateur de Devis/Proforma PDF — Format RAMYA TECHNOLOGIE"""

import os, json
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

TEAL = HexColor('#1a7a6d')
ORANGE = HexColor('#e8672a')


def number_to_words_fr(n):
    """Convertit un nombre en mots français (simplifié)."""
    units = ['', 'Un', 'Deux', 'Trois', 'Quatre', 'Cinq', 'Six', 'Sept', 'Huit', 'Neuf',
             'Dix', 'Onze', 'Douze', 'Treize', 'Quatorze', 'Quinze', 'Seize', 'Dix-sept',
             'Dix-huit', 'Dix-neuf']
    tens = ['', '', 'Vingt', 'Trente', 'Quarante', 'Cinquante', 'Soixante',
            'Soixante', 'Quatre-vingt', 'Quatre-vingt']
    
    if n == 0: return 'Zéro'
    if n < 0: return 'Moins ' + number_to_words_fr(-n)
    
    result = ''
    if n >= 1000000:
        m = n // 1000000
        result += ('Un Million' if m == 1 else number_to_words_fr(m) + ' Millions') + ' '
        n %= 1000000
    if n >= 1000:
        t = n // 1000
        result += ('Mille' if t == 1 else number_to_words_fr(t) + ' Mille') + ' '
        n %= 1000
    if n >= 100:
        c = n // 100
        result += ('Cent' if c == 1 else units[c] + ' Cent') + ' '
        n %= 100
    if n >= 20:
        d = n // 10
        if d == 7 or d == 9:
            result += tens[d] + '-' + units[10 + n % 10] + ' '
            n = 0
        else:
            result += tens[d]
            if n % 10 == 1 and d != 8:
                result += ' et Un '
            elif n % 10 > 0:
                result += '-' + units[n % 10] + ' '
            else:
                result += ' '
            n = 0
    if 0 < n < 20:
        result += units[n] + ' '
    
    return result.strip()


def fmt(amount):
    """Format number with thousands separator."""
    return f"{amount:,.0f}".replace(',', ' ')


def generate_devis_pdf(devis_data, output_path, logo_path=None, doc_params=None):
    """Génère un PDF de devis/proforma/facture au format RAMYA exact (cf. modèle).
    v162 : doc_params (entête/pied personnalisables via Admin → Paramètres documents)."""
    
    # ==== COULEURS exactes du modèle ====
    RAMYA_ORANGE = HexColor('#F29F2F')    # Header table + barre résumé
    RAMYA_GREEN = HexColor('#2FA85B')     # Barre Total TTC
    RAMYA_TEAL = HexColor('#1a7a6d')      # Texte RAMYA titre + footer
    ROW_BG = HexColor('#f8f8f8')          # fond alterné de ligne
    HEADER_GRAY = HexColor('#4B4B4B')     # texte DEVIS/PROFORMA/FACTURE
    
    def _draw_corners(canv, doc_):
        """Coins neutres — plus de triangles décoratifs (retirés à la demande du client)."""
        # Ligne pointillée séparatrice au-dessus du footer
        w, h = A4
        canv.saveState()
        canv.setStrokeColor(HexColor('#bbbbbb'))
        canv.setDash(2, 2)
        canv.line(15*mm, 22*mm, w-15*mm, 22*mm)
        canv.setDash()
        canv.restoreState()
    
    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=8*mm, bottomMargin=20*mm)
    
    story = []
    
    # Styles
    s_title = ParagraphStyle('title', fontSize=26, fontName='Helvetica-Bold',
                              alignment=TA_RIGHT, textColor=HEADER_GRAY, leading=28)
    s_ref_small = ParagraphStyle('ref_small', fontSize=9, alignment=TA_RIGHT,
                                  textColor=HexColor('#333'), leading=12)
    s_normal = ParagraphStyle('normal', fontSize=10, leading=13)
    s_bold = ParagraphStyle('bold', fontSize=10, fontName='Helvetica-Bold')
    s_small = ParagraphStyle('small', fontSize=8, textColor=HexColor('#888'))
    s_center = ParagraphStyle('center', fontSize=9, alignment=TA_CENTER)
    s_right = ParagraphStyle('right', fontSize=10, alignment=TA_RIGHT)
    
    # Document type
    raw_type = (devis_data.get('doc_type') or 'devis').lower()
    if raw_type in ('facture','invoice'):
        doc_type = 'FACTURE'; prefix = 'FAC'
    elif raw_type == 'proforma':
        doc_type = 'PROFORMA'; prefix = 'PRO'
    else:
        doc_type = 'DEVIS'; prefix = 'DEV'
    # v162 : sur une FACTURE, entête de tableau TEAL (modèle facture) au lieu de l'orange (devis)
    header_color = RAMYA_TEAL if doc_type == 'FACTURE' else RAMYA_ORANGE
    
    ref = devis_data.get('reference', '')
    # Date: preserve format if given else format from created_at
    date_str = devis_data.get('date') or devis_data.get('created_at','')[:10]
    if date_str and '-' in date_str and len(date_str) >= 10:
        # Convert YYYY-MM-DD to DD-MM-YYYY
        try:
            y,m,d = date_str[:10].split('-'); date_str = f"{d}-{m}-{y}"
        except: pass
    
    contact = devis_data.get('contact_commercial', '') or devis_data.get('commercial','')
    client_name = devis_data.get('client_name', '')
    client_code = devis_data.get('client_code', '')
    objet = devis_data.get('objet', '')
    items = devis_data.get('items_json') or devis_data.get('items') or '[]'
    if isinstance(items, str):
        try: items = json.loads(items)
        except: items = []
    
    total_ht = float(devis_data.get('total_ht', 0) or 0)
    petites_fourn = float(devis_data.get('petites_fournitures', 0) or 0)
    total_ttc = float(devis_data.get('total_ttc', 0) or 0)
    main_oeuvre = float(devis_data.get('main_oeuvre', 0) or 0)
    remise = float(devis_data.get('remise', 0) or 0)
    # TVA optionnelle
    tva_active = bool(devis_data.get('tva_active', False))
    tva_rate = float(devis_data.get('tva_rate', 18) or 18)  # 18% par défaut en CI
    tva_amount = float(devis_data.get('tva_amount', 0) or 0)
    # Si TVA active mais montant non calculé, on le calcule
    if tva_active and not tva_amount and total_ht:
        tva_amount = round(total_ht * tva_rate / 100, 0)
    
    # =========================================================
    # HEADER : Logo + nom société + liste services (comme modèle)
    # =========================================================
    # Logo with proper aspect ratio — logo is 1536x2008 (ratio ~0.76)
    logo_el = Paragraph("", ParagraphStyle('empty', fontSize=1))
    if logo_path and os.path.exists(logo_path):
        try:
            # Use aspect ratio from actual image to avoid distortion
            try:
                from PIL import Image as _PIL
                with _PIL.open(logo_path) as _im:
                    ratio = _im.width / _im.height
            except:
                ratio = 0.76  # fallback
            logo_w = 20*mm
            logo_h = logo_w / ratio
            logo_el = RLImage(logo_path, width=logo_w, height=logo_h)
        except: pass
    
    company_name = Paragraph(
        "<b>RAMYA</b><br/><b>TECHNOLOGIE &amp; INNOVATION</b>",
        ParagraphStyle('co', fontSize=10, fontName='Helvetica-Bold',
                       textColor=RAMYA_TEAL, leading=13, alignment=TA_CENTER)
    )
    
    # Services : texte puis puce (■) — tout aligné à droite.
    # v162 : sur une FACTURE, on n'affiche PAS la liste marketing (éléments inutiles en haut).
    if doc_type == 'FACTURE':
        services = Paragraph('', ParagraphStyle('svc', fontSize=8))
    else:
        services_html = (
            '<b>Caméras de surveillance</b> <font color="#F29F2F"><b>■</b></font><br/>'
            '<b>Clôture électrique</b> <font color="#F29F2F"><b>■</b></font><br/>'
            '<b>Kit visiophone alarme anti-intrusion</b> <font color="#F29F2F"><b>■</b></font><br/>'
            '<b>Domotique, Poignées intelligentes</b> <font color="#F29F2F"><b>■</b></font>'
        )
        services = Paragraph(services_html, ParagraphStyle('svc', fontSize=8, leading=11,
                              textColor=HexColor('#333'), alignment=TA_RIGHT))
    
    header_data = [[logo_el, company_name, services]]
    ht = Table(header_data, colWidths=[28*mm, 65*mm, 87*mm])
    ht.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        # White background behind logo to make it visible
        ('BACKGROUND', (0,0), (0,0), white),
        ('BOX', (0,0), (0,0), 0, white),
        ('TOPPADDING', (0,0), (0,0), 2),
        ('BOTTOMPADDING', (0,0), (0,0), 2),
        ('LEFTPADDING', (0,0), (0,0), 2),
        ('RIGHTPADDING', (0,0), (0,0), 2),
    ]))
    story.append(ht)
    story.append(Spacer(1, 4*mm))
    
    # =========================================================
    # TITRE DEVIS + référence à droite — taille réduite
    # =========================================================
    right_info_parts = [
        f"<font size='16' color='#4B4B4B'><b>{doc_type}</b></font>",
        f"<font size='8'># {ref}</font>",
        f"<font size='8'>Date: {date_str}</font>"
    ]
    # On retire "Contact commercial" — on garde seulement "Établi par : X le ..."
    redacteur = devis_data.get('redacteur', '') or contact or ''
    redacteur_date = devis_data.get('redacteur_date', '')
    if redacteur:
        rparts = [f"<i>Établi par :</i> <b>{redacteur}</b>"]
        if redacteur_date: rparts.append(f"<i>le</i> {redacteur_date}")
        right_info_parts.append(f"<font size='8' color='#888'>{' '.join(rparts)}</font>")
    
    right_info = Paragraph("<br/>".join(right_info_parts),
        ParagraphStyle('right_info', alignment=TA_RIGHT, leading=12))
    title_data = [['', right_info]]
    tt = Table(title_data, colWidths=[90*mm, 90*mm])
    tt.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),
                             ('LEFTPADDING',(0,0),(-1,-1),0),
                             ('RIGHTPADDING',(0,0),(-1,-1),0)]))
    story.append(tt)
    story.append(Spacer(1, 3*mm))
    
    # =========================================================
    # CLIENT (À gauche)
    # =========================================================
    story.append(Paragraph("<b>À</b>", ParagraphStyle('a', fontSize=10, fontName='Helvetica-Bold')))
    story.append(Paragraph(f"<b>{client_name}</b>",
        ParagraphStyle('cl', fontSize=11, fontName='Helvetica-Bold')))
    story.append(Spacer(1, 2*mm))
    if client_code:
        story.append(Paragraph(f"Code client: {client_code}",
            ParagraphStyle('cc', fontSize=10, textColor=HexColor('#333'))))
    story.append(Spacer(1, 2*mm))
    
    if objet:
        story.append(Paragraph(f"<b>Objet :</b> {objet}",
            ParagraphStyle('obj', fontSize=10, fontName='Helvetica-Bold')))
    story.append(Spacer(1, 3*mm))
    
    # =========================================================
    # TABLEAU DES ARTICLES — header ORANGE
    # =========================================================
    hdrs = ['#', 'Désignation', 'Qté.', 'Prix unitaire', 'Remise', 'Montant HT']
    th_style = ParagraphStyle('th', fontSize=9, fontName='Helvetica-Bold', textColor=white, alignment=TA_LEFT)
    th_style_c = ParagraphStyle('thc', fontSize=9, fontName='Helvetica-Bold', textColor=white, alignment=TA_CENTER)
    th_style_r = ParagraphStyle('thr', fontSize=9, fontName='Helvetica-Bold', textColor=white, alignment=TA_RIGHT)
    
    table_data = [[
        Paragraph('#', th_style_c),
        Paragraph('Désignation', th_style),
        Paragraph('Qté.', th_style_c),
        Paragraph('Prix unitaire', th_style_r),
        Paragraph('Remise', th_style_r),
        Paragraph('Montant HT', th_style_r),
    ]]
    
    for idx, item in enumerate(items, 1):
        desc = str(item.get('designation', ''))
        detail = str(item.get('detail', ''))
        full_desc = f"<b>{desc}</b>"
        if detail: full_desc += f"<br/>{detail}"
        
        qty = int(item.get('qty', 1) or 1)
        prix = float(item.get('prix', 0) or 0)
        rem = float(item.get('remise', 0) or 0)
        montant = qty * prix - rem
        
        table_data.append([
            Paragraph(str(item.get('num', idx)), s_center),
            Paragraph(full_desc, ParagraphStyle('desc', fontSize=9, leading=12)),
            Paragraph(str(qty), s_center),
            Paragraph(fmt(prix), s_right),
            Paragraph(fmt(rem) if rem else '', s_right),
            Paragraph(fmt(montant), s_right),
        ])
    
    col_widths = [10*mm, 82*mm, 14*mm, 26*mm, 20*mm, 28*mm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('LINEBELOW', (0, 0), (-1, 0), 1, header_color),
        ('LINEABOVE', (0, 1), (-1, -1), 0.25, HexColor('#eeeeee')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, 0), 4),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, ROW_BG]),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 1*mm))
    
    # =========================================================
    # TOTAUX : Total HT, petites fournitures, puis Total TTC en VERT
    # =========================================================
    # Totaux lignes (conditionnelles : on n'affiche que les lignes non nulles)
    totaux_rows = [
        ['', Paragraph("<b>Total HT (pièces)</b>", s_right),
         Paragraph(f"<b>{fmt(total_ht)}XOF</b>", s_right)],
    ]
    # v160 : main d'œuvre + remise affichées pour que le détail réconcilie le TTC
    if main_oeuvre > 0:
        totaux_rows.append(['', Paragraph("Main d'œuvre", s_right),
                            Paragraph(f"{fmt(main_oeuvre)}XOF", s_right)])
    if remise and remise > 0:
        totaux_rows.append(['', Paragraph("Remise", s_right),
                            Paragraph(f"-{fmt(remise)}XOF", s_right)])
    if petites_fourn > 0:
        totaux_rows.append(['', Paragraph("petites fournitures", s_right),
                            Paragraph(f"{fmt(petites_fourn)}XOF", s_right)])
    if tva_active and tva_amount > 0:
        totaux_rows.append(['', Paragraph(f"TVA ({tva_rate:.0f}%)", s_right),
                            Paragraph(f"{fmt(tva_amount)}XOF", s_right)])
    tot_t = Table(totaux_rows, colWidths=[110*mm, 32*mm, 38*mm])
    tot_t.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(tot_t)
    
    # Barre Total TTC VERTE
    ttc_row = [[
        '',
        Paragraph("<b>Total TTC</b>",
            ParagraphStyle('ttc_l', fontSize=12, fontName='Helvetica-Bold', textColor=white, alignment=TA_RIGHT)),
        Paragraph(f"<b>{fmt(total_ttc)}XOF</b>",
            ParagraphStyle('ttc_r', fontSize=12, fontName='Helvetica-Bold', textColor=white, alignment=TA_RIGHT))
    ]]
    ttc_t = Table(ttc_row, colWidths=[110*mm, 32*mm, 38*mm])
    ttc_t.setStyle(TableStyle([
        ('BACKGROUND', (1, 0), (-1, 0), RAMYA_GREEN),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (1, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (1, 0), (-1, 0), 6),
        ('LEFTPADDING', (1, 0), (-1, 0), 10),
        ('RIGHTPADDING', (1, 0), (-1, 0), 10),
    ]))
    story.append(ttc_t)
    story.append(Spacer(1, 4*mm))
    
    # =========================================================
    # BARRE RÉCAPITULATIVE ORANGE
    # =========================================================
    # v160 : total_ht stocké = pièces uniquement. Pièces = total_ht ; brut = pièces + main d'œuvre.
    total_pieces = total_ht
    total_brut = total_ht + main_oeuvre
    total_net = total_brut - remise
    
    summary_hdrs = ['TOTAL PIÈCES', "MAIN D'ŒUVRE", 'TOTAL BRUT', 'REMISE', 'TOTAL NET', 'PETITES FOURN.', 'TOTAL TTC']
    summary_vals = [
        fmt(total_pieces), fmt(main_oeuvre), fmt(total_brut),
        fmt(remise), fmt(total_net), fmt(petites_fourn), fmt(total_ttc)
    ]
    
    s_hdr = ParagraphStyle('sh', fontSize=6, fontName='Helvetica-Bold',
                            textColor=white, alignment=TA_CENTER, leading=7)
    s_val = ParagraphStyle('sv', fontSize=8, fontName='Helvetica-Bold',
                            textColor=HexColor('#333'), alignment=TA_CENTER)
    s_val_last = ParagraphStyle('svl', fontSize=8, fontName='Helvetica-Bold',
                                 textColor=HexColor('#000'), alignment=TA_CENTER)
    
    bar_data = [
        [Paragraph(h, s_hdr) for h in summary_hdrs],
        [Paragraph(f"{v}XOF", s_val if i < 6 else s_val_last) for i, v in enumerate(summary_vals)],
    ]
    bar_widths = [25*mm, 25*mm, 25*mm, 22*mm, 25*mm, 26*mm, 32*mm]
    bar = Table(bar_data, colWidths=bar_widths)
    bar.setStyle(TableStyle([
        # Ligne 0: header (TEAL pour facture, ORANGE pour devis)
        ('BACKGROUND', (0, 0), (-1, 0), header_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('TOPPADDING', (0, 0), (-1, 0), 4),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
        # Ligne 1: valeurs sur fond blanc
        ('BACKGROUND', (0, 1), (-1, 1), white),
        ('TOPPADDING', (0, 1), (-1, 1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 3),
        ('BOX', (0, 0), (-1, -1), 0.25, HexColor('#bbbbbb')),
        ('LINEBELOW', (0, 0), (-1, 0), 0.25, RAMYA_ORANGE),
    ]))
    story.append(bar)
    story.append(Spacer(1, 3*mm))
    
    # =========================================================
    # MONTANT EN LETTRES
    # =========================================================
    words = number_to_words_fr(int(total_ttc))
    story.append(Paragraph(
        f"<b>Sauf erreur, arrêté à la somme de: {words} Francs CFA</b>",
        ParagraphStyle('words', fontSize=9.5, alignment=TA_CENTER, textColor=HexColor('#000'))
    ))
    story.append(Spacer(1, 4*mm))
    
    # =========================================================
    # BAS DE PAGE : Note (haut) + Signature(gauche)+Visa(droite) en dessous
    # Même structure que le modèle fourni : tout sur une seule page
    # =========================================================
    story.append(Paragraph(
        "<b>Note:</b>",
        ParagraphStyle('noteL', fontSize=10, fontName='Helvetica-Bold')
    ))
    story.append(Paragraph(
        "<font size='9'>MODE DE REGLEMENT (Espèce, Chèque, Virement, Mobile money)</font>",
        ParagraphStyle('noteD', fontSize=9, textColor=HexColor('#333'))
    ))
    story.append(Spacer(1, 5*mm))
    
    # Signature gauche (avec signature électronique si présente), Visa droite
    # v74 : afficher la signature électronique si elle existe
    sig_data = devis_data.get('signature_data', '') or ''
    signed_by = devis_data.get('signed_by', '') or ''
    signed_at = devis_data.get('signed_at', '') or ''
    signed_role = devis_data.get('signed_role', '') or ''
    
    left_sig_elems = [Paragraph("<i>Signature autorisée</i>",
            ParagraphStyle('sigL', fontSize=10, fontName='Helvetica-Oblique'))]
    
    if sig_data and sig_data.startswith('data:image'):
        sig_img_ok = False
        try:
            import base64 as _b64, io as _io2, tempfile as _tmp2
            from PIL import Image as _PILImg
            b64part = sig_data.split(',', 1)[1] if ',' in sig_data else sig_data
            raw = _b64.b64decode(b64part)
            pil = _PILImg.open(_io2.BytesIO(raw))
            if pil.mode == 'RGBA':
                bg = _PILImg.new('RGB', pil.size, (255, 255, 255))
                bg.paste(pil, mask=pil.split()[3])
                pil = bg
            elif pil.mode != 'RGB':
                pil = pil.convert('RGB')
            # Garder le ratio, largeur cible 48mm
            w0, h0 = pil.size
            target_w = 48.0  # mm
            ratio = (h0 / w0) if w0 else 0.4
            target_h = max(8.0, min(28.0, target_w * ratio))
            _sf = _tmp2.NamedTemporaryFile(suffix='.jpg', delete=False)
            _sf.close()
            pil.save(_sf.name, 'JPEG', quality=92)
            left_sig_elems.append(Spacer(1, 2*mm))
            # v75 : utiliser RLImage (Image est importé comme RLImage dans ce module)
            left_sig_elems.append(RLImage(_sf.name, width=target_w*mm, height=target_h*mm))
            sig_img_ok = True
        except Exception as _sig_err:
            import traceback as _tb
            print(f"[generate_devis_pdf] Signature image error: {type(_sig_err).__name__}: {_sig_err}", flush=True)
            _tb.print_exc()
            sig_img_ok = False
        
        # Toujours afficher le nom/qualité/date du signataire (même si l'image échoue)
        sig_caption = []
        if not sig_img_ok:
            sig_caption.append("<font size='8' color='#2e7d32'>✓ Signé électroniquement</font>")
        if signed_by:
            sig_caption.append(f"<b>{signed_by}</b>")
        if signed_role:
            sig_caption.append(f"<font size='8'>{signed_role}</font>")
        if signed_at:
            sig_caption.append(f"<font size='8' color='#666'>Signé le {signed_at}</font>")
        if sig_caption:
            left_sig_elems.append(Spacer(1, 1*mm))
            left_sig_elems.append(Paragraph("<br/>".join(sig_caption),
                ParagraphStyle('sigcap', fontSize=9, leading=11)))
    
    sig_row = [[
        left_sig_elems,
        Paragraph("<b>Visa Client</b>",
            ParagraphStyle('visa', fontSize=10, fontName='Helvetica-Bold', alignment=TA_RIGHT))
    ]]
    sig_t = Table(sig_row, colWidths=[110*mm, 70*mm])
    sig_t.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(sig_t)
    
    # ==== FOOTER via canvas ====
    def _footer(canv, doc_):
        _draw_corners(canv, doc_)
        canv.saveState()
        canv.setFont('Helvetica-Bold', 7)
        canv.setFillColor(RAMYA_TEAL)
        w, _ = A4
        if doc_params:
            lines = []
            if doc_params.get('footer_text'): lines.append(doc_params['footer_text'])
            lines.append(f"{doc_params.get('company_name','')} · {doc_params.get('company_address','')} / N°RCCM : {doc_params.get('company_rccm','')} / NCC : {doc_params.get('company_ncc','')}")
            if doc_params.get('bank_for_payment'): lines.append(f"Compte bancaire : {doc_params['bank_for_payment']}")
            lines.append(f"Tél : {doc_params.get('company_phone','')} · Email : {doc_params.get('company_email','')} · {doc_params.get('company_website','')}")
            if doc_params.get('footer_legal'): lines.append(doc_params['footer_legal'])
        else:
            lines = [
                "Siège social ABIDJAN Cocody ABATTA derrière la station OLA ENERGY / N°RCCM : CI-ABJ-2017-A-25092 / NCC : 1746141.B",
                "Compte bancaire : Orabank N° : 033201001901 / Bdu N° : 20401160186 / Cel : + 225 2722204498 / 07 09 50 02 43 / 07 47 68 20 27",
                "Email: dg@ramyaci.tech - admin@ramyaci.tech - www.ramyatechnologie.com",
            ]
        y = 15*mm
        for ln in lines:
            canv.drawCentredString(w/2, y, ln); y -= 3.5*mm
        canv.restoreState()
    
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return output_path


# =============================================================================
# BON DE LIVRAISON — Document officiel remis au client en fin de chantier
# =============================================================================

def generate_bon_livraison_pdf(inter_data, output_path, logo_path=None):
    """Génère un bon de livraison PDF signé par coordinateur + client + technicien.
    
    inter_data dict must include:
      - delivery_bon_ref, reference (intervention ref), title, client_name, site_address
      - delivered_at, delivery_proposed_date
      - delivery_signed_client, delivery_signed_coordinator, delivery_signed_technicien
      - technician_name, delivered_by_name (coordinator), client_code
    """
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.colors import HexColor
    import base64, io, os
    
    RAMYA_TEAL = HexColor('#1A7A6D')
    RAMYA_ORANGE = HexColor('#F29F2F')
    
    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=25*mm)
    story = []
    
    # HEADER
    logo_el = Paragraph("", ParagraphStyle('empty', fontSize=1))
    if logo_path and os.path.exists(logo_path):
        try:
            try:
                from PIL import Image as _PIL
                with _PIL.open(logo_path) as _im:
                    ratio = _im.width / _im.height
            except: ratio = 0.76
            logo_w = 20*mm; logo_h = logo_w / ratio
            logo_el = RLImage(logo_path, width=logo_w, height=logo_h)
        except: pass
    
    company = Paragraph("<b>RAMYA</b><br/><b>TECHNOLOGIE &amp; INNOVATION</b>",
        ParagraphStyle('co', fontSize=10, fontName='Helvetica-Bold', textColor=RAMYA_TEAL, leading=13, alignment=TA_CENTER))
    
    title_right = Paragraph(
        f"<font size='22' color='#1A7A6D'><b>BON DE LIVRAISON</b></font><br/>"
        f"<font size='10'>N° <b>{inter_data.get('delivery_bon_ref','-')}</b></font><br/>"
        f"<font size='9'>Intervention : {inter_data.get('reference','-')}</font><br/>"
        f"<font size='9'>Date : {inter_data.get('delivered_at','')[:10]}</font>",
        ParagraphStyle('tr', alignment=TA_RIGHT, leading=14))
    
    ht = Table([[logo_el, company, title_right]], colWidths=[28*mm, 57*mm, 95*mm])
    ht.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('BACKGROUND',(0,0),(0,0), white),
        ('LEFTPADDING',(0,0),(-1,-1),4),
        ('RIGHTPADDING',(0,0),(-1,-1),4),
    ]))
    story.append(ht)
    story.append(Spacer(1, 6*mm))
    
    # CLIENT INFO
    client_block = Paragraph(
        f"<b>REMIS À :</b><br/>"
        f"<b>{inter_data.get('client_name','-')}</b>"
        f"{' (Code : ' + inter_data.get('client_code','') + ')' if inter_data.get('client_code') else ''}<br/>"
        f"{'Site : ' + inter_data.get('site_address','') if inter_data.get('site_address') else ''}",
        ParagraphStyle('cl', fontSize=11, leading=16))
    story.append(client_block)
    story.append(Spacer(1, 4*mm))
    
    # OBJET
    objet_para = Paragraph(
        f"<b>OBJET DE LA LIVRAISON :</b><br/>{inter_data.get('title','-')}",
        ParagraphStyle('obj', fontSize=11, leading=15))
    story.append(objet_para)
    story.append(Spacer(1, 4*mm))
    
    # DETAILS TABLE
    details = [
        ["Date de livraison", inter_data.get('delivery_proposed_date') or inter_data.get('delivered_at','')[:10]],
        ["Heure d'exécution", inter_data.get('delivered_at','')[11:16] if inter_data.get('delivered_at') else '-'],
        ["Type d'intervention", inter_data.get('type','-').capitalize()],
        ["Livré par (coordinateur)", inter_data.get('delivered_by_name','-')],
        ["Technicien responsable", inter_data.get('technician_name','-')],
    ]
    dt = Table(details, colWidths=[60*mm, 120*mm])
    dt.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(0,-1), HexColor('#f0f7f7')),
        ('TEXTCOLOR',(0,0),(0,-1), RAMYA_TEAL),
        ('FONTNAME',(0,0),(0,-1), 'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1), 10),
        ('GRID',(0,0),(-1,-1), 0.5, HexColor('#dddddd')),
        ('VALIGN',(0,0),(-1,-1), 'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1), 8),
        ('TOPPADDING',(0,0),(-1,-1), 6),
        ('BOTTOMPADDING',(0,0),(-1,-1), 6),
    ]))
    story.append(dt)
    story.append(Spacer(1, 6*mm))
    
    # NOTE CLIENT
    note_para = Paragraph(
        "<b>CONDITIONS DE LIVRAISON :</b><br/>"
        "Le client reconnaît avoir reçu la prestation décrite ci-dessus dans son intégralité, "
        "et déclare en avoir vérifié la conformité avec le cahier des charges et les fonctionnalités convenues. "
        "Toute réserve doit être consignée ci-dessous.",
        ParagraphStyle('nt', fontSize=9, leading=12, textColor=HexColor('#444')))
    story.append(note_para)
    story.append(Spacer(1, 4*mm))
    
    # OBSERVATIONS
    obs_box = Table([["Observations / Réserves éventuelles :", ""]], colWidths=[180*mm, 0*mm], rowHeights=[24*mm])
    obs_box.setStyle(TableStyle([
        ('BOX',(0,0),(-1,-1), 1, HexColor('#999999')),
        ('VALIGN',(0,0),(-1,-1), 'TOP'),
        ('LEFTPADDING',(0,0),(-1,-1), 8),
        ('TOPPADDING',(0,0),(-1,-1), 6),
        ('FONTSIZE',(0,0),(-1,-1), 9),
    ]))
    story.append(obs_box)
    story.append(Spacer(1, 6*mm))
    
    # ===== SIGNATURES (3 blocs côte-à-côte) =====
    def sig_cell(title, subtitle, sig_data):
        img_el = Paragraph("<font size='9' color='#888'><i>(signature)</i></font>",
            ParagraphStyle('s', fontSize=9, alignment=TA_CENTER, textColor=HexColor('#999')))
        if sig_data and sig_data.startswith('data:image'):
            try:
                b64 = sig_data.split(',',1)[1]
                raw = base64.b64decode(b64)
                img_el = RLImage(io.BytesIO(raw), width=50*mm, height=18*mm, kind='proportional')
            except: pass
        return [
            Paragraph(f"<b>{title}</b>", ParagraphStyle('t',fontSize=9,alignment=TA_CENTER, textColor=RAMYA_TEAL)),
            Paragraph(f"<font size='8' color='#666'>{subtitle}</font>",
                ParagraphStyle('s',fontSize=8,alignment=TA_CENTER)),
            Spacer(1, 2*mm),
            img_el,
        ]
    
    sig_data = [[
        sig_cell("Le Client", inter_data.get('client_name','')[:28], inter_data.get('delivery_signed_client')),
        sig_cell("Le Coordinateur", inter_data.get('delivered_by_name',''), inter_data.get('delivery_signed_coordinator')),
        sig_cell("Le Responsable Technique", inter_data.get('technician_name',''), inter_data.get('delivery_signed_technicien')),
    ]]
    st = Table(sig_data, colWidths=[60*mm, 60*mm, 60*mm])
    st.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1), 'TOP'),
        ('BOX',(0,0),(-1,-1), 0.5, HexColor('#cccccc')),
        ('INNERGRID',(0,0),(-1,-1), 0.5, HexColor('#eeeeee')),
        ('TOPPADDING',(0,0),(-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
        ('LEFTPADDING',(0,0),(-1,-1), 4),
        ('RIGHTPADDING',(0,0),(-1,-1), 4),
    ]))
    story.append(st)
    story.append(Spacer(1, 4*mm))
    
    # Footer
    def _footer(canv, doc_):
        canv.saveState()
        canv.setFont('Helvetica-Bold', 7)
        canv.setFillColor(RAMYA_TEAL)
        w, _ = A4
        lines = [
            "Siège social ABIDJAN Cocody ABATTA derrière la station OLA ENERGY / N°RCCM : CI-ABJ-2017-A-25092 / NCC : 1746141.B",
            "Compte bancaire : Orabank N° : 033201001901 / Bdu N° : 20401160186 / Cel : + 225 2722204498 / 07 09 50 02 43 / 07 47 68 20 27",
            "Email: dg@ramyaci.tech - admin@ramyaci.tech - www.ramyatechnologie.com",
        ]
        y = 15*mm
        for ln in lines:
            canv.drawCentredString(w/2, y, ln); y -= 3.5*mm
        canv.restoreState()
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return output_path


# =============================================================================
# ATTESTATION DE BONNE EXÉCUTION — Soumise au client pour signature
# =============================================================================

def generate_attestation_pdf(inter_data, output_path, logo_path=None):
    """Attestation de bonne exécution des travaux, pour signature client après livraison.
    """
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.colors import HexColor
    import base64, io, os
    from datetime import datetime
    
    RAMYA_TEAL = HexColor('#1A7A6D')
    
    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=30*mm)
    story = []
    
    # HEADER
    logo_el = Paragraph("", ParagraphStyle('empty', fontSize=1))
    if logo_path and os.path.exists(logo_path):
        try:
            try:
                from PIL import Image as _PIL
                with _PIL.open(logo_path) as _im:
                    ratio = _im.width / _im.height
            except: ratio = 0.76
            logo_w = 22*mm; logo_h = logo_w / ratio
            logo_el = RLImage(logo_path, width=logo_w, height=logo_h)
        except: pass
    
    company = Paragraph("<b>RAMYA</b><br/><b>TECHNOLOGIE &amp; INNOVATION</b>",
        ParagraphStyle('co', fontSize=10, fontName='Helvetica-Bold', textColor=RAMYA_TEAL, leading=13, alignment=TA_CENTER))
    
    ht = Table([[logo_el, company, ""]], colWidths=[28*mm, 60*mm, 86*mm])
    ht.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('BACKGROUND',(0,0),(0,0), white),
        ('LEFTPADDING',(0,0),(-1,-1),4),
    ]))
    story.append(ht)
    story.append(Spacer(1, 10*mm))
    
    # Title
    title = Paragraph(
        "<para alignment='center'><font size='22' color='#1A7A6D'><b>ATTESTATION DE BONNE EXÉCUTION</b></font></para>",
        ParagraphStyle('t', alignment=TA_CENTER, leading=28))
    story.append(title)
    story.append(Spacer(1, 4*mm))
    
    ref_line = Paragraph(
        f"<para alignment='center'><font size='10' color='#666'>"
        f"Référence intervention : <b>{inter_data.get('reference','-')}</b>"
        f"{' · Bon de livraison : ' + inter_data.get('delivery_bon_ref','') if inter_data.get('delivery_bon_ref') else ''}"
        f"</font></para>",
        ParagraphStyle('rl', alignment=TA_CENTER))
    story.append(ref_line)
    story.append(Spacer(1, 10*mm))
    
    # Body
    body_text = (
        f"Je soussigné(e), <b>___________________________________________</b>, "
        f"représentant légal de la société <b>{inter_data.get('client_name','-')}</b>"
        f"{' (Code client : ' + inter_data.get('client_code','') + ')' if inter_data.get('client_code') else ''}"
        f", atteste par la présente que la société <b>RAMYA TECHNOLOGIE &amp; INNOVATION</b> "
        f"a exécuté avec <b>satisfaction</b> les prestations suivantes :<br/><br/>"
        f"<b>Objet des travaux :</b> {inter_data.get('title','-')}<br/>"
        f"<b>Site d'intervention :</b> {inter_data.get('site_address') or '-'}<br/>"
        f"<b>Date de livraison :</b> {inter_data.get('delivery_proposed_date') or (inter_data.get('delivered_at','')[:10])}<br/><br/>"
        f"Les travaux ont été réalisés conformément au cahier des charges convenu, "
        f"dans les délais impartis et selon les règles de l'art. "
        f"Cette attestation est délivrée à RAMYA TECHNOLOGIE &amp; INNOVATION pour servir et valoir ce que de droit."
    )
    body = Paragraph(body_text, ParagraphStyle('body', fontSize=11, leading=18, alignment=TA_LEFT))
    story.append(body)
    story.append(Spacer(1, 14*mm))
    
    # Date + place
    place_date = Paragraph(
        f"Fait à ______________________, le ______________________",
        ParagraphStyle('pd', fontSize=11, leading=16))
    story.append(place_date)
    story.append(Spacer(1, 16*mm))
    
    # Signature zone client (grand cadre pour signature manuscrite)
    sig_title = Paragraph(
        "<para alignment='right'><b>Signature du client</b><br/><font size='8' color='#888'>(Nom, qualité, cachet)</font></para>",
        ParagraphStyle('st', alignment=TA_RIGHT, fontSize=11))
    story.append(sig_title)
    story.append(Spacer(1, 2*mm))
    
    sig_box = Table([[""]], colWidths=[90*mm], rowHeights=[35*mm])
    sig_box.setStyle(TableStyle([
        ('BOX',(0,0),(-1,-1), 0.8, HexColor('#666666')),
        ('BACKGROUND',(0,0),(-1,-1), HexColor('#fafafa')),
    ]))
    # Place signature box on the right
    sig_wrap = Table([['', sig_box]], colWidths=[84*mm, 90*mm])
    sig_wrap.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP')]))
    story.append(sig_wrap)
    
    # Footer
    def _footer(canv, doc_):
        canv.saveState()
        canv.setFont('Helvetica-Bold', 7)
        canv.setFillColor(RAMYA_TEAL)
        w, _ = A4
        lines = [
            "Siège social ABIDJAN Cocody ABATTA derrière la station OLA ENERGY / N°RCCM : CI-ABJ-2017-A-25092 / NCC : 1746141.B",
            "Compte bancaire : Orabank N° : 033201001901 / Bdu N° : 20401160186 / Cel : + 225 2722204498 / 07 09 50 02 43 / 07 47 68 20 27",
            "Email: dg@ramyaci.tech - admin@ramyaci.tech - www.ramyatechnologie.com",
        ]
        y = 15*mm
        for ln in lines:
            canv.drawCentredString(w/2, y, ln); y -= 3.5*mm
        canv.restoreState()
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return output_path


# ======================== RAPPORT SESSION (NOUVEAU v56) ========================

def gen_session_pages(story, emps_sessions, S, provider_name, provider_info, client_name, client_info, period, now):
    """
    Génère un rapport adapté pour le mode pointage SESSION.
    Pour chaque employé, affiche le détail de chaque session (heure prévue, début, fin, durée, statut).
    
    emps_sessions : liste de dicts par employé contenant :
        {
            'name': 'Alice MARTIN',
            'ref': 'alice',
            'sessions': [
                {'date': '2026-04-01', 'libelle': 'Mission matin',
                 'heure_prevue': '08:00', 'duree_minutes': 240,
                 'heure_debut': '08:05', 'heure_fin': '12:00',
                 'duree_reelle': 235, 'statut': 'ok' / 'retard' / 'non_effectuee'}
            ],
            'total_sessions': 12,
            'sessions_ok': 10, 'sessions_retard': 1, 'sessions_non_effectuee': 1,
            'total_duree_prevue': 2880, 'total_duree_reelle': 2750,
        }
    """
    total_emps = len(emps_sessions)
    
    for idx, emp in enumerate(emps_sessions):
        if idx > 0:
            story.append(SmartPageBreak())
        
        emp_num = idx + 1
        sessions = emp.get('sessions', [])
        
        story.append(make_header(S, provider_name, provider_info, client_name, client_info))
        story.append(Spacer(1, 1*mm))
        story.append(Paragraph(
            f"RAPPORT SESSIONS — {emp['name']} (Réf: {emp['ref']}) — Fiche {emp_num}/{total_emps} — {period}",
            ParagraphStyle('ti2', fontName='Helvetica-Bold', fontSize=11, textColor=TEAL,
                           alignment=TA_CENTER, spaceAfter=2)))
        story.append(Spacer(1, 1*mm))
        
        # === BANDEAU RÉSUMÉ ===
        total = emp.get('total_sessions', 0)
        ok = emp.get('sessions_ok', 0)
        retard = emp.get('sessions_retard', 0)
        non_eff = emp.get('sessions_non_effectuee', 0)
        durée_prev = emp.get('total_duree_prevue', 0)
        durée_reel = emp.get('total_duree_reelle', 0)
        
        def m2h(m):
            if not m: return '00:00'
            return f"{int(m)//60:02d}:{int(m)%60:02d}"
        
        sum_data = [[
            Paragraph("<b>Total sessions</b>", S['h']),
            Paragraph("<b>Effectuées<br/>à l'heure</b>", S['h']),
            Paragraph("<b>Effectuées<br/>en retard</b>", S['h']),
            Paragraph("<b>Non effectuées</b>", S['h']),
            Paragraph("<b>Durée prévue</b>", S['h']),
            Paragraph("<b>Durée réelle</b>", S['h']),
            Paragraph("<b>Taux<br/>d'efficacité</b>", S['h']),
        ], [
            Paragraph(f"{total}", S['cb']),
            Paragraph(f"{ok}", S['g']),
            Paragraph(f"{retard}", ParagraphStyle('ret', fontName='Helvetica-Bold', fontSize=9,
                                                   textColor=ORANGE, alignment=TA_CENTER, leading=11)),
            Paragraph(f"{non_eff}", S['r']),
            Paragraph(f"{m2h(durée_prev)}", S['cb']),
            Paragraph(f"{m2h(durée_reel)}", S['cb']),
            Paragraph(f"{(ok/total*100 if total else 0):.0f}%",
                     S['g'] if (ok/total if total else 0) >= 0.8 else S['r']),
        ]]
        sum_t = Table(sum_data, colWidths=[28*mm]*7)
        sum_t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), TEAL),
            ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(sum_t)
        story.append(Spacer(1, 4*mm))
        
        # === TABLEAU DE DÉTAIL DES SESSIONS ===
        if sessions:
            td = [[
                Paragraph("<b>N°</b>", S['h']),
                Paragraph("<b>Date</b>", S['h']),
                Paragraph("<b>Mission / Libellé</b>", S['h']),
                Paragraph("<b>Heure<br/>prévue</b>", S['h']),
                Paragraph("<b>Durée<br/>prévue</b>", S['h']),
                Paragraph("<b>Début<br/>réel</b>", S['h']),
                Paragraph("<b>Fin<br/>réelle</b>", S['h']),
                Paragraph("<b>Durée<br/>réelle</b>", S['h']),
                Paragraph("<b>Écart<br/>début</b>", S['h']),
                Paragraph("<b>Statut</b>", S['h']),
            ]]
            
            for i, s in enumerate(sessions, 1):
                # Calculer écart début (si retard)
                ecart = ''
                try:
                    if s.get('heure_debut') and s.get('heure_prevue'):
                        h1, m1 = map(int, s['heure_prevue'].split(':'))
                        h2, m2 = map(int, s['heure_debut'].split(':'))
                        diff = (h2 - h1) * 60 + (m2 - m1)
                        if diff > 0:
                            ecart = f"+{diff}m"
                        elif diff < 0:
                            ecart = f"{diff}m"
                        else:
                            ecart = '0m'
                except: pass
                
                statut = s.get('statut', '')
                if statut == 'ok':
                    statut_text = '✓ OK'
                    statut_style = S['g']
                elif statut == 'retard':
                    statut_text = '⚠ Retard'
                    statut_style = ParagraphStyle('rt', fontName='Helvetica-Bold', fontSize=7,
                                                   textColor=ORANGE, alignment=TA_CENTER, leading=8)
                elif statut == 'non_effectuee':
                    statut_text = '✗ Non eff.'
                    statut_style = S['r']
                else:
                    statut_text = statut or '-'
                    statut_style = S['c']
                
                td.append([
                    Paragraph(str(i), S['c']),
                    Paragraph(s.get('date', ''), S['c']),
                    Paragraph(s.get('libelle', '-'), S['c']),
                    Paragraph(s.get('heure_prevue', '-'), S['c']),
                    Paragraph(m2h(s.get('duree_minutes', 0)), S['c']),
                    Paragraph(s.get('heure_debut', '') or '-', S['c']),
                    Paragraph(s.get('heure_fin', '') or '-', S['c']),
                    Paragraph(m2h(s.get('duree_reelle', 0)), S['c']),
                    Paragraph(ecart, S['c']),
                    Paragraph(statut_text, statut_style),
                ])
            
            cw_s = [8*mm, 18*mm, 36*mm, 17*mm, 16*mm, 17*mm, 16*mm, 16*mm, 14*mm, 24*mm]
            dt = Table(td, colWidths=cw_s, repeatRows=1)
            sc = [('BACKGROUND', (0,0), (-1,0), TEAL),
                  ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
                  ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                  ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                  ('TOPPADDING', (0,0), (-1,-1), 1),
                  ('BOTTOMPADDING', (0,0), (-1,-1), 1)]
            for i in range(2, len(td), 2):
                sc.append(('BACKGROUND', (0,i), (-1,i), LGRAY))
            dt.setStyle(TableStyle(sc))
            story.append(dt)
        else:
            story.append(Paragraph("<i>Aucune session enregistrée pour cette période</i>",
                                   ParagraphStyle('none', fontSize=10, textColor=colors.grey,
                                                  alignment=TA_CENTER)))
        
        story.append(Spacer(1, 4*mm))
        
        # Footer line
        story.append(Paragraph(
            f"Généré le {now.strftime('%d/%m/%Y à %H:%M')} | {client_name} - Rapport sessions {emp['name']} {emp_num}/{total_emps}",
            S['ft']))




# ======================== PHARMACIE (v69 — refonte présence) ========================
# Format Excel attendu (vraies colonnes des systèmes de pointage type "Time Doctor", "Jibble"...) :
#   Prénom | Nom de famille | ID | Service | Date | Heure d'arrivée oblig. | Heure de départ oblig.
#   | Heure de contrôle d'arrivée | Sortie à | Durée
#
# Différences clé avec /dpci :
#   - 2 colonnes pour le nom (Prénom + Nom famille) -> on les concatène
#   - Heure d'arrivée obligatoire peut varier par jour (matin 07:30 ou aprem 13:00)
#   - Pas de colonne pause (durée = sortie - arrivée, pas de pause à soustraire)
#   - "-" dans Sortie / Durée = badge non sorti (absence ou oubli)


def _norm_excel_time(v):
    """Convertit une valeur Excel en HH:MM. Gère les décimales (0.5416... = 13:00)."""
    if v is None or v == '':
        return ''
    s = str(v).strip()
    if s == '-' or s == '':
        return ''
    # Format décimal Excel (fraction d'un jour)
    try:
        f = float(s)
        if 0 <= f < 1:
            total_min = int(round(f * 24 * 60))
            return f'{total_min // 60:02d}:{total_min % 60:02d}'
        # Si f >= 1, peut être un timestamp Excel
        if 1 <= f < 100:
            # Fractionnaire d'un jour avec partie entière
            day_frac = f - int(f)
            total_min = int(round(day_frac * 24 * 60))
            return f'{total_min // 60:02d}:{total_min % 60:02d}'
    except (ValueError, TypeError):
        pass
    # Objet datetime ou time
    if hasattr(v, 'strftime'):
        return v.strftime('%H:%M')
    # Si déjà HH:MM
    if len(s) >= 5 and s[2] == ':':
        return s[:5]
    return s


def _norm_excel_date(v):
    """Convertit en YYYY-MM-DD."""
    if v is None: return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if len(s) >= 10 and s[4] == '-' and s[7] == '-':
        return s[:10]
    if '/' in s:
        parts = s.split('/')
        if len(parts) == 3:
            try:
                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                if y < 100: y += 2000
                return f'{y:04d}-{m:02d}-{d:02d}'
            except: pass
    return s


def parse_pharma_excel(xlsx_path):
    """Parse un Excel de pointage pharmacie.
    
    Colonnes attendues :
      Prénom | Nom de famille | ID | Service | Date | H.arrivée oblig. 
      | H.départ oblig. | H.contrôle arrivée | Sortie à | Durée
    
    Retourne (emps, period) où :
      emps = [{
        'name': 'Prénom Nom',
        'id': 'PHCIE01',
        'service': 'PHARMACIE GNIMAH',
        'records': [{
          'date': '2026-04-30',
          'sched_start': '07:30',     # heure arrivée oblig.
          'sched_end': '20:00',       # heure départ oblig.
          'arrival': '06:24',          # heure réelle arrivée
          'departure': '13:32',        # heure réelle sortie ('' si "-")
          'duration': '07:08',         # durée déjà calculée par le système
        }, ...]
      }, ...]
      period = '2026-04' (ou range si chevauche)
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    # Détection en-têtes (ligne contenant Prénom + Nom + Date)
    header_row = None
    col_map = {}
    for r in range(1, min(10, ws.max_row + 1)):
        vals = [(str(ws.cell(r, c).value or '')).strip().lower() for c in range(1, ws.max_column + 1)]
        joined = ' | '.join(vals)
        if 'prénom' in joined or 'prenom' in joined:
            if 'nom' in joined and 'date' in joined:
                header_row = r
                for c, v in enumerate(vals, 1):
                    if 'prénom' in v or 'prenom' in v: col_map['prenom'] = c
                    elif 'nom' in v and 'famille' in v: col_map['nom'] = c
                    elif v == 'nom': col_map['nom'] = c  # si juste "Nom"
                    elif v == 'id' or 'matricule' in v: col_map['id'] = c
                    elif 'service' in v or 'département' in v or 'departement' in v: col_map['service'] = c
                    elif v == 'date': col_map['date'] = c
                    elif 'arrivée' in v and ('oblig' in v or 'prévue' in v or 'prevue' in v or 'théo' in v): col_map['sched_start'] = c
                    elif 'départ' in v and ('oblig' in v or 'prévue' in v or 'prevue' in v or 'théo' in v): col_map['sched_end'] = c
                    elif 'contrôle' in v and 'arrivée' in v: col_map['arrival'] = c
                    elif 'pointage' in v and 'arrivée' in v: col_map['arrival'] = c
                    elif v.startswith('arrivée') and 'oblig' not in v and 'prévue' not in v: col_map['arrival'] = c
                    elif 'sortie' in v: col_map['departure'] = c
                    elif v == 'durée' or 'durée' in v: col_map['duration'] = c
                break
    
    # Si pas trouvé via "Prénom", chercher au moins "Nom" + "Date"
    if header_row is None:
        for r in range(1, min(10, ws.max_row + 1)):
            vals = [(str(ws.cell(r, c).value or '')).strip().lower() for c in range(1, ws.max_column + 1)]
            if any('nom' in v for v in vals) and any(v == 'date' for v in vals):
                header_row = r
                for c, v in enumerate(vals, 1):
                    if 'nom' in v and 'prenom' not in v: col_map['nom'] = c
                    elif v == 'date': col_map['date'] = c
                    elif 'arrivée' in v and 'oblig' in v: col_map['sched_start'] = c
                    elif 'départ' in v and 'oblig' in v: col_map['sched_end'] = c
                    elif 'arrivée' in v: col_map['arrival'] = c
                    elif 'sortie' in v: col_map['departure'] = c
                    elif 'durée' in v: col_map['duration'] = c
                break
    
    if header_row is None or 'date' not in col_map:
        raise ValueError("Fichier Excel invalide : impossible de trouver les colonnes attendues. "
                         "Le fichier doit contenir au moins : Nom, Date, Heure d'arrivée, Sortie.")
    
    # Auto-fallback : si pas de "nom" mais on a "prenom", inverser
    if 'nom' not in col_map and 'prenom' in col_map:
        col_map['nom'] = col_map['prenom']
        col_map.pop('prenom', None)
    
    employees = {}
    dates_seen = set()
    
    for r in range(header_row + 1, ws.max_row + 1):
        # Construire le nom complet
        prenom = str(ws.cell(r, col_map.get('prenom', 0)).value or '').strip() if 'prenom' in col_map else ''
        nom_fam = str(ws.cell(r, col_map['nom']).value or '').strip()
        if not nom_fam and not prenom: continue
        full_name = (prenom + ' ' + nom_fam).strip()
        if not full_name: continue
        
        date = _norm_excel_date(ws.cell(r, col_map['date']).value)
        if not date or len(date) < 8: continue
        
        ident = str(ws.cell(r, col_map.get('id', 0)).value or '').strip() if 'id' in col_map else ''
        service = str(ws.cell(r, col_map.get('service', 0)).value or '').strip() if 'service' in col_map else ''
        
        sched_start = _norm_excel_time(ws.cell(r, col_map.get('sched_start', 0)).value) if 'sched_start' in col_map else ''
        sched_end = _norm_excel_time(ws.cell(r, col_map.get('sched_end', 0)).value) if 'sched_end' in col_map else ''
        arrival = _norm_excel_time(ws.cell(r, col_map.get('arrival', 0)).value) if 'arrival' in col_map else ''
        departure = _norm_excel_time(ws.cell(r, col_map.get('departure', 0)).value) if 'departure' in col_map else ''
        duration = _norm_excel_time(ws.cell(r, col_map.get('duration', 0)).value) if 'duration' in col_map else ''
        
        # Si "-" → vide
        for fld in ['arrival', 'departure', 'duration']:
            pass  # déjà géré dans _norm_excel_time
        
        if full_name not in employees:
            employees[full_name] = {
                'name': full_name,
                'id': ident,
                'ref': ident,  # alias pour DPCI/Traitement
                'service': service,
                'department': service,  # alias DPCI-compatible
                'records': []
            }
        if not employees[full_name].get('id') and ident:
            employees[full_name]['id'] = ident
            employees[full_name]['ref'] = ident
        if not employees[full_name].get('service') and service:
            employees[full_name]['service'] = service
            employees[full_name]['department'] = service
        
        employees[full_name]['records'].append({
            'date': date,
            # Format DPCI-compatible : arrival/departure/pause_start/pause_end
            'arrival': arrival,
            'pause_start': '',  # pas de pause dans le format Pharmacie source
            'pause_end': '',
            'departure': departure,
            # Champs additionnels Pharmacie
            'sched_start': sched_start,
            'sched_end': sched_end,
            'duration': duration,
        })
        dates_seen.add(date)
    
    # Trier par date
    for emp in employees.values():
        emp['records'].sort(key=lambda r: r['date'])
    
    # Période
    period = ''
    if dates_seen:
        sd = sorted(dates_seen)
        if sd[0][:7] == sd[-1][:7]:
            period = sd[0][:7]
        else:
            period = f'{sd[0]} → {sd[-1]}'
    
    return list(employees.values()), period


def _classify_pharma_day(date_str, sched_start, feries_set, custom_classifications=None):
    """Classifie un jour : 'normal', 'garde_nuit', 'garde_we', 'ferie', ou 'astreinte'.
    
    Priorité :
      1. Override manuel (custom_classifications)
      2. Jour férié (date dans feries_set)
      3. Garde nuit (heure arrivée >= 18h ou < 6h)
      4. Garde week-end (samedi/dimanche)
      5. Normal
    """
    custom_classifications = custom_classifications or {}
    if date_str in custom_classifications:
        return custom_classifications[date_str]
    
    # Jour férié
    if date_str in feries_set:
        return 'ferie'
    
    # Date pour test WE
    try:
        from datetime import datetime as _dt
        d = _dt.strptime(date_str[:10], '%Y-%m-%d')
        weekday = d.weekday()
    except:
        weekday = 0
    
    # Garde de nuit : si l'heure d'arrivée prévue est >= 18:00 ou < 06:00
    if sched_start:
        try:
            h = int(sched_start.split(':')[0])
            if h >= 18 or h < 6:
                return 'garde_nuit'
        except: pass
    
    # Week-end
    if weekday >= 5:
        return 'garde_we'
    
    return 'normal'


def calc_pharma_employee_stats(emp, types_by_key, feries_set,
                                 default_hourly_rate=0, employee_rate=0,
                                 custom_classifications=None,
                                 hp=8):
    """Calcule les stats de présence pour un pharmacien.
    
    custom_classifications = {date_str: type_code} pour overrider la classif auto
    hp = heures obligatoires par défaut (utilisé si sched_start/sched_end absents)
    """
    custom_classifications = custom_classifications or {}
    
    enriched = []
    by_type = {}
    total_min_worked = 0    # heures effectivement travaillées (hors majoration)
    total_min_normal = 0    # heures normales (taux 1.0)
    total_min_majored = 0   # heures qui ont eu une majoration
    total_late_min = 0
    total_required_min = 0
    days_present = 0
    days_absent = 0
    days_late = 0
    days_no_dep = 0  # jours sans heure de sortie (badge oublié)
    total_primes = 0
    total_cost = 0
    
    rate = employee_rate or default_hourly_rate or 0
    
    counts_by_type = {'normal': 0, 'garde_nuit': 0, 'garde_we': 0, 'ferie': 0, 'astreinte': 0}
    
    for rec in emp['records']:
        date_str = rec['date']
        sched_s = rec.get('sched_start', '')
        sched_e = rec.get('sched_end', '')
        arrival = rec.get('arrival', '')
        departure = rec.get('departure', '')
        duration_str = rec.get('duration', '')
        
        # v69 : priorité 1 = _forced_type (programmation manuelle via UI)
        # priorité 2 = custom_classifications (rétrocompatibilité)
        # priorité 3 = détection auto (jour férié / WE / nuit)
        if rec.get('_forced_type'):
            type_code = rec['_forced_type']
        else:
            type_code = _classify_pharma_day(date_str, sched_s, feries_set, custom_classifications)
        counts_by_type[type_code] = counts_by_type.get(type_code, 0) + 1
        
        # Trouver les paramètres du type
        type_data = types_by_key.get(type_code) or types_by_key.get(type_code.replace('_', ' ')) or {}
        taux_maj = float(type_data.get('taux_majoration', 0) or 0)
        prime = float(type_data.get('prime_fixe', 0) or 0)
        type_libelle = type_data.get('libelle', type_code.replace('_', ' ').title())
        type_couleur = type_data.get('couleur', '#1A7A6D')
        
        # Calcul de la durée requise (en minutes)
        required_min = 0
        try:
            h1, m1 = map(int, sched_s.split(':'))
            h2, m2 = map(int, sched_e.split(':'))
            req = (h2 * 60 + m2) - (h1 * 60 + m1)
            if req < 0: req += 24 * 60  # traverse minuit
            required_min = req
        except:
            required_min = hp * 60 if hp else 0
        total_required_min += required_min
        
        # Statut & calculs
        worked_min = 0
        late_min = 0
        is_present = False
        is_absent = False
        no_dep = False
        
        if not arrival or arrival == '-':
            # Absence
            is_absent = True
            days_absent += 1
            state = 'ABS'
        elif not departure or departure == '-':
            # Badge non sorti
            is_present = True
            no_dep = True
            days_present += 1
            days_no_dep += 1
            state = 'ARR'  # Arrivé mais pas sorti
            # Si duration fournie, l'utiliser
            if duration_str:
                try:
                    dh, dm = duration_str.split(':')
                    worked_min = int(dh) * 60 + int(dm)
                except: pass
        else:
            is_present = True
            days_present += 1
            state = 'OK'
            # Durée
            if duration_str:
                try:
                    dh, dm = duration_str.split(':')
                    worked_min = int(dh) * 60 + int(dm)
                except: pass
            if worked_min == 0:
                # Calculer depuis arrival / departure
                try:
                    ah, am = map(int, arrival.split(':'))
                    dh, dm = map(int, departure.split(':'))
                    w = (dh * 60 + dm) - (ah * 60 + am)
                    if w < 0: w += 24 * 60
                    worked_min = w
                except: pass
            
            # Retard ?
            if sched_s and arrival:
                try:
                    sh, sm = map(int, sched_s.split(':'))
                    ah, am = map(int, arrival.split(':'))
                    diff = (ah * 60 + am) - (sh * 60 + sm)
                    if diff > 5:  # tolérance 5 min
                        late_min = diff
                        days_late += 1
                        state = 'RETARD'
                except: pass
        
        total_min_worked += worked_min
        total_late_min += late_min
        
        # Coût
        heures = worked_min / 60.0
        if taux_maj > 0:
            total_min_majored += worked_min
        else:
            total_min_normal += worked_min
        cout_heures = heures * rate * (1 + taux_maj)
        cout_jour = cout_heures + (prime if is_present else 0)
        total_cost += cout_jour
        if is_present:
            total_primes += prime
        
        # Stats par type
        if type_libelle not in by_type:
            by_type[type_libelle] = {'nb': 0, 'duree_min': 0, 'cout': 0,
                                     'couleur': type_couleur, 'code': type_code,
                                     'taux_maj': taux_maj}
        by_type[type_libelle]['nb'] += 1
        by_type[type_libelle]['duree_min'] += worked_min
        by_type[type_libelle]['cout'] += cout_jour
        
        # Jour de la semaine
        jour_lbl = ''
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(date_str[:10], '%Y-%m-%d')
            _DAYS = ['Lun','Mar','Mer','Jeu','Ven','Sam','Dim']
            jour_lbl = _DAYS[d.weekday()]
        except: pass
        
        enriched.append({
            **rec,
            'jour': jour_lbl,
            'type_code': type_code,
            'type_libelle': type_libelle,
            'type_couleur': type_couleur,
            'taux_maj': taux_maj,
            'prime': prime if is_present else 0,
            'required_min': required_min,
            'worked_min': worked_min,
            'worked_str': f'{worked_min // 60:02d}:{worked_min % 60:02d}' if worked_min > 0 else '00:00',
            'late_min': late_min,
            'late_str': f'{late_min // 60:02d}:{late_min % 60:02d}' if late_min > 0 else '',
            'state': state,
            'is_present': is_present,
            'is_absent': is_absent,
            'no_dep': no_dep,
            'cout_jour': cout_jour,
            'rate': rate,
        })
    
    nb_total_jours = len(emp['records'])
    
    return {
        'emp': emp,
        'enriched': enriched,
        'by_type': by_type,
        'rate': rate,
        'nb_total_jours': nb_total_jours,
        'days_present': days_present,
        'days_absent': days_absent,
        'days_late': days_late,
        'days_no_dep': days_no_dep,
        'total_min_worked': total_min_worked,
        'total_min_normal': total_min_normal,
        'total_min_majored': total_min_majored,
        'total_late_min': total_late_min,
        'total_required_min': total_required_min,
        'total_primes': total_primes,
        'total_cost': total_cost,
        'taux_presence': round(100 * days_present / nb_total_jours, 1) if nb_total_jours > 0 else 0,
        'counts_by_type': counts_by_type,
    }


def generate_pharma_pdf(emps, output_path, pharmacy_name='PHARMACIE', period='',
                         types_by_key=None, feries_set=None,
                         default_hourly_rate=0, employee_rates=None,
                         custom_classifications=None,
                         logo_path=None, hp=8):
    """Génère le PDF rapport de présence Pharmacie.
    
    Police lisible (8pt minimum), 1 page par employé (compactage si nécessaire).
    """
    if types_by_key is None: types_by_key = {}
    if feries_set is None: feries_set = set()
    if employee_rates is None: employee_rates = {}
    if custom_classifications is None: custom_classifications = {}
    
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=10*mm, bottomMargin=10*mm)
    story = []
    
    PURPLE = HexColor('#7b1fa2')
    NIGHT = HexColor('#3949ab')
    WE_COLOR = HexColor('#f29f2f')
    FERIE_COLOR = HexColor('#c53030')
    NORMAL_COLOR = HexColor('#1A7A6D')
    
    # === Styles LISIBLES (taille augmentée v69) ===
    s_title = ParagraphStyle('s_title', fontSize=15, fontName='Helvetica-Bold', textColor=PURPLE, alignment=TA_CENTER, leading=18)
    s_subtitle = ParagraphStyle('s_sub', fontSize=11, fontName='Helvetica-Bold', textColor=HexColor('#444'), leading=14)
    s_emp = ParagraphStyle('s_emp', fontSize=12, fontName='Helvetica-Bold', textColor=PURPLE, leading=15)
    s_kpi_n = ParagraphStyle('s_kn', fontSize=18, fontName='Helvetica-Bold', textColor=PURPLE, alignment=TA_CENTER, leading=20)
    s_kpi_l = ParagraphStyle('s_kl', fontSize=8, textColor=HexColor('#666'), alignment=TA_CENTER, leading=10)
    
    # Styles tableau (8pt par défaut, 7pt si dense)
    def build_styles(font_size):
        leading = font_size + 1.5
        return {
            'h': ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=font_size, textColor=white, alignment=TA_CENTER, leading=leading),
            'c': ParagraphStyle('c', fontName='Helvetica', fontSize=font_size, alignment=TA_CENTER, leading=leading),
            'cb': ParagraphStyle('cb', fontName='Helvetica-Bold', fontSize=font_size, alignment=TA_CENTER, leading=leading),
            'cl': ParagraphStyle('cl', fontName='Helvetica', fontSize=font_size, alignment=TA_LEFT, leading=leading),
            'g': ParagraphStyle('g', fontName='Helvetica-Bold', fontSize=font_size, textColor=HexColor('#2e7d32'), alignment=TA_CENTER, leading=leading),
            'r': ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=font_size, textColor=FERIE_COLOR, alignment=TA_CENTER, leading=leading),
            'o': ParagraphStyle('o', fontName='Helvetica-Bold', fontSize=font_size, textColor=HexColor('#b8561f'), alignment=TA_CENTER, leading=leading),
            'b': ParagraphStyle('b', fontName='Helvetica-Bold', fontSize=font_size, textColor=NIGHT, alignment=TA_CENTER, leading=leading),
        }
    
    def make_header(emp_name=None):
        title = '💊 RAPPORT DE PRÉSENCE PHARMACIE'
        if emp_name: title += f' — {emp_name}'
        cells = []
        left_cell = ''
        if logo_path and os.path.exists(logo_path):
            from reportlab.platypus import Image as _Img
            try: left_cell = _Img(logo_path, width=22*mm, height=22*mm)
            except: left_cell = ''
        title_cell = Paragraph(
            f"<b>{pharmacy_name}</b><br/><font size='10' color='#666'>{title}</font><br/>"
            f"<font size='10' color='#666'>Période : {period}</font>",
            ParagraphStyle('hd', fontSize=15, fontName='Helvetica-Bold', textColor=PURPLE, leading=18))
        ht = Table([[left_cell or '', title_cell]], colWidths=[28*mm, 162*mm])
        ht.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ]))
        return ht
    
    # Calculs stats
    all_stats = []
    for emp in emps:
        emp_rate = employee_rates.get(emp['name'], default_hourly_rate)
        stats = calc_pharma_employee_stats(emp, types_by_key, feries_set,
                                            default_hourly_rate=default_hourly_rate,
                                            employee_rate=emp_rate,
                                            custom_classifications=custom_classifications,
                                            hp=hp)
        all_stats.append(stats)
    
    # === UNE PAGE PAR PHARMACIEN ===
    total_emps = len(emps)
    for idx, (emp, stats) in enumerate(zip(emps, all_stats)):
        if idx > 0:
            story.append(PageBreak())
        
        story.append(make_header(emp['name']))
        story.append(Spacer(1, 3*mm))
        
        nb_jours = stats['nb_total_jours']
        # Choisir taille de police selon densité
        # v69 : police minimum 7pt même en mode dense (lisibilité prioritaire)
        is_dense = nb_jours > 22
        font_size = 7 if is_dense else 8
        S2 = build_styles(font_size)
        
        # === Bloc identification ===
        rate = stats['rate']
        rate_str = f"{rate:,.0f} F/h" if rate > 0 else '—'
        info_data = [[
            Paragraph(f"<b>Pharmacien :</b> {emp['name']}", s_emp),
            Paragraph(f"<b>ID :</b> {emp.get('id','—')}<br/><b>Service :</b> {emp.get('service','—')[:40]}",
                     ParagraphStyle('si', fontSize=9, leading=12)),
            Paragraph(f"<b>Taux horaire :</b> {rate_str}",
                     ParagraphStyle('si2', fontSize=10, fontName='Helvetica-Bold', textColor=NORMAL_COLOR, alignment=TA_RIGHT, leading=13)),
        ]]
        info_t = Table(info_data, colWidths=[70*mm, 75*mm, 45*mm])
        info_t.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOX', (0,0), (-1,-1), 0.6, PURPLE),
            ('BACKGROUND', (0,0), (-1,-1), HexColor('#faf4fc')),
            ('TOPPADDING', (0,0), (-1,-1), 5),('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),('RIGHTPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(info_t)
        story.append(Spacer(1, 3*mm))
        
        # === BANDEAU KPIs ===
        kpis = [
            [
                Paragraph(f"<b>{stats['nb_total_jours']}</b>", s_kpi_n),
                Paragraph(f"<b>{stats['days_present']}</b>",
                         ParagraphStyle('k2', fontSize=18, fontName='Helvetica-Bold', textColor=NORMAL_COLOR, alignment=TA_CENTER, leading=20)),
                Paragraph(f"<b>{stats['days_absent']}</b>",
                         ParagraphStyle('k3', fontSize=18, fontName='Helvetica-Bold', textColor=FERIE_COLOR, alignment=TA_CENTER, leading=20)),
                Paragraph(f"<b>{stats['days_late']}</b>",
                         ParagraphStyle('k4', fontSize=18, fontName='Helvetica-Bold', textColor=WE_COLOR, alignment=TA_CENTER, leading=20)),
                Paragraph(f"<b>{stats['total_min_worked']//60}h{stats['total_min_worked']%60:02d}</b>",
                         ParagraphStyle('k5', fontSize=16, fontName='Helvetica-Bold', textColor=NORMAL_COLOR, alignment=TA_CENTER, leading=20)),
                Paragraph(f"<b>{stats['taux_presence']}%</b>",
                         ParagraphStyle('k6', fontSize=18, fontName='Helvetica-Bold', textColor=PURPLE, alignment=TA_CENTER, leading=20)),
                Paragraph(f"<b>{stats['total_cost']:,.0f} F</b>",
                         ParagraphStyle('k7', fontSize=13, fontName='Helvetica-Bold', textColor=PURPLE, alignment=TA_CENTER, leading=18)),
            ],
            [
                Paragraph("Jours<br/>enregistrés", s_kpi_l),
                Paragraph("Présent(e)", s_kpi_l),
                Paragraph("Absent(e)", s_kpi_l),
                Paragraph("En retard", s_kpi_l),
                Paragraph("Heures<br/>travaillées", s_kpi_l),
                Paragraph("Taux<br/>présence", s_kpi_l),
                Paragraph("Paie<br/>estimée", s_kpi_l),
            ]
        ]
        kpi_t = Table(kpis, colWidths=[27*mm]*7)
        kpi_t.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.8, PURPLE),
            ('INNERGRID', (0,0), (-1,-1), 0.3, HexColor('#ccc')),
            ('BACKGROUND', (0,0), (-1,-1), HexColor('#f3e5f5')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(kpi_t)
        story.append(Spacer(1, 4*mm))
        
        # === RÉPARTITION PAR TYPE ===
        if stats['by_type']:
            story.append(Paragraph("<b>📊 Répartition par type de service</b>", s_subtitle))
            story.append(Spacer(1, 1*mm))
            rep_data = [[
                Paragraph('<b>Type</b>', S2['h']),
                Paragraph('<b>Jours</b>', S2['h']),
                Paragraph('<b>Heures travaillées</b>', S2['h']),
                Paragraph('<b>Majoration</b>', S2['h']),
                Paragraph('<b>Coût</b>', S2['h']),
            ]]
            for lib, info in sorted(stats['by_type'].items(), key=lambda x: -x[1]['duree_min']):
                maj = info.get('taux_maj', 0)
                rep_data.append([
                    Paragraph(f'<font color="{info["couleur"]}"><b>● {lib}</b></font>',
                             ParagraphStyle('rl', fontSize=font_size+1, leading=font_size+3)),
                    Paragraph(str(info['nb']), S2['c']),
                    Paragraph(f"{info['duree_min']//60}h{info['duree_min']%60:02d}", S2['cb']),
                    Paragraph(f"+{int(maj*100)}%" if maj > 0 else '—', S2['c']),
                    Paragraph(f"{info['cout']:,.0f} F", S2['cb']),
                ])
            rep_t = Table(rep_data, colWidths=[60*mm, 25*mm, 35*mm, 28*mm, 42*mm])
            rep_t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), PURPLE),
                ('GRID', (0,0), (-1,-1), 0.3, HexColor('#ccc')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING', (0,0), (-1,-1), 4),('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(rep_t)
            story.append(Spacer(1, 4*mm))
        
        # === DÉTAIL DES JOURS ===
        story.append(Paragraph("<b>📅 Détail des présences</b>", s_subtitle))
        story.append(Spacer(1, 1*mm))
        
        # 11 colonnes : N° | Date | J | Type | H.oblig | Arrivée | Sortie | Durée | Retard | Statut | Coût
        det_hdrs = ['N°','Date','J','Type','H.oblig.','Arrivée','Sortie','Durée','Retard','Statut','Coût']
        det_cw = [8*mm, 18*mm, 10*mm, 28*mm, 19*mm, 14*mm, 14*mm, 14*mm, 14*mm, 14*mm, 27*mm]
        
        det_data = [[Paragraph(x, S2['h']) for x in det_hdrs]]
        
        for i, r in enumerate(stats['enriched'], 1):
            # Jour avec couleur weekend
            jour_str = r['jour']
            if r['jour'] in ('Sam','Dim'):
                jour_str = f'<font color="#f29f2f"><b>{r["jour"]}</b></font>'
            if r['type_code'] == 'ferie':
                jour_str = f'<font color="#c53030"><b>{r["jour"]}*</b></font>'
            
            # Type coloré
            type_str = f'<font color="{r["type_couleur"]}"><b>● {r["type_libelle"]}</b></font>'
            
            # Heure oblig (start - end)
            hoblig_str = f"{r['sched_start']}<br/>{r['sched_end']}" if r['sched_start'] and r['sched_end'] else '—'
            
            # Arrivée / Sortie
            arr_str = r['arrival'] or '—'
            dep_str = r['departure'] or '—'
            
            # Durée
            dur_str = r['worked_str'] if r['worked_min'] > 0 else '—'
            
            # Retard
            late_str = r['late_str'] if r['late_min'] > 0 else '—'
            
            # Statut
            if r['state'] == 'OK':
                statut = Paragraph('OK', S2['g'])
            elif r['state'] == 'RETARD':
                statut = Paragraph('RETARD', S2['o'])
            elif r['state'] == 'ABS':
                statut = Paragraph('ABSENT', S2['r'])
            elif r['state'] == 'ARR':
                statut = Paragraph('NON<br/>SORTI', S2['o'])
            else:
                statut = Paragraph(r['state'], S2['c'])
            
            det_data.append([
                Paragraph(str(i), S2['c']),
                Paragraph(r['date'], S2['c']),
                Paragraph(jour_str, S2['c']),
                Paragraph(type_str, ParagraphStyle('tc', fontSize=font_size, alignment=TA_LEFT, leading=font_size+2)),
                Paragraph(hoblig_str, S2['c']),
                Paragraph(arr_str, S2['c']),
                Paragraph(dep_str, S2['c']),
                Paragraph(dur_str, S2['cb']),
                Paragraph(late_str, S2['o'] if r['late_min'] > 0 else S2['c']),
                statut,
                Paragraph(f"{r['cout_jour']:,.0f} F" if r['cout_jour'] > 0 else '—', S2['cb']),
            ])
        
        tpad = 2 if is_dense else 3
        det_t = Table(det_data, colWidths=det_cw, repeatRows=1)
        det_t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PURPLE),
            ('GRID', (0,0), (-1,-1), 0.3, HexColor('#ccc')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING', (0,0), (-1,-1), tpad),
            ('BOTTOMPADDING', (0,0), (-1,-1), tpad),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ]))
        # Alternance lignes
        for i in range(2, len(det_data), 2):
            det_t.setStyle(TableStyle([('BACKGROUND', (0,i), (-1,i), HexColor('#fafafa'))]))
        story.append(det_t)
        story.append(Spacer(1, 3*mm))
        
        # === TOTAUX ===
        total_data = [[
            Paragraph(f"<b>HEURES TRAVAILLÉES : {stats['total_min_worked']//60}h{stats['total_min_worked']%60:02d}</b>",
                     ParagraphStyle('t1', fontSize=10, fontName='Helvetica-Bold', textColor=NORMAL_COLOR, alignment=TA_CENTER, leading=12)),
            Paragraph(f"<b>RETARDS : {stats['total_late_min']//60}h{stats['total_late_min']%60:02d}</b>",
                     ParagraphStyle('t2', fontSize=10, fontName='Helvetica-Bold', textColor=WE_COLOR, alignment=TA_CENTER, leading=12)),
            Paragraph(f"<b>PRIMES : {stats['total_primes']:,.0f} F</b>",
                     ParagraphStyle('t3', fontSize=10, fontName='Helvetica-Bold', textColor=NIGHT, alignment=TA_CENTER, leading=12)),
            Paragraph(f"<b>PAIE ESTIMÉE : {stats['total_cost']:,.0f} F</b>",
                     ParagraphStyle('t4', fontSize=11, fontName='Helvetica-Bold', textColor=PURPLE, alignment=TA_CENTER, leading=13)),
        ]]
        total_t = Table(total_data, colWidths=[48*mm, 40*mm, 42*mm, 60*mm])
        total_t.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 1.2, PURPLE),
            ('INNERGRID', (0,0), (-1,-1), 0.4, PURPLE),
            ('BACKGROUND', (0,0), (-1,-1), HexColor('#fff3e0')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(total_t)
        story.append(Spacer(1, 2*mm))
        
        # Légende
        story.append(Paragraph(
            "<font size='7' color='#888'><i>* = Jour férié | RETARD = arrivée &gt; 5 min après l'heure obligatoire | "
            "NON SORTI = badge de sortie non enregistré | Coût = heures × taux × (1+majoration) + prime fixe</i></font>",
            ParagraphStyle('lg', fontSize=7, textColor=HexColor('#888'))))
        story.append(Spacer(1, 1*mm))
        story.append(Paragraph(
            f"<font size='8' color='#888'>Pharmacien {idx+1}/{total_emps}</font>",
            ParagraphStyle('pg', fontSize=8, textColor=HexColor('#888'), alignment=TA_RIGHT)))
    
    # === PAGE RÉCAPITULATIF GLOBAL ===
    if len(emps) > 1:
        story.append(PageBreak())
        story.append(make_header())
        story.append(Spacer(1, 5*mm))
        story.append(Paragraph("📊 RÉCAPITULATIF GLOBAL",
                              ParagraphStyle('rg', fontSize=18, fontName='Helvetica-Bold',
                                            textColor=PURPLE, alignment=TA_CENTER, leading=22)))
        story.append(Spacer(1, 5*mm))
        
        S3 = build_styles(8)
        
        # Tableau récap par employé
        recap_hdrs = ['Pharmacien','Jours','Prés.','Abs.','Retards','H. travail.','Taux','Primes','Paie estimée']
        recap_data = [[Paragraph(f"<b>{x}</b>", S3['h']) for x in recap_hdrs]]
        
        tot_jours = tot_pres = tot_abs = tot_late = 0
        tot_min = tot_primes = tot_cost = 0
        
        for emp, st in zip(emps, all_stats):
            recap_data.append([
                Paragraph(emp['name'][:30], ParagraphStyle('rn', fontSize=9, fontName='Helvetica-Bold', leading=11)),
                Paragraph(str(st['nb_total_jours']), S3['c']),
                Paragraph(str(st['days_present']), S3['g']),
                Paragraph(str(st['days_absent']), S3['r']),
                Paragraph(str(st['days_late']), S3['o']),
                Paragraph(f"{st['total_min_worked']//60}h{st['total_min_worked']%60:02d}", S3['cb']),
                Paragraph(f"{st['taux_presence']}%", S3['cb']),
                Paragraph(f"{st['total_primes']:,.0f} F", S3['c']),
                Paragraph(f"{st['total_cost']:,.0f} F", S3['cb']),
            ])
            tot_jours += st['nb_total_jours']
            tot_pres += st['days_present']
            tot_abs += st['days_absent']
            tot_late += st['days_late']
            tot_min += st['total_min_worked']
            tot_primes += st['total_primes']
            tot_cost += st['total_cost']
        
        # Ligne TOTAL
        recap_data.append([
            Paragraph("<b>TOTAL</b>", ParagraphStyle('rt', fontSize=10, fontName='Helvetica-Bold', textColor=PURPLE)),
            Paragraph(f"<b>{tot_jours}</b>", S3['cb']),
            Paragraph(f"<b>{tot_pres}</b>", S3['cb']),
            Paragraph(f"<b>{tot_abs}</b>", S3['cb']),
            Paragraph(f"<b>{tot_late}</b>", S3['cb']),
            Paragraph(f"<b>{tot_min//60}h{tot_min%60:02d}</b>", S3['cb']),
            Paragraph(f"<b>{round(100*tot_pres/tot_jours,1) if tot_jours else 0}%</b>", S3['cb']),
            Paragraph(f"<b>{tot_primes:,.0f} F</b>", S3['cb']),
            Paragraph(f"<b>{tot_cost:,.0f} F</b>",
                     ParagraphStyle('rtc', fontSize=10, fontName='Helvetica-Bold', textColor=PURPLE, alignment=TA_CENTER)),
        ])
        
        recap_cw = [42*mm, 14*mm, 14*mm, 14*mm, 18*mm, 22*mm, 16*mm, 22*mm, 28*mm]
        recap_t = Table(recap_data, colWidths=recap_cw)
        recap_t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PURPLE),
            ('BACKGROUND', (0,-1), (-1,-1), HexColor('#fff3e0')),
            ('GRID', (0,0), (-1,-1), 0.4, HexColor('#ccc')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('BOX', (0,-1), (-1,-1), 1.5, PURPLE),
        ]))
        story.append(recap_t)
    
    doc.build(story)
    return output_path


def generate_pharma_recap_page(pdf_path, emps, pharmacy_name='PHARMACIE', period='',
                                 types_by_key=None, feries_set=None,
                                 default_hourly_rate=0, employee_rates=None,
                                 emps_classifications=None,
                                 logo_path=None):
    """Ajoute une (ou plusieurs) page(s) Pharmacie spécifiques à un PDF existant.
    
    Pages ajoutées :
      1. Récap global Pharmacie (paie estimée avec majorations par type)
      2. Détail par pharmacien (heures par type + coût total)
    
    Utilise pypdf pour fusionner le PDF existant avec les nouvelles pages.
    """
    if not os.path.exists(pdf_path):
        return False
    if types_by_key is None: types_by_key = {}
    if feries_set is None: feries_set = set()
    if employee_rates is None: employee_rates = {}
    if emps_classifications is None: emps_classifications = {}
    
    import tempfile, io
    from pypdf import PdfWriter, PdfReader
    
    # Générer un PDF temporaire avec les pages Pharmacie
    tmp_pharma = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    tmp_pharma.close()
    
    PURPLE = HexColor('#7b1fa2')
    NIGHT_BLUE = HexColor('#3949ab')
    ORANGE = HexColor('#f29f2f')
    RED_C = HexColor('#c53030')
    TEAL_C = HexColor('#1A7A6D')
    
    doc = SimpleDocTemplate(tmp_pharma.name, pagesize=A4,
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=10*mm, bottomMargin=10*mm)
    story = []
    S = make_styles()
    
    # === Header ===
    def make_header():
        cells = []
        left = ''
        if logo_path and os.path.exists(logo_path):
            try:
                from reportlab.platypus import Image as _Img
                left = _Img(logo_path, width=22*mm, height=22*mm)
            except: left = ''
        title = Paragraph(
            f"<b>{pharmacy_name}</b><br/>"
            f"<font size='11' color='#666'>💊 Récapitulatif Pharmacie — paie estimée</font><br/>"
            f"<font size='10' color='#666'>Période : {period}</font>",
            ParagraphStyle('phd', fontSize=15, fontName='Helvetica-Bold', textColor=PURPLE, leading=18))
        ht = Table([[left or '', title]], colWidths=[28*mm, 162*mm])
        ht.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ]))
        return ht
    
    story.append(make_header())
    story.append(Spacer(1, 5*mm))
    
    # === Titre section ===
    story.append(Paragraph("💰 RÉCAPITULATIF PHARMACIE — Paie estimée avec majorations",
                          ParagraphStyle('rgh', fontSize=16, fontName='Helvetica-Bold',
                                        textColor=PURPLE, alignment=TA_CENTER, leading=20)))
    story.append(Spacer(1, 3*mm))
    
    # === Légende des types de service ===
    legende_data = [[
        Paragraph("<b>📋 TYPES DE SERVICE & MAJORATIONS</b>",
                 ParagraphStyle('lh', fontSize=10, fontName='Helvetica-Bold', textColor=white, alignment=TA_CENTER, leading=12))
    ]]
    legende_t = Table(legende_data, colWidths=[190*mm])
    legende_t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), PURPLE),
        ('TOPPADDING', (0,0), (-1,-1), 3),('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(legende_t)
    
    types_for_legend = []
    seen_codes = set()
    for k, t in types_by_key.items():
        code = (t.get('code') or '').lower()
        if code in seen_codes: continue
        seen_codes.add(code)
        types_for_legend.append(t)
    
    types_for_legend.sort(key=lambda t: t.get('ordre_affichage', 99))
    
    types_data = [[Paragraph(f"<b>{x}</b>", S['h']) for x in
                  ['Code', 'Libellé', 'Couleur', 'Majoration', 'Prime fixe', 'Description']]]
    for t in types_for_legend:
        code = t.get('code', '')
        lib = t.get('libelle', '')
        couleur = t.get('couleur', '#1A7A6D')
        maj = float(t.get('taux_majoration', 0) or 0)
        prime = float(t.get('prime_fixe', 0) or 0)
        desc = t.get('description', '')
        types_data.append([
            Paragraph(f'<code>{code}</code>', ParagraphStyle('cc', fontSize=8, leading=10)),
            Paragraph(f'<font color="{couleur}"><b>{lib}</b></font>', ParagraphStyle('lc', fontSize=9, leading=11)),
            Paragraph(f'<font color="{couleur}">●</font> <font size="7">{couleur}</font>', ParagraphStyle('cs', fontSize=9, leading=11)),
            Paragraph(f'+{int(maj*100)}%' if maj > 0 else '—', S['cb']),
            Paragraph(f'{prime:,.0f} F' if prime > 0 else '—', S['c']),
            Paragraph(desc, ParagraphStyle('dc', fontSize=8, leading=10)),
        ])
    types_table = Table(types_data, colWidths=[20*mm, 35*mm, 30*mm, 22*mm, 25*mm, 58*mm])
    types_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PURPLE),
        ('GRID', (0,0), (-1,-1), 0.3, HexColor('#ccc')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(types_table)
    story.append(Spacer(1, 5*mm))
    
    # === Calcul pour chaque pharmacien ===
    all_pharma_stats = []
    global_by_type = {}
    
    for emp in emps:
        emp_rate = employee_rates.get(emp['name'], default_hourly_rate)
        emp_classif = emps_classifications.get(emp['name'], {})
        
        # Stats pharmacie : pour chaque record on calcule type + coût avec majoration
        nb_by_type = {}  # libelle -> {nb_jours, total_min, cout, couleur, taux_maj, prime}
        total_min = 0
        total_cost = 0
        total_primes = 0
        
        for rec in emp['records']:
            # Classifier
            forced = emp_classif.get(rec['date'])
            if forced:
                type_code = forced
            else:
                type_code = _classify_pharma_day(rec['date'], rec.get('sched_start',''), feries_set)
            
            type_data = types_by_key.get(type_code) or {}
            type_libelle = type_data.get('libelle', type_code.replace('_',' ').title())
            type_couleur = type_data.get('couleur', '#1A7A6D')
            taux_maj = float(type_data.get('taux_majoration', 0) or 0)
            prime = float(type_data.get('prime_fixe', 0) or 0)
            
            # Calculer durée travaillée
            worked_min = 0
            dur_str = rec.get('duration', '')
            if dur_str and dur_str != '-' and ':' in dur_str:
                try:
                    h, m = dur_str.split(':')
                    worked_min = int(h) * 60 + int(m)
                except: pass
            
            if worked_min == 0 and rec.get('arrival') and rec.get('departure'):
                try:
                    ah, am = map(int, rec['arrival'].split(':'))
                    dh, dm = map(int, rec['departure'].split(':'))
                    w = (dh*60 + dm) - (ah*60 + am)
                    if w < 0: w += 24*60
                    worked_min = w
                except: pass
            
            is_present = bool(rec.get('arrival') and rec.get('arrival') != '-')
            
            cout_jour = (worked_min / 60.0) * emp_rate * (1 + taux_maj) + (prime if is_present else 0)
            
            if type_libelle not in nb_by_type:
                nb_by_type[type_libelle] = {
                    'nb_jours': 0, 'nb_presents': 0, 'total_min': 0, 'cout': 0,
                    'primes': 0, 'couleur': type_couleur, 'taux_maj': taux_maj, 'prime_fixe': prime,
                    'code': type_code,
                }
            nb_by_type[type_libelle]['nb_jours'] += 1
            if is_present:
                nb_by_type[type_libelle]['nb_presents'] += 1
                nb_by_type[type_libelle]['primes'] += prime
            nb_by_type[type_libelle]['total_min'] += worked_min
            nb_by_type[type_libelle]['cout'] += cout_jour
            
            total_min += worked_min
            total_cost += cout_jour
            if is_present: total_primes += prime
            
            # Global
            if type_libelle not in global_by_type:
                global_by_type[type_libelle] = {
                    'nb_jours': 0, 'total_min': 0, 'cout': 0,
                    'couleur': type_couleur, 'taux_maj': taux_maj,
                }
            global_by_type[type_libelle]['nb_jours'] += 1
            global_by_type[type_libelle]['total_min'] += worked_min
            global_by_type[type_libelle]['cout'] += cout_jour
        
        all_pharma_stats.append({
            'name': emp['name'], 'id': emp.get('id',''), 'rate': emp_rate,
            'by_type': nb_by_type, 'total_min': total_min,
            'total_primes': total_primes, 'total_cost': total_cost,
        })
    
    # === Tableau global par pharmacien ===
    story.append(Paragraph("<b>📊 Récapitulatif par pharmacien</b>",
                          ParagraphStyle('rph', fontSize=11, fontName='Helvetica-Bold',
                                        textColor=PURPLE, leading=14)))
    story.append(Spacer(1, 2*mm))
    
    recap_hdrs = ['Pharmacien', 'ID', 'Taux/h', 'Heures totales', 'Primes', 'Paie estimée']
    recap_data = [[Paragraph(f"<b>{x}</b>", S['h']) for x in recap_hdrs]]
    
    total_emp_min = 0
    total_emp_primes = 0
    total_emp_cost = 0
    
    for st in all_pharma_stats:
        recap_data.append([
            Paragraph(st['name'][:35], ParagraphStyle('en', fontSize=9, fontName='Helvetica-Bold', leading=11)),
            Paragraph(st['id'] or '—', ParagraphStyle('id', fontSize=8, leading=10)),
            Paragraph(f"{st['rate']:,.0f} F" if st['rate'] else '—', S['c']),
            Paragraph(f"{st['total_min']//60}h{st['total_min']%60:02d}", S['cb']),
            Paragraph(f"{st['total_primes']:,.0f} F" if st['total_primes'] else '—', S['c']),
            Paragraph(f"<b>{st['total_cost']:,.0f} F</b>",
                     ParagraphStyle('tc', fontSize=10, fontName='Helvetica-Bold', textColor=PURPLE, alignment=TA_CENTER)),
        ])
        total_emp_min += st['total_min']
        total_emp_primes += st['total_primes']
        total_emp_cost += st['total_cost']
    
    # Ligne TOTAL
    recap_data.append([
        Paragraph("<b>TOTAL</b>", ParagraphStyle('rt', fontSize=10, fontName='Helvetica-Bold', textColor=PURPLE)),
        Paragraph(f"<b>{len(all_pharma_stats)} pharm.</b>", ParagraphStyle('rt2', fontSize=9, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph('', S['c']),
        Paragraph(f"<b>{total_emp_min//60}h{total_emp_min%60:02d}</b>", S['cb']),
        Paragraph(f"<b>{total_emp_primes:,.0f} F</b>", S['cb']),
        Paragraph(f"<b>{total_emp_cost:,.0f} F</b>",
                 ParagraphStyle('rtc', fontSize=11, fontName='Helvetica-Bold', textColor=PURPLE, alignment=TA_CENTER)),
    ])
    
    recap_cw = [50*mm, 22*mm, 22*mm, 28*mm, 28*mm, 40*mm]
    recap_t = Table(recap_data, colWidths=recap_cw)
    recap_t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PURPLE),
        ('BACKGROUND', (0,-1), (-1,-1), HexColor('#fff3e0')),
        ('GRID', (0,0), (-1,-1), 0.4, HexColor('#ccc')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 5),('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('BOX', (0,-1), (-1,-1), 1.5, PURPLE),
    ]))
    story.append(recap_t)
    story.append(Spacer(1, 5*mm))
    
    # === Répartition globale par type ===
    if global_by_type:
        story.append(Paragraph("<b>📊 Répartition globale par type de service</b>",
                              ParagraphStyle('gh', fontSize=11, fontName='Helvetica-Bold', textColor=PURPLE, leading=14)))
        story.append(Spacer(1, 2*mm))
        gbl_hdrs = ['Type de service', 'Jours', 'Heures travaillées', 'Majoration', 'Coût total']
        gbl_data = [[Paragraph(f"<b>{x}</b>", S['h']) for x in gbl_hdrs]]
        for lib, info in sorted(global_by_type.items(), key=lambda x: -x[1]['total_min']):
            gbl_data.append([
                Paragraph(f'<font color="{info["couleur"]}"><b>● {lib}</b></font>',
                         ParagraphStyle('gl1', fontSize=10, leading=12)),
                Paragraph(str(info['nb_jours']), S['c']),
                Paragraph(f"{info['total_min']//60}h{info['total_min']%60:02d}", S['cb']),
                Paragraph(f"+{int(info['taux_maj']*100)}%" if info['taux_maj'] > 0 else '—', S['c']),
                Paragraph(f"{info['cout']:,.0f} F", S['cb']),
            ])
        gbl_t = Table(gbl_data, colWidths=[60*mm, 25*mm, 40*mm, 28*mm, 37*mm])
        gbl_t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PURPLE),
            ('GRID', (0,0), (-1,-1), 0.3, HexColor('#ccc')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(gbl_t)
        story.append(Spacer(1, 4*mm))
    
    # === Détail par pharmacien : tableau heures par type ===
    if len(all_pharma_stats) <= 30:  # éviter PDF interminable
        story.append(PageBreak())
        story.append(make_header())
        story.append(Spacer(1, 5*mm))
        story.append(Paragraph("📋 DÉTAIL PAR PHARMACIEN — Heures et coût par type de service",
                              ParagraphStyle('dgh', fontSize=14, fontName='Helvetica-Bold',
                                            textColor=PURPLE, alignment=TA_CENTER, leading=18)))
        story.append(Spacer(1, 4*mm))
        
        for st in all_pharma_stats:
            if not st['by_type']: continue
            story.append(Paragraph(f"<b>👤 {st['name']}</b> <font size='9' color='#666'>— ID: {st['id'] or '—'} — Taux : {st['rate']:,.0f} F/h</font>",
                                  ParagraphStyle('ed', fontSize=11, fontName='Helvetica-Bold', textColor=PURPLE, leading=14)))
            story.append(Spacer(1, 1*mm))
            
            byt_hdrs = ['Type', 'Jours', 'Présents', 'Heures', 'Majoration', 'Primes', 'Coût']
            byt_data = [[Paragraph(f"<b>{x}</b>", S['h']) for x in byt_hdrs]]
            for lib, info in sorted(st['by_type'].items(), key=lambda x: -x[1]['total_min']):
                byt_data.append([
                    Paragraph(f'<font color="{info["couleur"]}"><b>● {lib}</b></font>',
                             ParagraphStyle('bl1', fontSize=9, leading=11)),
                    Paragraph(str(info['nb_jours']), S['c']),
                    Paragraph(str(info['nb_presents']), S['c']),
                    Paragraph(f"{info['total_min']//60}h{info['total_min']%60:02d}", S['cb']),
                    Paragraph(f"+{int(info['taux_maj']*100)}%" if info['taux_maj'] > 0 else '—', S['c']),
                    Paragraph(f"{info['primes']:,.0f} F" if info['primes'] else '—', S['c']),
                    Paragraph(f"{info['cout']:,.0f} F", S['cb']),
                ])
            # Total
            byt_data.append([
                Paragraph("<b>TOTAL</b>", ParagraphStyle('bt', fontSize=9, fontName='Helvetica-Bold', textColor=PURPLE)),
                Paragraph(f"<b>{sum(i['nb_jours'] for i in st['by_type'].values())}</b>", S['cb']),
                Paragraph(f"<b>{sum(i['nb_presents'] for i in st['by_type'].values())}</b>", S['cb']),
                Paragraph(f"<b>{st['total_min']//60}h{st['total_min']%60:02d}</b>", S['cb']),
                Paragraph('', S['c']),
                Paragraph(f"<b>{st['total_primes']:,.0f} F</b>", S['cb']),
                Paragraph(f"<b>{st['total_cost']:,.0f} F</b>",
                         ParagraphStyle('btc', fontSize=10, fontName='Helvetica-Bold', textColor=PURPLE, alignment=TA_CENTER)),
            ])
            byt_t = Table(byt_data, colWidths=[55*mm, 18*mm, 22*mm, 22*mm, 25*mm, 22*mm, 28*mm])
            byt_t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), PURPLE),
                ('BACKGROUND', (0,-1), (-1,-1), HexColor('#fff3e0')),
                ('GRID', (0,0), (-1,-1), 0.3, HexColor('#ccc')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING', (0,0), (-1,-1), 3),('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('BOX', (0,-1), (-1,-1), 1, PURPLE),
            ]))
            story.append(byt_t)
            story.append(Spacer(1, 3*mm))
    
    doc.build(story)
    
    # === Fusionner avec le PDF existant ===
    writer = PdfWriter()
    # Ajouter les pages du PDF existant
    base_reader = PdfReader(pdf_path)
    for page in base_reader.pages:
        writer.add_page(page)
    # Ajouter les pages Pharmacie
    pharma_reader = PdfReader(tmp_pharma.name)
    for page in pharma_reader.pages:
        writer.add_page(page)
    
    # Écraser le fichier d'origine
    with open(pdf_path, 'wb') as fp:
        writer.write(fp)
    
    # Nettoyer
    try: os.remove(tmp_pharma.name)
    except: pass
    
    return True
